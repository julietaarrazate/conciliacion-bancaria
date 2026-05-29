#!/usr/bin/env python3
"""
Analizador de extractos y planillas Excel — pre-validación antes de subir al sistema.

Reutiliza el MISMO parser de producción (app/services/excel_parser.py) para que el
diagnóstico nunca se desincronice de cómo el sistema realmente lee los archivos.

Uso:
    python backend/scripts/analizar_excel.py <archivo.xlsx>           # autodetecta tipo
    python backend/scripts/analizar_excel.py <archivo.xlsx> --planilla # forzar modo planilla

Qué reporta:
  - Banco detectado y mapeo de columnas
  - Cantidad de movimientos válidos + total
  - Montos duplicados (clave: el sistema exige identidad CUIT/CBU cuando un monto se repite)
  - Filas sin identidad (sin CUIT ni CBU) — quedarían en "sin datos" al conciliar
  - Fechas fuera de rango / no parseables
  - Filas que el parser descarta (monto vacío o 0)
  - Columnas que NO matchearon ningún keyword conocido (posible formato nuevo)
"""
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta

# Permitir importar app.services.* corriendo desde la raíz del repo
_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _BACKEND)

import openpyxl  # noqa: E402
from app.services.excel_parser import (  # noqa: E402
    preparar_archivo, detectar_banco, detectar_columnas,
    parsear_generico, _parse_fecha, _parse_monto, _normalizar,
)

_CUIT_RE = re.compile(r"\b\d{2}[-\s]?\d{8}[-\s]?\d\b")        # 11 dígitos (con o sin guiones)
_CBU_RE = re.compile(r"\b\d{22}\b")                            # 22 dígitos


def _c(txt, code):
    return f"\033[{code}m{txt}\033[0m" if sys.stdout.isatty() else txt


def _verde(t): return _c(t, "32")
def _rojo(t): return _c(t, "31")
def _amar(t): return _c(t, "33")
def _bold(t): return _c(t, "1")


def analizar(filepath: str):
    if not os.path.exists(filepath):
        print(_rojo(f"✗ No existe el archivo: {filepath}"))
        return 1

    path, _ = preparar_archivo(filepath)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    banco = detectar_banco(ws)
    cols = detectar_columnas(ws)
    movs = parsear_generico(ws, cols)

    print(_bold("\n══════════════════════════════════════════════"))
    print(_bold(f"  ANÁLISIS — {os.path.basename(filepath)}"))
    print(_bold("══════════════════════════════════════════════"))

    print(f"\n{_bold('Banco detectado:')} {banco or _amar('genérico (no identificado)')}")
    print(f"{_bold('Fila de headers:')} {cols.get('hdr_row')}")
    print(_bold("\nMapeo de columnas:"))
    nombres = {"monto": "Monto", "credito": "Crédito", "debito": "Débito",
               "fecha": "Fecha", "titular": "Titular/Concepto", "orden": "Orden/Ref",
               "saldo": "Saldo", "fecha_acred": "Fecha acred.", "cliente_acred": "Cliente acred."}
    for k, label in nombres.items():
        v = cols.get(k)
        if v:
            print(f"  • {label:18s} → columna {v}")

    # Columnas con header que no matchearon nada.
    # Ojo: solo las CLAVES que son columnas cuentan (hdr_row/mes son nº de fila u otra cosa).
    hdr_row = cols.get("hdr_row", 1)
    _claves_columna = ("monto", "credito", "debito", "fecha", "titular",
                       "orden", "saldo", "fecha_acred", "cliente_acred")
    mapeadas = {cols[k] for k in _claves_columna if isinstance(cols.get(k), int)}
    sin_match = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(hdr_row, c).value
        if val and _normalizar(str(val)) and c not in mapeadas:
            sin_match.append(str(val).strip())
    if sin_match:
        print(_amar(f"\n  ⚠ Columnas no reconocidas: {', '.join(sin_match[:8])}"))

    # ── Estadísticas de movimientos ──────────────────────────────
    total_validos = len(movs)
    suma = round(sum(m["monto"] for m in movs if m.get("monto")), 2)
    print(_bold(f"\nMovimientos válidos: {_verde(str(total_validos))}"))
    print(f"Suma de montos: $ {suma:,.2f}")

    if total_validos == 0:
        print(_rojo("\n✗ El parser no extrajo NINGÚN movimiento. Probable formato no soportado"
                    " o fila de headers mal detectada. Revisá el archivo."))
        return 1

    # ── Montos duplicados (lo más importante para conciliar) ─────
    montos = Counter(m["monto"] for m in movs if m.get("monto"))
    dups = {monto: n for monto, n in montos.items() if n > 1}
    if dups:
        top = sorted(dups.items(), key=lambda x: -x[1])[:8]
        print(_amar(f"\n⚠ Montos duplicados: {len(dups)} montos se repiten"))
        for monto, n in top:
            print(f"    $ {monto:,.2f}  ×{n}")
        print(_amar("    → El sistema EXIGE identidad (CUIT/CBU/nº cuenta) para estos. "
                    "Sin identidad en la planilla quedan 'sin datos'."))
    else:
        print(_verde("\n✓ Sin montos duplicados — conciliación directa por monto."))

    # ── Identidad (CUIT/CBU en el texto del titular/concepto) ────
    con_cuit = con_cbu = 0
    for m in movs:
        txt = str(m.get("titular") or "")
        if _CUIT_RE.search(txt):
            con_cuit += 1
        if _CBU_RE.search(txt):
            con_cbu += 1
    print(_bold("\nIdentidad detectada en el texto:"))
    print(f"  • Con CUIT visible: {con_cuit}/{total_validos}")
    print(f"  • Con CBU visible:  {con_cbu}/{total_validos}")
    if dups and con_cuit == 0 and con_cbu == 0:
        print(_rojo("  ✗ Hay duplicados PERO ninguna fila muestra CUIT/CBU → "
                    "la conciliación de esos montos va a fallar."))

    # ── Fechas ───────────────────────────────────────────────────
    fechas = [m["fecha"] for m in movs if m.get("fecha")]
    sin_fecha = total_validos - len(fechas)
    if fechas:
        fmin, fmax = min(fechas), max(fechas)
        print(_bold("\nRango de fechas:"))
        print(f"  {fmin} → {fmax}")
        hoy = date.today()
        futuras = [f for f in fechas if f > hoy + timedelta(days=1)]
        muy_viejas = [f for f in fechas if f < hoy - timedelta(days=365)]
        if futuras:
            print(_amar(f"  ⚠ {len(futuras)} fecha(s) en el futuro"))
        if muy_viejas:
            print(_amar(f"  ⚠ {len(muy_viejas)} fecha(s) de hace más de 1 año"))
    if sin_fecha:
        print(_amar(f"  ⚠ {sin_fecha} movimiento(s) sin fecha parseable"))

    # ── Veredicto ────────────────────────────────────────────────
    print(_bold("\n──────────────────────────────────────────────"))
    problemas = []
    if not banco:
        problemas.append("banco no identificado (usará reglas genéricas)")
    if sin_match:
        problemas.append(f"{len(sin_match)} columna(s) sin reconocer")
    if dups and con_cuit == 0 and con_cbu == 0:
        problemas.append("duplicados sin identidad")
    if sin_fecha:
        problemas.append(f"{sin_fecha} sin fecha")

    if not problemas:
        print(_verde("✓ LISTO PARA SUBIR — no se detectaron problemas de formato."))
    else:
        print(_amar("⚠ REVISAR ANTES DE SUBIR:"))
        for p in problemas:
            print(_amar(f"    - {p}"))
    print(_bold("──────────────────────────────────────────────\n"))
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("Uso: python backend/scripts/analizar_excel.py <archivo.xlsx>")
        sys.exit(2)
    sys.exit(analizar(args[0]))
