"""
Parseo de archivos Excel para extractos bancarios y planillas de clientes.
Soporta .xlsx, .xls y .csv.
Soporta Banco Macro, BBVA, Santander, Galicia, ICBC y formatos genericos.
"""

import re
import os
import tempfile
from typing import Tuple, Optional, List, Dict
from datetime import date, datetime
import openpyxl


def _convertir_xls_a_xlsx(filepath: str) -> str:
    """Convierte .xls (formato viejo) a .xlsx para procesarlo con openpyxl."""
    import xlrd
    from openpyxl import Workbook

    xls = xlrd.open_workbook(filepath)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in xls.sheet_names():
        xls_sheet = xls.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = datetime(*xlrd.xldate_as_tuple(val, xls.datemode))
                    except Exception:
                        pass
                ws.cell(row=row + 1, column=col + 1, value=val)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    wb.close()
    xls.release_resources()
    return tmp.name


def _convertir_csv_a_xlsx(filepath: str) -> str:
    """Convierte .csv a .xlsx para procesarlo con openpyxl."""
    import csv
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    encodings = ['utf-8', 'latin-1', 'cp1252']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()

    delimiter = ';' if content.count(';') > content.count(',') else ','
    # Patrón para montos argentinos: "1.234,56" o "-1.234,56"
    _ARG_MONTO_RE = re.compile(r'^-?\d{1,3}(\.\d{3})*(,\d+)?$')

    def _normalizar_numero(s: str) -> str:
        """Convierte "1.234,56" → "1234.56"; deja otros strings intactos."""
        if _ARG_MONTO_RE.match(s):
            # Quitar punto separador de miles y cambiar coma decimal por punto
            return re.sub(r'\.(?=\d{3}([,\s]|$))', '', s).replace(',', '.')
        return s

    reader = csv.reader(content.splitlines(), delimiter=delimiter)
    for row_idx, row in enumerate(reader, 1):
        for col_idx, val in enumerate(row, 1):
            val = val.strip()
            try:
                ws.cell(row=row_idx, column=col_idx, value=float(_normalizar_numero(val)))
            except ValueError:
                ws.cell(row=row_idx, column=col_idx, value=val)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    wb.close()
    return tmp.name


def preparar_archivo(filepath: str) -> Tuple[str, bool]:
    """
    Si el archivo es .xls o .csv, lo convierte a .xlsx.
    Retorna (ruta_xlsx, necesita_cleanup).
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return _convertir_xls_a_xlsx(filepath), True
    if ext in ('.csv', '.tsv'):
        return _convertir_csv_a_xlsx(filepath), True
    return filepath, False

INDICADORES_BANCO = {
    "macro":       ["macro", "banco macro"],
    "bbva":        ["bbva", "frances", "banco frances"],
    "santander":   ["santander", "rio", "santander rio"],
    "galicia":     ["galicia", "banco galicia"],
    "icbc":        ["icbc", "industrial"],
    "nacion":      ["nacion", "banco de la nacion"],
    "provincia":   ["provincia", "bapro"],
    "ciudad":      ["ciudad", "banco ciudad"],
    "hsbc":        ["hsbc"],
    # ── Bancos agregados (v3.13+): Mercado Pago + 6 bancos regionales ──────
    "mercadopago": ["mercado pago", "mercadopago", "mercado libre",
                    "mercadolibre", "money_release", "dineromp"],
    "credicoop":   ["credicoop", "banco credicoop"],
    "supervielle": ["supervielle", "banco supervielle"],
    "patagonia":   ["patagonia", "banco patagonia"],
    "bancor":      ["bancor", "banco de cordoba",
                    "banco de la provincia de cordoba"],
    "rioja":       ["banco rioja", "banco de la rioja",
                    "nuevo banco de la rioja"],
    "lapampa":     ["banco de la pampa", "bancodelapampa", "banco la pampa"],
}

def detectar_banco(ws) -> str:
    texto = ""
    # Escanea más filas/columnas: algunos formatos (Credicoop, Supervielle,
    # Patagonia) ponen el nombre del banco recién en la fila 3-8.
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, min(12, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v:
                texto += " " + str(v).lower()
    texto = _normalizar(texto)
    # Matchear primero las keywords MÁS LARGAS/específicas para evitar que un
    # token corto y goloso (ej. "rio" de Santander) capture a otro banco
    # (ej. "Banco de La Rioja"). Se ordena (banco, keyword) por longitud desc.
    candidatos = []
    for banco, palabras in INDICADORES_BANCO.items():
        for p in palabras:
            candidatos.append((banco, _normalizar(p)))
    candidatos.sort(key=lambda bp: len(bp[1]), reverse=True)
    for banco, p in candidatos:
        # "rio" suelto no debe matchear "rioja" / "prioritario" / etc.
        if p == "rio" and ("rioja" in texto or "prio" in texto):
            continue
        if p in texto:
            return banco
    # Mercado Pago por estructura: el CSV de liberaciones no siempre
    # contiene el texto "Mercado Pago", pero sí columnas características.
    if _es_mercadopago_por_headers(ws):
        return "mercadopago"
    return "generico"


def _es_mercadopago_por_headers(ws) -> bool:
    """True si la planilla tiene 2+ cabeceras típicas del reporte MP."""
    hits = set()
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, min(40, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if not v:
                continue
            h = _normalizar(str(v).lower().strip())
            for kw in KEYWORDS_MP_HEADER:
                if kw in h:
                    hits.add(kw)
        if len(hits) >= 2:
            return True
    return len(hits) >= 2

def _parse_fecha(v):
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def _parse_monto(v):
    # Devuelve float; la conversión a Decimal la hace SQLAlchemy en columnas Numeric(12,2).
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str):
        s = v.strip().replace("$","").replace("\xa0","").replace(" ","")
        if not s or s in ("-",""):
            return None
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".","").replace(",",".")
            else:
                s = s.replace(",","")
        elif "," in s:
            s = s.replace(",",".")
        try:
            return round(float(s), 2)
        except ValueError:
            return None
    return None

KEYWORDS_MONTO   = ["importe","monto","amount","transaction_amount","net_amount"]
KEYWORDS_CREDITO = ["credito","haber","credit"]
KEYWORDS_DEBITO  = ["debito","debe","debit","cargo"]
KEYWORDS_FECHA   = ["fecha","date","vencimiento","release_date","date_created","money_release_date"]
KEYWORDS_TITULAR = ["titular","concepto","descripcion","glosa","detalle","nombre","beneficiario","ordenante",
                    "transaction_type","payer_name","description"]
KEYWORDS_ORDEN   = ["orden","nro. de referencia","nro","numero","secuencia","referencia",
                    "operation_id","source_id","external_reference","payment_id"]
KEYWORDS_SALDO   = ["saldo","balance","balance_amount"]
KEYWORDS_MES     = ["mes","period","periodo"]

# Cabeceras características del CSV de liberaciones/settlement de Mercado Pago.
# Si aparecen 2+ de estas, es un reporte MP aunque no diga "Mercado Pago".
KEYWORDS_MP_HEADER = ["transaction_amount", "transaction_type", "balance_amount",
                      "source_id", "release_date", "money_release_date",
                      "external_reference", "payment_method_type"]
KEYWORDS_CLIENTE_ACRED = ["cliente acreditado", "cliente acredit"]
KEYWORDS_FECHA_ACRED   = ["fecha acred", "fecha de acred"]

# Bloquear "cliente" / "acred" como TITULAR (no confundir con concepto),
# pero seguir permitiendolos como columnas propias de cliente_acred/fecha_acred.
BLOCKLIST = {"acred", "acreditad", "cliente"}

def _normalizar(s: str) -> str:
    """Quita acentos y diacríticos para que 'Débitos' matchee 'debito'."""
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def _match_kw(header, keywords):
    h = _normalizar(header.lower().strip())
    if any(b in h for b in BLOCKLIST):
        return False
    return any(k in h for k in keywords)

def _match_kw_raw(header, keywords):
    h = _normalizar(header.lower().strip())
    return any(k in h for k in keywords)

def detectar_columnas(ws):
    cols = {"hdr_row":1,"monto":None,"credito":None,"debito":None,"fecha":None,
            "titular":None,"orden":None,"saldo":None,"mes":None,
            "cliente_acred":None,"fecha_acred":None}
    for r in range(1, min(13, ws.max_row + 1)):
        encontrados = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(r, c).value or "").lower().strip()
            if not h:
                continue
            # Cliente acreditado / Fecha acred (columnas propias del extracto exportado)
            if _match_kw_raw(h, KEYWORDS_CLIENTE_ACRED):
                encontrados.setdefault("cliente_acred", c)
                continue
            if _match_kw_raw(h, KEYWORDS_FECHA_ACRED):
                encontrados.setdefault("fecha_acred", c)
                continue
            if _match_kw(h, KEYWORDS_MONTO):
                encontrados.setdefault("monto", c)
            elif _match_kw(h, KEYWORDS_CREDITO):
                encontrados.setdefault("credito", c)
            elif _match_kw(h, KEYWORDS_DEBITO):
                encontrados.setdefault("debito", c)
            elif _match_kw(h, KEYWORDS_FECHA):
                encontrados.setdefault("fecha", c)
            elif _match_kw(h, KEYWORDS_TITULAR):
                encontrados.setdefault("titular", c)
            elif _match_kw(h, KEYWORDS_ORDEN):
                encontrados.setdefault("orden", c)
            elif _match_kw(h, KEYWORDS_SALDO):
                encontrados.setdefault("saldo", c)
            elif _match_kw(h, KEYWORDS_MES):
                encontrados.setdefault("mes", c)
        # Fila de headers válida si tiene monto unitario O par crédito/débito
        if "monto" in encontrados or "credito" in encontrados or "debito" in encontrados:
            cols["hdr_row"] = r
            cols.update(encontrados)
            return cols
    # Fallback Banco Macro extracto
    cols["hdr_row"] = 2
    cols["orden"]   = 2
    cols["fecha"]   = 3
    cols["mes"]     = 4
    cols["titular"] = 5
    cols["monto"]   = 6
    cols["saldo"]   = 7
    return cols

def parsear_generico(ws, cols):
    movimientos = []
    hdr = cols["hdr_row"]
    for row in range(hdr + 1, ws.max_row + 1):
        if cols["monto"]:
            # Columna única de monto (Banco Macro y formato genérico)
            monto = _parse_monto(ws.cell(row, cols["monto"]).value)
        else:
            # Columnas separadas crédito/débito (BBVA, Santander, Galicia)
            credito = _parse_monto(ws.cell(row, cols["credito"]).value) if cols.get("credito") else None
            debito  = _parse_monto(ws.cell(row, cols["debito"]).value)  if cols.get("debito")  else None
            credito = abs(credito) if credito is not None else 0.0
            debito  = abs(debito)  if debito  is not None else 0.0
            if credito == 0.0 and debito == 0.0:
                monto = None
            elif credito > 0.0 and debito == 0.0:
                monto = credito
            elif debito > 0.0 and credito == 0.0:
                monto = -debito
            else:
                # Ambas columnas con valor: neto (poco común, pero posible)
                monto = round(credito - debito, 2)
        if monto is None or abs(monto) < 0.01:
            continue
        fecha   = _parse_fecha(ws.cell(row, cols["fecha"]).value)   if cols["fecha"]   else None
        titular = str(ws.cell(row, cols["titular"]).value or "")     if cols["titular"] else ""
        orden   = ws.cell(row, cols["orden"]).value                  if cols["orden"]   else None
        saldo   = _parse_monto(ws.cell(row, cols["saldo"]).value)    if cols["saldo"]   else None
        mes_val = str(ws.cell(row, cols["mes"]).value or "")         if cols["mes"]     else None

        cliente_acred = None
        fecha_acred = None
        if cols.get("cliente_acred"):
            v = ws.cell(row, cols["cliente_acred"]).value
            if v is not None and str(v).strip():
                cliente_acred = str(v).strip()
        if cols.get("fecha_acred"):
            fecha_acred = _parse_fecha(ws.cell(row, cols["fecha_acred"]).value)

        # Mes: SOLO el numero (ej "5"), no "Mayo 2026".
        # Si en el Excel viene "Mayo" o "5" o "05/2026", normalizamos a numero.
        def _mes_normalizado(v, f):
            if v:
                s = str(v).strip()
                # si ya es un numero
                try:
                    n = int(float(s))
                    if 1 <= n <= 12:
                        return str(n)
                except ValueError:
                    pass
            return str(f.month) if f else None

        movimientos.append({
            "orden":   int(orden) if isinstance(orden,(int,float)) else None,
            "fecha":   fecha,
            "mes":     _mes_normalizado(mes_val, fecha),
            "titular": titular.strip() or None,
            "monto":   monto,
            "saldo":   saldo,
            "cliente_acreditado": cliente_acred,
            "fecha_acred": fecha_acred,
        })
    return movimientos

def parsear_extracto_bancario(filepath: str) -> dict:
    """
    Parsea un extracto bancario detectando automaticamente el banco y formato.
    Soporta .xlsx, .xls y .csv.
    Prueba todas las hojas del archivo y devuelve la que tiene mas movimientos.
    """
    xlsx_path, necesita_cleanup = preparar_archivo(filepath)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        mejor_movs = []
        banco_detectado = "generico"
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 3:
                continue
            banco = detectar_banco(ws)
            cols  = detectar_columnas(ws)
            movs  = parsear_generico(ws, cols)
            if len(movs) > len(mejor_movs):
                mejor_movs = movs
                banco_detectado = banco
        wb.close()
        return {"movimientos": mejor_movs, "total": len(mejor_movs), "banco_detectado": banco_detectado}
    finally:
        if necesita_cleanup and os.path.exists(xlsx_path):
            os.remove(xlsx_path)

def detectar_header(ws):
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(r, c).value or '').lower().strip()
            if 'monto' in h or 'importe' in h:
                data_start = r + 1
                for test_col in [c, c + 1]:
                    hits = sum(
                        1 for dr in range(data_start, min(data_start + 6, ws.max_row + 1))
                        if isinstance(ws.cell(dr, test_col).value, (int, float))
                        and 0 < ws.cell(dr, test_col).value < 5_000_000
                    )
                    if hits >= 1:
                        return r, test_col
    return 2, 6

def detectar_cuit_col(ws, hdr_row):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hdr_row, c).value or '').lower().strip()
        if 'cuit' in h or 'cuil' in h:
            return c
    return None

def detectar_titular_col(ws, hdr_row):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hdr_row, c).value or '').lower().strip()
        if 'titular' in h or 'nombre' in h or 'cliente' in h:
            return c
    return None

def parsear_planilla_cliente(filepath: str) -> dict:
    xlsx_path, necesita_cleanup = preparar_archivo(filepath)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        hdr_row, imp_col = detectar_header(ws)
        cuit_col    = detectar_cuit_col(ws, hdr_row)
        titular_col = detectar_titular_col(ws, hdr_row)
        filas = []
        for row in range(hdr_row + 1, ws.max_row + 1):
            monto = ws.cell(row, imp_col).value
            if monto is None:
                continue
            cuit    = ws.cell(row, cuit_col).value    if cuit_col    else None
            titular = ws.cell(row, titular_col).value if titular_col else None
            filas.append({"monto": monto, "cuit": str(cuit) if cuit else None, "titular": str(titular) if titular else None})
        wb.close()
        return {"filas": filas, "total": len(filas), "header_row": hdr_row, "imp_col": imp_col, "cuit_col": cuit_col, "titular_col": titular_col}
    finally:
        if necesita_cleanup and os.path.exists(xlsx_path):
            os.remove(xlsx_path)