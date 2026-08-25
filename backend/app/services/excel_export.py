"""Generacion de archivos Excel para download (movimientos, historial)"""

import io
from decimal import Decimal
from typing import List
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ARG = ZoneInfo('America/Argentina/Buenos_Aires')

def _now() -> datetime:
    """Hora actual en zona horaria Argentina."""
    return datetime.now(_ARG)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
HDR_FILL = PatternFill("solid", fgColor="3483FA")
ML_YELLOW = PatternFill("solid", fgColor="FFE600")
TITLE_FONT = Font(bold=True, size=14, color="333333")

# ── Paleta de marca Cuadra (verde) — usada en el export de planilla conciliada ──
CUADRA_GREEN = "16A34A"        # verde de marca
CUADRA_GREEN_DARK = "15803D"   # verde oscuro para headers (contraste con texto blanco)
CUADRA_GREEN_TINT = "DCFCE7"   # verde muy claro para fila de total
AMBER_TINT = "FFF3CD"          # ámbar claro (diferencia / estado ambiguo)
AMBER_FONT = "92650A"
RED_TINT = "F8D7DA"
RED_FONT = "842029"
GREEN_FONT = "155724"
HDR_FILL_CUADRA = PatternFill("solid", fgColor=CUADRA_GREEN_DARK)
MONEY_FMT_ARS = '"$ "#,##0.00'   # formato es-AR: $ 1.234.567,89 (Excel adapta separadores al locale)
TOTAL_BORDER = Border(top=Side(style="medium", color=CUADRA_GREEN_DARK))


def _mes_a_int(v):
    """Convierte el campo 'mes' a int para evitar el warning 'numero como texto'
    en Excel. Si no es numerico (ej. None), devuelve None."""
    if v is None:
        return None
    try:
        n = int(float(str(v).strip()))
        return n if 1 <= n <= 12 else None
    except (ValueError, TypeError):
        return None


def _autosize(ws, max_col):
    """Ajustar ancho de columnas al contenido"""
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 8
        for row in ws[letter]:
            v = row.value
            if v is not None:
                length = len(str(v))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def _hdr(ws, row, headers):
    """Aplicar formato a fila de header"""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def export_movimientos(extracto_nombre: str, movimientos: List[dict]) -> bytes:
    """Genera xlsx con la lista de movimientos filtrados"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Titulo
    ws.cell(row=1, column=1, value="Movimientos del extracto").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Extracto: {extracto_nombre}").font = Font(italic=True, color="666666")
    ws.cell(row=3, column=1, value=f"Generado: {_now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="666666")

    headers = ["Orden", "Fecha", "Mes", "Titular", "Importe", "Saldo", "Cliente acreditado", "Fecha acred."]
    _hdr(ws, 5, headers)

    for i, m in enumerate(movimientos, start=6):
        ws.cell(row=i, column=1, value=m.get("orden"))
        f = m.get("fecha")
        ws.cell(row=i, column=2, value=f).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=3, value=_mes_a_int(m.get("mes")))
        ws.cell(row=i, column=4, value=m.get("titular"))
        ws.cell(row=i, column=5, value=m.get("monto")).number_format = '"$"#,##0.00'
        s = m.get("saldo")
        if s is not None:
            ws.cell(row=i, column=6, value=s).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=7, value=m.get("cliente_acreditado"))
        fa = m.get("fecha_acred")
        if fa:
            ws.cell(row=i, column=8, value=fa).number_format = "DD/MM/YYYY"

        for col in range(1, 9):
            cell = ws.cell(row=i, column=col)
            cell.border = BORDER
            # wrap_text=False fuerza que el titular largo NO agrande la fila
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "general",
                                       vertical="center", wrap_text=False)
        # Alto 15 estandar Excel (= ~25px en pantalla)
        ws.row_dimensions[i].height = 15

    ws.auto_filter.ref = f"A5:H{5 + len(movimientos)}"
    _autosize(ws, 8)
    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_planilla_conciliada(planilla_data: dict, movimientos_acreditados: List[dict]) -> bytes:
    """
    Genera un xlsx con 2 hojas:
      Hoja 1 - Planilla del cliente con columna de estado (ok / no esta / faltan datos / etc.)
      Hoja 2 - Movimientos bancarios acreditados a esta planilla

    `planilla_data` puede incluir opcionalmente "total_declarado" (Decimal | None):
    el total que el cliente declaró en su planilla, usado para la fila de cuadre
    al final de la hoja 1. No cambia columnas ni orden existentes (compatibilidad
    con el contador) — solo agrega encabezado de marca, formato y fila de totales.
    """
    wb = openpyxl.Workbook()
    now = _now()

    # ── HOJA 1: planilla del cliente con estado ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Planilla cliente"

    ws1.cell(row=1, column=1, value=f"Cliente: {planilla_data['cliente_nombre']}").font = TITLE_FONT
    ws1.cell(row=2, column=1, value=f"Archivo: {planilla_data['nombre_archivo']}").font = Font(italic=True, color="666666")
    ws1.cell(row=3, column=1, value=f"Generado: {now.strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="666666")
    marca = ws1.cell(row=1, column=9, value="Cuadra")
    marca.font = Font(italic=True, bold=True, size=10, color=CUADRA_GREEN)
    marca.alignment = Alignment(horizontal="right")

    # Hoja 1: columnas del cliente + movimiento del extracto + Estado AL FINAL
    h1 = ["#", "Importe", "CUIT", "Titular planilla", "Orden mov.", "Titular extracto", "Fecha mov.", "Fecha acred.", "Estado"]
    for col, h in enumerate(h1, start=1):
        c = ws1.cell(row=5, column=col, value=h)
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.fill = HDR_FILL_CUADRA
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    STATUS_COLORS = {
        "ok": ("D4EDDA", GREEN_FONT),
        "ok (aprendido)": ("D4EDDA", GREEN_FONT),
        "no está": (RED_TINT, RED_FONT),
        "VENCIDO": (RED_TINT, RED_FONT),
        "faltan datos": (AMBER_TINT, AMBER_FONT),
        "EN_REVISION": (AMBER_TINT, AMBER_FONT),
        "duplicado": (AMBER_TINT, AMBER_FONT),
        "PAGO_PARCIAL": (AMBER_TINT, AMBER_FONT),
        "CONCILIADO_CON_DIFERENCIA": (AMBER_TINT, AMBER_FONT),
    }

    rows_data = planilla_data["rows"]
    total_monto = Decimal("0")
    for i, row in enumerate(rows_data, start=6):
        monto = row["monto"] if row.get("monto") is not None else Decimal("0")
        total_monto += Decimal(str(monto))
        ws1.cell(row=i, column=1, value=i - 5)
        ws1.cell(row=i, column=2, value=monto).number_format = MONEY_FMT_ARS
        ws1.cell(row=i, column=3, value=row.get("cuit") or "")
        ws1.cell(row=i, column=4, value=row.get("titular") or "")
        ws1.cell(row=i, column=5, value=row.get("orden_movimiento_acreditado"))
        ws1.cell(row=i, column=6, value=row.get("mov_titular") or "")
        f = row.get("mov_fecha")
        ws1.cell(row=i, column=7, value=f).number_format = "DD/MM/YYYY"
        fa = row.get("mov_fecha_acred")
        ws1.cell(row=i, column=8, value=fa).number_format = "DD/MM/YYYY"
        # Estado ULTIMA columna con color
        st = row["status"]
        status_cell = ws1.cell(row=i, column=9, value=st)
        fill_color, font_color = STATUS_COLORS.get(st, ("FFFFFF", "333333"))
        if isinstance(st, str) and st.startswith("acreditado"):
            fill_color, font_color = STATUS_COLORS["duplicado"]
        status_cell.fill = PatternFill("solid", fgColor=fill_color)
        status_cell.font = Font(color=font_color)
        for col in range(1, 10):
            ws1.cell(row=i, column=col).border = BORDER

    # ── Fila de TOTAL ─────────────────────────────────────────────────────────
    last_data_row = 5 + len(rows_data)
    total_row = last_data_row + 1
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    tc = ws1.cell(row=total_row, column=2, value=total_monto)
    tc.number_format = MONEY_FMT_ARS
    tc.font = Font(bold=True)
    for col in range(1, 10):
        cell = ws1.cell(row=total_row, column=col)
        cell.fill = PatternFill("solid", fgColor=CUADRA_GREEN_TINT)
        cell.border = TOTAL_BORDER

    total_declarado = planilla_data.get("total_declarado")
    if total_declarado is not None:
        total_declarado = Decimal(str(total_declarado))
        fila_decl = total_row + 1
        ws1.cell(row=fila_decl, column=1, value="Total declarado por el cliente").font = Font(italic=True)
        dc = ws1.cell(row=fila_decl, column=2, value=total_declarado)
        dc.number_format = MONEY_FMT_ARS
        dc.font = Font(italic=True)

        diferencia = total_declarado - total_monto
        fila_dif = fila_decl + 1
        ws1.cell(row=fila_dif, column=1, value="Diferencia").font = Font(bold=True)
        difc = ws1.cell(row=fila_dif, column=2, value=diferencia)
        difc.number_format = MONEY_FMT_ARS
        difc.font = Font(bold=True, color=GREEN_FONT if diferencia == 0 else AMBER_FONT)
        color_dif = CUADRA_GREEN_TINT if diferencia == 0 else AMBER_TINT
        for col in range(1, 10):
            ws1.cell(row=fila_dif, column=col).fill = PatternFill("solid", fgColor=color_dif)

    ws1.auto_filter.ref = f"A5:I{last_data_row}"
    ws1.freeze_panes = "A6"
    _autosize(ws1, 9)

    # ── HOJA 2: movimientos del extracto acreditados ─────────────────────────
    ws2 = wb.create_sheet("Movimientos acreditados")

    ws2.cell(row=1, column=1, value="Movimientos bancarios acreditados a esta planilla").font = TITLE_FONT

    # Headers estilo extracto Macro (azul)
    h2 = [None, "Orden", "Fecha", "Mes", "Titular / CUIT", "Importe Pesos", "Saldo", "Cliente", "Fecha acred."]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        if h:
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")

    for i, m in enumerate(movimientos_acreditados, start=4):
        ws2.cell(row=i, column=2, value=m.get("orden"))
        f = m.get("fecha")
        ws2.cell(row=i, column=3, value=f).number_format = "DD/MM/YYYY"
        ws2.cell(row=i, column=4, value=_mes_a_int(m.get("mes")))
        ws2.cell(row=i, column=5, value=m.get("titular"))
        ws2.cell(row=i, column=6, value=m.get("monto")).number_format = '"$"#,##0.00'
        s = m.get("saldo")
        if s is not None:
            ws2.cell(row=i, column=7, value=s).number_format = '"$"#,##0.00'
        ws2.cell(row=i, column=8, value=m.get("cliente_acreditado"))
        fa = m.get("fecha_acred")
        ws2.cell(row=i, column=9, value=fa).number_format = "DD/MM/YYYY"
        for col in range(2, 10):
            ws2.cell(row=i, column=col).border = BORDER

    ws2.freeze_panes = "B4"
    _autosize(ws2, 9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_extracto_contador(extracto_nombre: str, movimientos: List[dict]) -> bytes:
    """
    Export para el contador.
    Hoja 1: idéntica al extracto original (header azul Macro, orden descendente,
             filtros Excel activos, editable) + columnas cliente acreditado y fecha acred.
    Hoja 2: resumen estadístico.
    """
    # Colores Banco Macro (igual que en pantalla)
    MACRO_BLUE   = PatternFill("solid", fgColor="3483FA")
    MACRO_FONT   = Font(color="FFFFFF", bold=True, size=10)
    GREEN_FILL   = PatternFill("solid", fgColor="D4EDDA")
    GREEN_FONT   = Font(color="155724")
    TOTAL_FONT   = Font(bold=True, size=11)
    TOTAL_FILL   = PatternFill("solid", fgColor="E2EAF7")
    THIN_BLUE    = Side(style="thin", color="3483FA")
    Border(bottom=THIN_BLUE)

    now_str = _now().strftime('%d/%m/%Y %H:%M')

    # Ordenar: mayor orden primero (último movimiento arriba, como el original)
    movs_sorted = sorted(
        movimientos,
        key=lambda m: (m.get("orden") or 0),
        reverse=True
    )

    acreditados = [m for m in movimientos if m.get("cliente_acreditado")]
    libres      = [m for m in movimientos if not m.get("cliente_acreditado")]
    total_acred_monto = sum(m.get("monto", 0) for m in acreditados)

    wb = openpyxl.Workbook()

    # ── Hoja 1: extracto igual al original ───────────────────────────────────
    ws = wb.active
    ws.title = "Extracto conciliado"

    # Fila 1: título igual al banco
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value=extracto_nombre)
    t.font = Font(bold=True, size=12, color="1B3F73")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Fila 2: fecha de generación
    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1,
            value=f"Generado: {now_str}  ·  {len(acreditados)} acreditados  ·  {len(libres)} libres"
            ).font = Font(italic=True, size=9, color="666666")

    # Fila 3: headers azul Macro
    HDR_ROW = 3
    headers = ["Orden", "Fecha", "Mes", "Titular", "Importe", "Saldo", "Cliente acreditado", "Fecha acred."]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.font   = MACRO_FONT
        c.fill   = MACRO_BLUE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        c.border = Border(
            left=Side(style="thin", color="2968C8"),
            right=Side(style="thin", color="2968C8"),
            top=Side(style="thin", color="2968C8"),
            bottom=Side(style="thin", color="2968C8"),
        )
    ws.row_dimensions[HDR_ROW].height = 18

    # AutoFilter en fila 3 (para trabajarlo desde Excel)
    ws.auto_filter.ref = f"A{HDR_ROW}:H{HDR_ROW + len(movs_sorted)}"
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # Datos
    for i, m in enumerate(movs_sorted, start=HDR_ROW + 1):
        acred = bool(m.get("cliente_acreditado"))
        fill  = GREEN_FILL if acred else None

        def wr(col, value, fmt=None, align=None):
            c = ws.cell(row=i, column=col, value=value)
            if fill:
                c.fill = fill
            c.border = BORDER
            if fmt:
                c.number_format = fmt
            # vertical=center + wrap_text=False: la fila NO se agranda
            # aunque el titular sea largo (estilo Excel Macro original).
            c.alignment = Alignment(
                horizontal=align or "general",
                vertical="center",
                wrap_text=False,
            )
            return c

        wr(1, m.get("orden"), align="center")
        wr(2, m.get("fecha"), "DD/MM/YYYY")
        wr(3, _mes_a_int(m.get("mes")))
        wr(4, m.get("titular"))
        wr(5, m.get("monto"), '"$"#,##0.00', "right")
        saldo = m.get("saldo")
        wr(6, saldo if saldo is not None else None,
           '"$"#,##0.00' if saldo is not None else None, "right")
        c7 = wr(7, m.get("cliente_acreditado") or "")
        if acred:
            c7.font = GREEN_FONT
        wr(8, m.get("fecha_acred"), "DD/MM/YYYY")

        # Alto 15 estandar Excel (= ~25px en pantalla)
        ws.row_dimensions[i].height = 15

    # Fila de totales al pie
    tot_row = HDR_ROW + len(movs_sorted) + 1
    ws.cell(row=tot_row, column=4, value="TOTAL ACREDITADO").font = TOTAL_FONT
    tc = ws.cell(row=tot_row, column=5, value=total_acred_monto)
    tc.font = TOTAL_FONT
    tc.number_format = '"$"#,##0.00'
    tc.fill = TOTAL_FILL
    tc.alignment = Alignment(horizontal="right")
    ws.cell(row=tot_row, column=7,
            value=f"{len(acreditados)} de {len(movimientos)} movimientos acreditados"
            ).font = Font(italic=True, size=9, color="3483FA")
    for col in [4, 5, 6, 7, 8]:
        ws.cell(row=tot_row, column=col).fill = TOTAL_FILL

    # Anchos fijos similares al original
    ws.column_dimensions["A"].width = 8   # Orden
    ws.column_dimensions["B"].width = 12  # Fecha
    ws.column_dimensions["C"].width = 10  # Mes
    ws.column_dimensions["D"].width = 38  # Titular
    ws.column_dimensions["E"].width = 14  # Importe
    ws.column_dimensions["F"].width = 14  # Saldo
    ws.column_dimensions["G"].width = 18  # Cliente acreditado
    ws.column_dimensions["H"].width = 14  # Fecha acred.

    # ── Hoja 2: resumen ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    ws2.cell(row=1, column=1, value="Resumen del extracto").font = Font(bold=True, size=13, color="1B3F73")
    ws2.cell(row=2, column=1, value=f"Extracto: {extracto_nombre}").font = Font(italic=True, color="666666")
    ws2.cell(row=3, column=1, value=f"Generado: {now_str}").font = Font(italic=True, color="666666")

    pct = round(len(acreditados) / len(movimientos) * 100, 1) if movimientos else 0
    stats = [
        ("Total movimientos", len(movimientos)),
        ("Movimientos acreditados", len(acreditados)),
        ("Movimientos libres", len(libres)),
        ("% acreditado", f"{pct}%"),
        ("", ""),
        ("Total $ acreditado", total_acred_monto),
        ("Total $ libre", sum(m.get("monto", 0) for m in libres)),
    ]
    for j, (label, val) in enumerate(stats, start=5):
        ws2.cell(row=j, column=1, value=label).font = Font(bold=bool(label))
        c = ws2.cell(row=j, column=2, value=val)
        if isinstance(val, float):
            c.number_format = '"$"#,##0.00'
            c.font = TOTAL_FONT

    ws2.cell(row=13, column=1, value="Detalle por cliente").font = Font(bold=True, size=11, color="3483FA")
    clientes: dict = {}
    for m in acreditados:
        cl = m.get("cliente_acreditado", "")
        clientes[cl] = clientes.get(cl, 0) + 1
    for k, (cl, cnt) in enumerate(sorted(clientes.items()), start=14):
        ws2.cell(row=k, column=1, value=cl)
        ws2.cell(row=k, column=2, value=cnt)
        ws2.cell(row=k, column=2).number_format = '0" movimientos"'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_historial_planillas(planillas: List[dict]) -> bytes:
    """Genera xlsx con el historial de planillas reconciliadas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    ws.cell(row=1, column=1, value="Historial de reconciliaciones").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generado: {_now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="666666")

    headers = ["Cliente", "Archivo", "Fecha carga", "Usuario", "Total filas", "Acreditadas", "No encontradas", "Duplicadas", "Sin datos", "% acred."]
    _hdr(ws, 4, headers)

    for i, p in enumerate(planillas, start=5):
        ws.cell(row=i, column=1, value=p["cliente_nombre"])
        ws.cell(row=i, column=2, value=p["nombre_archivo"])
        ws.cell(row=i, column=3, value=p["fecha_carga"]).number_format = "DD/MM/YYYY HH:MM"
        ws.cell(row=i, column=4, value=p["usuario_nombre"])
        ws.cell(row=i, column=5, value=p["total_filas"])
        ws.cell(row=i, column=6, value=p["acreditadas"])
        ws.cell(row=i, column=7, value=p["no_encontradas"])
        ws.cell(row=i, column=8, value=p["duplicadas"])
        ws.cell(row=i, column=9, value=p["sin_datos"])
        pct = (p["acreditadas"] / p["total_filas"]) if p["total_filas"] > 0 else 0
        ws.cell(row=i, column=10, value=pct).number_format = "0%"

        for col in range(1, 11):
            ws.cell(row=i, column=col).border = BORDER

    _autosize(ws, 10)
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_backup_completo(org_nombre: str, planillas: List[dict], extractos: List[dict]) -> bytes:
    """Backup completo: Hoja1=planillas conciliadas, Hoja2=extractos"""
    wb = openpyxl.Workbook()
    now_str = _now().strftime('%d/%m/%Y %H:%M')

    # ── Hoja 1: Planillas ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Planillas"
    ws1.cell(row=1, column=1, value=f"Backup — {org_nombre}").font = TITLE_FONT
    ws1.cell(row=2, column=1, value=f"Generado: {now_str}").font = Font(italic=True, color="666666")

    hdrs1 = ["ID","Cliente","Archivo","Fecha carga","Usuario","Total","OK","No está","Dup.","Sin datos"]
    _hdr(ws1, 4, hdrs1)
    for i, p in enumerate(planillas, start=5):
        ws1.cell(row=i, column=1, value=p.get("id"))
        ws1.cell(row=i, column=2, value=p.get("cliente_nombre"))
        ws1.cell(row=i, column=3, value=p.get("nombre_archivo"))
        ws1.cell(row=i, column=4, value=p.get("fecha_carga")).number_format = "DD/MM/YYYY HH:MM"
        ws1.cell(row=i, column=5, value=p.get("usuario_nombre"))
        ws1.cell(row=i, column=6, value=p.get("total_filas"))
        ws1.cell(row=i, column=7, value=p.get("acreditadas"))
        ws1.cell(row=i, column=8, value=p.get("no_encontradas"))
        ws1.cell(row=i, column=9, value=p.get("duplicadas"))
        ws1.cell(row=i, column=10, value=p.get("sin_datos"))
        for col in range(1, 11):
            ws1.cell(row=i, column=col).border = BORDER
    ws1.auto_filter.ref = f"A4:J{4+len(planillas)}"
    _autosize(ws1, 10)
    ws1.freeze_panes = "A5"

    # ── Hoja 2: Extractos ──────────────────────────────────────
    ws2 = wb.create_sheet("Extractos")
    ws2.cell(row=1, column=1, value=f"Extractos — {org_nombre}").font = TITLE_FONT
    hdrs2 = ["ID","Archivo","Fecha creación","Movimientos"]
    _hdr(ws2, 3, hdrs2)
    for i, e in enumerate(extractos, start=4):
        ws2.cell(row=i, column=1, value=e.get("id"))
        ws2.cell(row=i, column=2, value=e.get("nombre_archivo"))
        ws2.cell(row=i, column=3, value=e.get("fecha_creacion")).number_format = "DD/MM/YYYY HH:MM"
        ws2.cell(row=i, column=4, value=e.get("total_movimientos"))
        for col in range(1, 5):
            ws2.cell(row=i, column=col).border = BORDER
    _autosize(ws2, 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_liquidacion_excel(liquidacion, revisiones) -> bytes:
    """
    Excel de liquidacion con 3 hojas:
    Hoja 1 — Resumen ejecutivo
    Hoja 2 — Detalle por cliente
    Hoja 3 — Log de revisiones manuales
    """
    ESTADO_FILL = {
        "borrador": PatternFill("solid", fgColor="FEF3C7"),
        "aprobada": PatternFill("solid", fgColor="D1FAE5"),
        "pagada":   PatternFill("solid", fgColor="DBEAFE"),
    }
    now_str = _now().strftime('%d/%m/%Y %H:%M')
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ejecutivo ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen ejecutivo"

    ws1.merge_cells("A1:D1")
    ws1.cell(row=1, column=1, value="LIQUIDACIÓN").font = Font(bold=True, size=14, color="1B3F73")
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal="left")
    ws1.cell(row=2, column=1, value=f"Período: {liquidacion.periodo_inicio} → {liquidacion.periodo_fin}").font = Font(italic=True, color="666666")
    ws1.cell(row=3, column=1, value=f"Generado: {now_str}").font = Font(italic=True, color="666666")

    # Estado
    estado_cell = ws1.cell(row=4, column=1, value=f"Estado: {liquidacion.estado.upper()}")
    estado_cell.font = Font(bold=True, size=11)
    fill = ESTADO_FILL.get(liquidacion.estado, PatternFill("solid", fgColor="F3F4F6"))
    for col in range(1, 5):
        ws1.cell(row=4, column=col).fill = fill

    # Totales
    totales = [
        ("", ""),
        ("Total conciliado", liquidacion.total_conciliado),
        ("Total comisión", liquidacion.total_comision),
        ("Total neto", liquidacion.total_neto),
    ]
    for i, (label, val) in enumerate(totales, start=6):
        if not label:
            continue
        c_label = ws1.cell(row=i, column=1, value=label)
        c_label.font = Font(bold=True)
        c_val = ws1.cell(row=i, column=2, value=val)
        c_val.number_format = '"$"#,##0.00'
        c_val.font = Font(bold=True, size=12)
        c_val.alignment = Alignment(horizontal="right")
        if label == "Total neto":
            c_val.font = Font(bold=True, size=13, color="155724")

    if liquidacion.notas:
        ws1.cell(row=11, column=1, value="Notas:").font = Font(bold=True)
        ws1.cell(row=12, column=1, value=liquidacion.notas).font = Font(italic=True, color="666666")

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 18

    # ── Hoja 2: Detalle por cliente ───────────────────────────────────────
    ws2 = wb.create_sheet("Detalle por cliente")
    ws2.cell(row=1, column=1, value=f"Detalle — {liquidacion.periodo_inicio} → {liquidacion.periodo_fin}").font = TITLE_FONT

    headers2 = ["Cliente", "Monto conciliado", "% Comisión", "Monto comisión", "Monto neto", "Observaciones"]
    _hdr(ws2, 3, headers2)

    for i, d in enumerate(liquidacion.detalles, start=4):
        ws2.cell(row=i, column=1, value=d.cliente_nombre)
        ws2.cell(row=i, column=2, value=d.monto_conciliado).number_format = '"$"#,##0.00'
        ws2.cell(row=i, column=3, value=d.porcentaje_comision).number_format = '0.00"%"'
        ws2.cell(row=i, column=4, value=d.monto_comision).number_format = '"$"#,##0.00'
        ws2.cell(row=i, column=5, value=d.monto_neto).number_format = '"$"#,##0.00'
        ws2.cell(row=i, column=6, value=d.observaciones or "")
        for col in range(1, 7):
            ws2.cell(row=i, column=col).border = BORDER

    # Fila totales
    tot = len(liquidacion.detalles) + 4
    ws2.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=tot, column=2, value=liquidacion.total_conciliado).number_format = '"$"#,##0.00'
    ws2.cell(row=tot, column=4, value=liquidacion.total_comision).number_format = '"$"#,##0.00'
    ws2.cell(row=tot, column=5, value=liquidacion.total_neto).number_format = '"$"#,##0.00'
    for col in [1,2,3,4,5]:
        ws2.cell(row=tot, column=col).fill = PatternFill("solid", fgColor="E2EAF7")
        ws2.cell(row=tot, column=col).font = Font(bold=True)

    ws2.auto_filter.ref = f"A3:F{tot-1}"
    _autosize(ws2, 6)
    ws2.freeze_panes = "A4"

    # ── Hoja 3: Log de revisiones manuales ───────────────────────────────
    ws3 = wb.create_sheet("Revisiones manuales")
    ws3.cell(row=1, column=1, value="Log de revisiones manuales del período").font = TITLE_FONT

    if not revisiones:
        ws3.cell(row=3, column=1, value="Sin revisiones manuales en este período").font = Font(italic=True, color="888888")
    else:
        headers3 = ["Planilla", "Cliente", "Monto", "Estado anterior", "Estado final", "Comentario"]
        _hdr(ws3, 3, headers3)
        for i, (row, planilla) in enumerate(revisiones, start=4):
            ws3.cell(row=i, column=1, value=planilla.nombre_archivo)
            ws3.cell(row=i, column=2, value=planilla.cliente.nombre if planilla.cliente else "")
            ws3.cell(row=i, column=3, value=row.monto).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=4, value="sin datos")
            ws3.cell(row=i, column=5, value=row.status)
            ws3.cell(row=i, column=6, value=row.comentario_revision or "")
            for col in range(1, 7):
                ws3.cell(row=i, column=col).border = BORDER
        _autosize(ws3, 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_eft_historial(ops: list, periodo: str = "") -> bytes:
    """
    Exporta el historial EFT identico a la planilla manual de Julieta.

    Hoja 1 - Historico: Fecha | Cliente | Importe (mismo formato que pago eft.xlsx)
    Hoja 2 - Tabla dinamica: totales por fecha
    """
    GREEN_HEADER = PatternFill("solid", fgColor="92D050")  # verde del original
    GREEN_FONT   = Font(bold=True, color="FFFFFF", size=11)
    MONEY_FMT    = '"$"#,##0.00'
    _now().strftime('%d/%m/%Y %H:%M')

    wb = openpyxl.Workbook()

    # ── Hoja 1: Historico ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Historico"

    # Headers identicos al original (verde, bold, blanco)
    for col, label in enumerate(["Fecha", "Cliente", "Importe"], start=1):
        c = ws1.cell(row=1, column=col, value=label)
        c.fill   = GREEN_HEADER
        c.font   = GREEN_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws1.row_dimensions[1].height = 16

    # Datos
    for i, op in enumerate(ops, start=2):
        fecha_val = op.get("fecha")
        ws1.cell(row=i, column=1, value=fecha_val).number_format = "DD-MMM"
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal="left")
        ws1.cell(row=i, column=2, value=op.get("cliente_nombre", ""))
        monto_cell = ws1.cell(row=i, column=3, value=op.get("importe", 0))
        monto_cell.number_format = MONEY_FMT
        for col in range(1, 4):
            ws1.cell(row=i, column=col).border = BORDER

    # Total al pie
    if ops:
        tot_row = len(ops) + 2
        ws1.cell(row=tot_row, column=2, value="TOTAL").font = Font(bold=True)
        tc = ws1.cell(row=tot_row, column=3,
                      value=sum(op.get("importe", 0) for op in ops))
        tc.number_format = MONEY_FMT
        tc.font = Font(bold=True)
        tc.fill = PatternFill("solid", fgColor="E2EFDA")

    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 16
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:C{len(ops) + 1}"

    # ── Hoja 2: Totales por fecha (tabla dinamica manual) ────────────────────
    ws2 = wb.create_sheet("Diario")

    # Agrupar por fecha
    por_fecha: dict = {}
    for op in ops:
        f = str(op.get("fecha", ""))
        por_fecha.setdefault(f, 0)
        por_fecha[f] += op.get("importe", 0)

    ws2.cell(row=1, column=1, value="Etiquetas de fila").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Suma de Importe").font = Font(bold=True)
    for col in [1, 2]:
        ws2.cell(row=1, column=col).fill   = GREEN_HEADER
        ws2.cell(row=1, column=col).font   = GREEN_FONT
        ws2.cell(row=1, column=col).border = BORDER

    for i, (fecha_str, total) in enumerate(sorted(por_fecha.items()), start=2):
        ws2.cell(row=i, column=1, value=fecha_str).border = BORDER
        c = ws2.cell(row=i, column=2, value=total)
        c.number_format = MONEY_FMT
        c.border = BORDER

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_cierre_mensual_xlsx(planillas: list, anio: int, mes: int) -> bytes:
    """Excel de cierre mensual: hoja Resumen + una hoja por cliente."""
    MESES_NOMBRES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    wb = openpyxl.Workbook()

    # Hoja resumen
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(1, 1, f"Cierre mensual — {MESES_NOMBRES[mes]} {anio}").font = TITLE_FONT
    ws.cell(2, 1, f"Generado: {_now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="666666")
    _hdr(ws, 4, ["Cliente", "Planilla", "Total filas", "Acreditadas", "Pendientes", "Monto OK", "Monto Pendiente"])
    row = 5
    for p in planillas:
        rows = p.rows if hasattr(p, 'rows') else []
        ok = [r for r in rows if r.status in ('ok', 'OK', 'PAGO_PARCIAL', 'CONCILIADO_CON_DIFERENCIA')]
        pendiente = [r for r in rows if r.status not in ('ok', 'OK', 'PAGO_PARCIAL', 'CONCILIADO_CON_DIFERENCIA', 'duplicado')]
        cliente_nombre = p.cliente.nombre if p.cliente else '—'
        ws.cell(row, 1, cliente_nombre).border = BORDER
        ws.cell(row, 2, p.nombre_archivo).border = BORDER
        ws.cell(row, 3, len(rows)).border = BORDER
        ws.cell(row, 4, len(ok)).border = BORDER
        ws.cell(row, 5, len(pendiente)).border = BORDER
        c6 = ws.cell(row, 6, float(sum(r.monto for r in ok)))
        c6.number_format = '"$"#,##0.00'
        c6.border = BORDER
        c7 = ws.cell(row, 7, float(sum(r.monto for r in pendiente)))
        c7.number_format = '"$"#,##0.00'
        c7.border = BORDER
        ws.row_dimensions[row].height = 15
        row += 1
    _autosize(ws, 7)
    ws.freeze_panes = "A5"

    # Una hoja por cliente
    for p in planillas:
        rows = p.rows if hasattr(p, 'rows') else []
        cliente_nombre = (p.cliente.nombre if p.cliente else 'Cliente')[:31]
        sheet_name = cliente_nombre[:28] + (f"_{p.id}" if len(cliente_nombre) >= 28 else "")
        # ensure unique sheet names
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:25] + f"_{p.id}"
        ws2 = wb.create_sheet(title=sheet_name)
        ws2.cell(1, 1, f"{cliente_nombre} — {p.nombre_archivo}").font = TITLE_FONT
        _hdr(ws2, 3, ["#", "CUIT", "Titular", "Monto", "Estado", "Titular extracto", "Fecha acred."])
        for i, r in enumerate(rows, start=4):
            ws2.cell(i, 1, i - 3).border = BORDER
            ws2.row_dimensions[i].height = 15
            ws2.cell(i, 2, r.cuit or '—').border = BORDER
            ws2.cell(i, 3, r.titular or '—').border = BORDER
            c = ws2.cell(i, 4, float(r.monto) if r.monto is not None else 0)
            c.number_format = '"$"#,##0.00'
            c.border = BORDER
            ws2.cell(i, 5, r.status).border = BORDER
            ws2.cell(i, 6, r.mov_titular or '—').border = BORDER
            fa = r.mov_fecha_acred
            cfa = ws2.cell(i, 7, fa)
            if fa:
                cfa.number_format = "DD/MM/YYYY"
            cfa.border = BORDER
        _autosize(ws2, 7)
        ws2.freeze_panes = "A4"

    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    return buf2.getvalue()
