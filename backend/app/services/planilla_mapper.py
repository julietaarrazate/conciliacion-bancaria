"""Pipeline de estandarización de planillas de clientes → esquema canónico.

Estandariza CUALQUIER planilla (formatos heterogéneos: Tucu, Green, Dani, …) al
mismo esquema que consume el motor de conciliación:

    {monto: Decimal, cuit: str|None, titular: str|None, referencia: str|None, fecha: date|None}

Pipeline de 3 capas — la primera que resuelve con confianza alta gana:

  1. PERFIL / MANUAL: si viene un `mapeo` (dict con columnas + header_row) se usa
     directo. Si trae `fingerprint` y coincide con el del archivo → origen "perfil"
     (perfil aprendido del cliente). Si no trae fingerprint → origen "manual"
     (corrección del usuario desde el modal).
  2. HEURÍSTICA: detecta la fila de headers (escanea 1..15, soporta offset de
     columna), mapea columnas por diccionario de sinónimos y VALIDA el contenido
     (monto parsea, cuit son dígitos, titular no es constante — trampa "Caneland").
  3. IA (Gemini): solo si la heurística quedó con confianza < umbral. Degradación
     total: sin GEMINI_API_KEY o si Gemini falla → cae a la mejor apuesta de la
     heurística. NUNCA lanza excepción por la IA.

Índices de columna 1-based (estilo openpyxl) en todo el módulo.
"""

from __future__ import annotations

import os
import re
import json
import hashlib
import tempfile
import unicodedata

import openpyxl

from app.services.decimal_utils import to_decimal
from app.services.excel_parser import preparar_archivo, _parse_monto, _parse_fecha

# Campos canónicos, en orden de asignación (monto es el ancla).
FIELDS = ["monto", "cuit", "titular", "fecha", "referencia"]

# Umbral de confianza: por debajo se intenta la capa IA.
UMBRAL_CONFIANZA = 0.7

# Límites de escaneo (defensivos: las planillas tipo Tucu tienen ~1M filas vacías).
MAX_COLS = 60
MAX_HEADER_SCAN = 15
MAX_DATA_ROWS = 200_000
CORTE_FILAS_VACIAS = 3  # cortar tras N filas consecutivas con monto vacío


def _norm(s) -> str:
    """Normaliza un header: sin acentos, minúsculas, solo [a-z0-9]
    (quita espacios, barras, puntos, guiones)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


# Sinónimos ya normalizados con _norm. "cliente" y "cuenta" NO están en titular
# a propósito (trampa Caneland: son columnas constantes, no el titular real).
SINONIMOS = {
    "monto": {"monto", "importe", "total", "neto", "pago", "valor", "montotransferido"},
    "cuit": {"cuit", "cuil", "dni", "cuitcuildni", "cuitcuil", "documento", "cuitcuit"},
    "titular": {"titular", "titulardelacuenta", "titulartransferir", "nombre",
                "razonsocial", "beneficiario", "ordenante"},
    "fecha": {"fecha", "fechatransferencia", "fechadetransferencia", "fechadepago"},
    "referencia": {"referencia", "comprobante", "operacion", "orden", "id", "nrooperacion"},
}


def _score(header_norm: str, field: str) -> int:
    """0 = no matchea; 1 = header contenido en un sinónimo; 2 = un sinónimo
    contenido en el header; 3 = igualdad exacta. Mayor = mejor evidencia."""
    if not header_norm:
        return 0
    syns = SINONIMOS[field]
    if header_norm in syns:
        return 3
    for syn in syns:
        # Solo sinónimos de 3+ chars para containment, para que "id"/"dni" cortos
        # no matcheen por casualidad dentro de headers largos.
        if len(syn) >= 3 and syn in header_norm:
            return 2
    for syn in syns:
        if len(header_norm) >= 3 and header_norm in syn:
            return 1
    return 0


def _solo_digitos(v) -> str | None:
    """Extrae los dígitos de un valor (CUIT/CUIL/DNI). Acepta 10 u 11 dígitos."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        s = str(v)
    elif isinstance(v, float):
        s = str(int(v)) if v.is_integer() else repr(v)
    else:
        s = str(v).strip()
    d = re.sub(r"\D", "", s)
    return d or None


def _texto(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ── Carga del workbook desde bytes (soporta xlsx / xls / csv) ─────────────────

def _cargar_ws(contenido: bytes):
    """Escribe los bytes a un temp, los convierte a xlsx si hace falta y devuelve
    (wb, ws_activa, cleanup)."""
    head = contenido[:8]
    if head[:2] == b"PK":
        suffix = ".xlsx"
    elif head[:4] == b"\xD0\xCF\x11\xE0":
        suffix = ".xls"
    else:
        suffix = ".csv"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(contenido)
    tmp.close()

    xlsx_path, _need = preparar_archivo(tmp.name)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    def cleanup():
        try:
            wb.close()
        except Exception:
            pass
        for p in {tmp.name, xlsx_path}:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except OSError:
                    pass

    return wb, ws, cleanup


def _maxcol(ws) -> int:
    return min(ws.max_column or 1, MAX_COLS)


def _fingerprint(ws, header_row: int, maxcol: int) -> str:
    """sha1 de los headers normalizados (incluye posiciones vacías, así el offset
    de columna queda codificado)."""
    parts = []
    for c in range(1, maxcol + 1):
        parts.append(_norm(ws.cell(header_row, c).value))
    base = "|".join(parts).rstrip("|")
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


# ── Capa 2: heurística ────────────────────────────────────────────────────────

def _detectar_header_row(ws, maxcol: int) -> int:
    """Elige la fila (1..15) con más columnas que matchean el diccionario de
    sinónimos. No asume fila 1. Empate → la fila más arriba."""
    mejor_row, mejor_hits = 1, -1
    for r in range(1, min(MAX_HEADER_SCAN, ws.max_row or 1) + 1):
        hits = 0
        for c in range(1, maxcol + 1):
            h = _norm(ws.cell(r, c).value)
            if h and any(_score(h, f) > 0 for f in FIELDS):
                hits += 1
        if hits > mejor_hits:
            mejor_hits, mejor_row = hits, r
    return mejor_row


def _muestras_col(ws, header_row: int, col: int, n: int = 12) -> list:
    """Primeros `n` valores no vacíos de la columna, desde la fila de datos."""
    vals = []
    r = header_row + 1
    scanned = 0
    limite = (ws.max_row or header_row) + 1
    while r < limite and scanned < n * 4 and len(vals) < n:
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() != "":
            vals.append(v)
        r += 1
        scanned += 1
    return vals


def _valida_contenido(field: str, vals: list) -> bool:
    """Valida que el contenido de la columna sea coherente con el campo."""
    if not vals:
        return False
    if field == "monto":
        ok = sum(1 for v in vals if _parse_monto(v) is not None)
        return ok >= 1 and ok / len(vals) >= 0.6
    if field == "cuit":
        ok = 0
        for v in vals:
            d = _solo_digitos(v)
            if d and 7 <= len(d) <= 11:
                ok += 1
        return ok / len(vals) >= 0.5
    if field == "titular":
        textos = [str(v).strip() for v in vals if str(v).strip()]
        if not textos:
            return False
        # Trampa Caneland: columna constante (mismo valor en todas las filas) NO
        # es el titular.
        if len(textos) >= 2 and len({t.lower() for t in textos}) == 1:
            return False
        return True
    if field == "fecha":
        ok = sum(1 for v in vals if _parse_fecha(v) is not None)
        return ok / len(vals) >= 0.6
    if field == "referencia":
        return True
    return False


def _asignar_columnas(ws, header_row: int, maxcol: int):
    """Mapea cada campo canónico a una columna, validando contenido.
    Devuelve (columnas, validados, confianza)."""
    normh = {c: _norm(ws.cell(header_row, c).value) for c in range(1, maxcol + 1)}
    header_cols = [c for c in range(1, maxcol + 1) if normh[c]]

    columnas = {f: None for f in FIELDS}
    validados: set[str] = set()
    usados: set[int] = set()

    for field in FIELDS:
        cands = [c for c in header_cols if c not in usados and _score(normh[c], field) > 0]
        cands.sort(key=lambda c: _score(normh[c], field), reverse=True)
        chosen, chosen_valid = None, False
        for c in cands:
            if _valida_contenido(field, _muestras_col(ws, header_row, c)):
                chosen, chosen_valid = c, True
                break
        # monto es el ancla: si ninguna candidata valida, quedarse con la mejor
        # apuesta igual (sin marcarla validada → baja la confianza).
        if chosen is None and field == "monto" and cands:
            chosen = cands[0]
        if chosen is not None:
            columnas[field] = chosen
            usados.add(chosen)
            if chosen_valid:
                validados.add(field)

    conf = 0.0
    if "monto" in validados:
        conf += 0.5
    elif columnas["monto"]:
        conf += 0.2
    if "cuit" in validados:
        conf += 0.25
    if "titular" in validados:
        conf += 0.2
    if "fecha" in validados:
        conf += 0.05
    conf = min(round(conf, 4), 1.0)

    return columnas, validados, conf


# ── Capa 3: IA (Gemini) — degradación total, nunca lanza ──────────────────────

def _intentar_ia(ws, maxcol: int):
    """Pide a Gemini el índice de columna de cada campo canónico. Devuelve
    {"header_row": int, "columnas": {...}} o None si no hay key / falla."""
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        return None
    try:
        import google.generativeai as genai

        # Grilla compacta de las primeras filas para que el modelo vea la estructura.
        lineas = []
        for r in range(1, min(10, ws.max_row or 1) + 1):
            celdas = []
            for c in range(1, maxcol + 1):
                v = ws.cell(r, c).value
                if v is None or str(v).strip() == "":
                    continue
                celdas.append(f"col{c}={str(v).strip()[:40]}")
            if celdas:
                lineas.append(f"fila{r}: " + " | ".join(celdas))
        grilla = "\n".join(lineas)

        prompt = (
            "Sos un parser de planillas de pagos. La grilla siguiente son las "
            "primeras filas de una planilla (índices de columna 1-based). Devolvé "
            "SOLO un JSON con la fila de encabezados y el índice de columna de cada "
            "campo canónico (null si no existe). NO confundas columnas constantes "
            "(mismo valor en todas las filas, ej. nombre de cuenta) con el titular.\n"
            'Formato exacto: {"header_row": int, "columnas": {"monto": int|null, '
            '"cuit": int|null, "titular": int|null, "fecha": int|null, '
            '"referencia": int|null}}\n\n' + grilla
        )
        model = genai.GenerativeModel(os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"))
        resp = model.generate_content(prompt)
        texto = (resp.text or "").strip()
        if texto.startswith("```"):
            texto = "\n".join(l for l in texto.split("\n") if not l.startswith("```")).strip()
        data = json.loads(texto)
        if not isinstance(data, dict) or "columnas" not in data:
            return None
        return data
    except Exception:
        return None


# ── Parseo de filas + armado del resultado ────────────────────────────────────

def _coerce_columnas(raw: dict) -> dict:
    out = {f: None for f in FIELDS}
    for f in FIELDS:
        v = raw.get(f) if raw else None
        try:
            out[f] = int(v) if v not in (None, "", 0) else None
        except (TypeError, ValueError):
            out[f] = None
    return out


def _parsear_filas(ws, header_row: int, columnas: dict):
    """Itera datos desde header_row+1, corta tras N filas consecutivas con monto
    vacío. Devuelve (filas, descartadas)."""
    filas = []
    descartadas = 0
    vacias = 0
    cm = columnas.get("monto")
    mapped = [c for c in columnas.values() if c]
    limite = (ws.max_row or header_row) + 1

    r = header_row + 1
    scanned = 0
    while r < limite and scanned < MAX_DATA_ROWS:
        scanned += 1
        monto = _parse_monto(ws.cell(r, cm).value) if cm else None
        if monto is None:
            fila_vacia = all(
                (v := ws.cell(r, c).value) is None or str(v).strip() == "" for c in mapped
            )
            if not fila_vacia:
                descartadas += 1
            vacias += 1
            if vacias >= CORTE_FILAS_VACIAS:
                break
            r += 1
            continue
        vacias = 0
        filas.append({
            "monto": to_decimal(str(monto)),
            "cuit": _solo_digitos(ws.cell(r, columnas["cuit"]).value) if columnas.get("cuit") else None,
            "titular": _texto(ws.cell(r, columnas["titular"]).value) if columnas.get("titular") else None,
            "referencia": _texto(ws.cell(r, columnas["referencia"]).value) if columnas.get("referencia") else None,
            "fecha": _parse_fecha(ws.cell(r, columnas["fecha"]).value) if columnas.get("fecha") else None,
        })
        r += 1

    return filas, descartadas


def _columnas_disponibles(ws, header_row: int, maxcol: int) -> list:
    """Para el modal: todas las columnas con header, con 3 muestras de datos."""
    out = []
    for c in range(1, maxcol + 1):
        header = ws.cell(header_row, c).value
        if header is None or str(header).strip() == "":
            continue
        muestras = [str(v) for v in _muestras_col(ws, header_row, c, n=3)]
        out.append({"idx": c, "header": str(header).strip(), "muestras": muestras})
    return out


def _preview(ws, header_row: int, maxcol: int, n: int = 8) -> list:
    """Primeras ~n filas de datos crudas (para mostrar en el modal)."""
    out = []
    r = header_row + 1
    limite = (ws.max_row or header_row) + 1
    while r < limite and len(out) < n:
        valores = [
            ("" if (v := ws.cell(r, c).value) is None else str(v))
            for c in range(1, maxcol + 1)
        ]
        if any(x.strip() for x in valores):
            out.append({"fila": r, "valores": valores})
        r += 1
    return out


def _armar_resultado(ws, header_row, columnas, maxcol, origen, confianza, fingerprint) -> dict:
    filas, descartadas = _parsear_filas(ws, header_row, columnas)
    return {
        "origen": origen,
        "confianza": round(float(confianza), 4),
        "header_row": header_row,
        "columnas": {f: columnas.get(f) for f in FIELDS},
        "columnas_disponibles": _columnas_disponibles(ws, header_row, maxcol),
        "preview": _preview(ws, header_row, maxcol),
        "filas": filas,
        "filas_totales": len(filas),
        "filas_descartadas": descartadas,
        "fingerprint": fingerprint,
    }


# ── API pública ───────────────────────────────────────────────────────────────

def estandarizar_planilla(contenido: bytes, mapeo: dict | None = None,
                          usar_ia: bool = True) -> dict:
    """Estandariza una planilla al esquema canónico. Ver docstring del módulo."""
    wb, ws, cleanup = _cargar_ws(contenido)
    try:
        maxcol = _maxcol(ws)

        # ── Capa 1: perfil / manual ──
        # Aceptar dos formas de `mapeo`: anidada {header_row, columnas:{...}} (perfil
        # guardado) o plana {monto, cuit, titular, fecha, referencia, header_row}
        # (corrección manual del frontend). Se normaliza a la anidada.
        if mapeo and isinstance(mapeo, dict) and not mapeo.get("columnas"):
            _campos = ("monto", "cuit", "titular", "fecha", "referencia")
            if any(k in mapeo for k in _campos):
                mapeo = {
                    "header_row": mapeo.get("header_row"),
                    "fingerprint": mapeo.get("fingerprint"),
                    "columnas": {k: mapeo.get(k) for k in _campos},
                }
        if mapeo and isinstance(mapeo, dict) and mapeo.get("columnas"):
            hr = int(mapeo.get("header_row") or 1)
            fp_file = _fingerprint(ws, hr, maxcol)
            fp_stored = mapeo.get("fingerprint")
            # Si el perfil guardado no matchea este archivo, se descarta y se cae
            # a la heurística (el cliente puede haber cambiado de formato).
            if not fp_stored or fp_stored == fp_file:
                columnas = _coerce_columnas(mapeo["columnas"])
                origen = "perfil" if fp_stored else "manual"
                return _armar_resultado(ws, hr, columnas, maxcol, origen, 1.0, fp_file)

        # ── Capa 2: heurística ──
        hr = _detectar_header_row(ws, maxcol)
        columnas, _validados, conf = _asignar_columnas(ws, hr, maxcol)
        fp = _fingerprint(ws, hr, maxcol)

        # ── Capa 3: IA (solo si la heurística no alcanzó el umbral) ──
        if usar_ia and conf < UMBRAL_CONFIANZA:
            ia = _intentar_ia(ws, maxcol)
            if ia:
                hr2 = int(ia.get("header_row") or hr)
                cols2 = _coerce_columnas(ia.get("columnas") or {})
                if cols2.get("monto"):
                    return _armar_resultado(
                        ws, hr2, cols2, maxcol, "ia",
                        max(conf, 0.85), _fingerprint(ws, hr2, maxcol),
                    )

        return _armar_resultado(ws, hr, columnas, maxcol, "heuristica", conf, fp)
    finally:
        cleanup()
