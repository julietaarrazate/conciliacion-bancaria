from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, and_, or_, text, func
from datetime import date, datetime
from zoneinfo import ZoneInfo as _ZI
from slowapi import Limiter
from slowapi.util import get_remote_address
_ARG = _ZI('America/Argentina/Buenos_Aires')
from typing import Optional
import tempfile, os, io, hashlib

from app.database import get_db
from app.models.extracto import ExtractoBancario, MovimientoBanco
from app.models.planilla import Planilla
from app.models.user import User
from app.schemas.extracto import (
    ExtractoBancarioResponse, ExtractoListResponse,
    MergeUMResponse, MovimientosFiltradosResponse
)
from app.services.excel_parser import parsear_extracto_bancario
from app.services.extracto_merger import mergear_movimientos
from app.services.auditoria import registrar_log
from app.services.excel_export import export_movimientos, export_extracto_contador
import logging
from app.middleware.auth import get_current_user, require_permission, can_switch_org
from app.config import get_settings


def _extracto_for_user(db: Session, extracto_id: int, current_user: User,
                       include_deleted: bool = False) -> ExtractoBancario:
    """Resuelve un extracto con aislamiento multi-tenant.
    Superadmin ve cualquier org; el resto solo la propia. 404 si no existe o es de otra org."""
    q = db.query(ExtractoBancario).filter(ExtractoBancario.id == extracto_id)
    if not include_deleted:
        q = q.filter(ExtractoBancario.deleted_at.is_(None))
    if not current_user.is_superadmin:
        q = q.filter(ExtractoBancario.organizacion_id == current_user.organizacion_id)
    extracto = q.first()
    if not extracto:
        raise HTTPException(404, "Extracto no encontrado")
    return extracto

settings = get_settings()
logger = logging.getLogger(__name__)
router = APIRouter(prefix="/extractos", tags=["extractos"])
limiter = Limiter(key_func=get_remote_address)

# Mapeo de clave interna del parser → nombre de banco legible
_BANCO_NOMBRE = {
    "macro":     "Banco Macro",
    "bbva":      "BBVA",
    "santander": "Santander",
    "galicia":   "Galicia",
    "icbc":      "ICBC",
    "nacion":    "Banco Nación",
    "provincia": "Bapro",
    "ciudad":    "Banco Ciudad",
    "hsbc":      "HSBC",
}

def _resolver_banco(banco_param: str, banco_detectado: str) -> str:
    """
    Usa el banco detectado por el parser cuando es más específico que el default.
    Si el usuario pasó un banco explícito distinto del default, se respeta siempre.
    """
    if banco_detectado != "generico":
        nombre_detectado = _BANCO_NOMBRE.get(banco_detectado, banco_detectado.capitalize())
        # Usar detectado si el parámetro es el default ("Banco Macro") o coincide
        if banco_param == "Banco Macro" or banco_param == nombre_detectado:
            return nombre_detectado
    return banco_param


def _fingerprint(movimientos: list) -> str:
    """Huella digital del extracto: hash de total+ordenes+suma_montos"""
    total = len(movimientos)
    if total == 0:
        return "empty"
    primer_orden = movimientos[0].get("orden") or 0
    ultimo_orden = movimientos[-1].get("orden") or 0
    suma = round(sum(m.get("monto", 0) for m in movimientos), 2)
    raw = f"{total}|{primer_orden}|{ultimo_orden}|{suma}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


@router.get("", response_model=ExtractoListResponse)
def list_extractos(skip: int = 0, limit: int = 50,
                   org_id: Optional[int] = Query(None),
                   db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    q = db.query(ExtractoBancario).filter(ExtractoBancario.deleted_at.is_(None))
    # Superadmin puede filtrar por org; usuarios normales ven solo su org
    if can_switch_org(current_user, org_id) and org_id:
        q = q.filter(ExtractoBancario.organizacion_id == org_id)
    elif not current_user.is_superadmin:
        q = q.filter(ExtractoBancario.organizacion_id == (current_user.organizacion_id or 1))
    total = q.count()
    rows = q.order_by(desc(ExtractoBancario.fecha_creacion)).offset(skip).limit(limit).all()
    items = [{"id": e.id, "nombre_archivo": e.nombre_archivo,
              "fecha_creacion": e.fecha_creacion, "total_movimientos": len(e.movimientos),
              "banco": e.banco or "Banco Macro"} for e in rows]
    return {"total": total, "items": items}


@router.post("/upload", response_model=ExtractoBancarioResponse)
@limiter.limit("10/minute")
async def upload_extracto(request: Request,
                          file: UploadFile = File(...),
                          banco: str = Query("Banco Macro"),
                          org_id: Optional[int] = Query(None),
                          db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    # Org destino: la activa seleccionada por el usuario (si tiene permiso para
    # operar ahí), si no la propia. Antes esto se ignoraba y todo se guardaba
    # con el default del modelo (organizacion_id=1) sin importar qué org tenía
    # seleccionada el superadmin — además de archivar mal el extracto, un
    # segundo intento podía pisar el unique index (fingerprint, organizacion_id)
    # de un upload previo y tirar un error genérico de "procesar el archivo".
    org_destino = org_id if (org_id and can_switch_org(current_user, org_id)) else (current_user.organizacion_id or 1)
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        raise HTTPException(400, "Formatos aceptados: .xlsx, .xls, .csv")
    tmp_path = None
    try:
        content = await file.read()
        if len(content) > settings.max_file_size:
            raise HTTPException(413, f"El archivo supera el límite de {settings.max_file_size // (1024 * 1024)} MB.")
        # Validación de magic bytes: evita que archivos renombrados pasen la validación de extensión
        _XLSX_MAGIC = b'PK\x03\x04'       # ZIP (XLSX internamente es ZIP)
        _XLS_MAGIC  = b'\xd0\xcf\x11\xe0'  # OLE2 (formato binario .xls)
        header = content[:8]
        if ext == '.xlsx' and not header.startswith(_XLSX_MAGIC):
            raise HTTPException(400, "El archivo no es un Excel válido (.xlsx debe ser un archivo ZIP/OOXML)")
        if ext == '.xls' and not header.startswith(_XLS_MAGIC):
            raise HTTPException(400, "El archivo no es un Excel válido (.xls debe ser formato OLE2)")
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        parsed = parsear_extracto_bancario(tmp_path)
        movs = parsed["movimientos"]
        if len(movs) == 0:
            raise HTTPException(400,
                "El archivo no contiene movimientos reconocibles. "
                "Verificá que sea un extracto bancario en formato .xlsx (Macro, BBVA, Santander, Galicia)."
            )
        suma_abs = sum(abs(float(m.get("monto") or 0)) for m in movs)
        if suma_abs == 0:
            raise HTTPException(400,
                "El archivo se procesó pero todos los montos son cero. "
                "El formato puede no coincidir con el esperado — intentá exportarlo nuevamente desde el banco."
            )
        banco_detectado = parsed.get("banco_detectado", "generico")
        banco_final = _resolver_banco(banco, banco_detectado)
        fp = _fingerprint(movs)

        # Si ya existe un extracto identico (fingerprint match) → upsert de acreditados:
        # actualizar cliente_acreditado y fecha_acred en movs existentes cuando el archivo
        # nuevo trae info que el guardado no tiene. NO tirar 409: el usuario espera que
        # subir el extracto "actualice" lo que ya esta.
        # Solo matchea extractos ACTIVOS: re-subir el mismo archivo ACTUALIZA los
        # acreditados del extracto vigente. Si el archivo fue borrado, el dedupe NO
        # lo encuentra → se crea uno nuevo y limpio (numerado desde 1), en vez de
        # resucitar la fila vieja con su numeración heredada.
        existente = db.query(ExtractoBancario).filter(
            ExtractoBancario.fingerprint == fp,
            ExtractoBancario.organizacion_id == org_destino,
            ExtractoBancario.deleted_at.is_(None),
        ).first()
        if existente:
            actualizados = 0
            for m in existente.movimientos:
                cliente_nuevo = None
                fecha_nueva = None
                # buscar el correspondiente en el nuevo archivo por (orden) o (saldo, monto).
                # m.saldo/m.monto son Decimal (Numeric en la DB) y n[...] viene float del
                # parser → comparar en float evita "unsupported operand Decimal - float"
                # (bug recurrente Decimal/float, ver BUGS.md).
                for n in movs:
                    if (m.orden is not None and n.get("orden") == m.orden) or (
                        m.saldo is not None and n.get("saldo") is not None
                        and abs(float(m.saldo or 0) - float(n["saldo"])) < 0.01
                        and abs(float(m.monto or 0) - float(n.get("monto") or 0)) < 0.01
                    ):
                        cliente_nuevo = n.get("cliente_acreditado")
                        fecha_nueva = n.get("fecha_acred")
                        break
                if cliente_nuevo and not m.cliente_acreditado:
                    m.cliente_acreditado = cliente_nuevo
                    actualizados += 1
                if fecha_nueva and not m.fecha_acred:
                    m.fecha_acred = fecha_nueva
            # Actualizar banco si el detectado es más específico que el guardado
            if banco_final and banco_final != "Banco Macro" and not existente.banco:
                existente.banco = banco_final
            db.commit()
            db.refresh(existente)
            registrar_log(db, current_user.id, "extractos_bancarios", existente.id, "UPSERT_ACRED",
                          {"archivo": file.filename, "actualizados": actualizados})
            return existente

        extracto = ExtractoBancario(nombre_archivo=file.filename, creado_por=current_user.id, fingerprint=fp,
                                     banco=banco_final, organizacion_id=org_destino)
        db.add(extracto)
        db.flush()
        # Orden: contador secuencial POR EXTRACTO — NO usar valores del banco.
        # Cada extracto arranca su propia numeración desde 1 (un extracto recién
        # creado no tiene movimientos → max 0 → el más antiguo recibe 1, el más
        # reciente N). Mismo criterio que el append de UM (extracto_merger), y evita
        # que un extracto de otra empresa/banco continúe la numeración de otro.
        from sqlalchemy import func as _func
        max_global = (
            db.query(_func.max(MovimientoBanco.orden))
            .filter(MovimientoBanco.extracto_id == extracto.id)
            .scalar() or 0
        )
        n_movs = len(movs)
        for i, m in enumerate(movs):
            # i=0 es el más reciente → orden más alto
            orden_asignado = max_global + (n_movs - i)
            db.add(MovimientoBanco(
                extracto_id=extracto.id, orden=orden_asignado,
                fecha=m.get("fecha"), mes=m.get("mes"),
                titular=m.get("titular"), monto=m.get("monto"), saldo=m.get("saldo"),
                cliente_acreditado=m.get("cliente_acreditado"),
                fecha_acred=m.get("fecha_acred"),
                organizacion_id=org_destino,
            ))
        db.commit()
        db.refresh(extracto)
        registrar_log(db, current_user.id, "extractos_bancarios", extracto.id, "INSERT",
                      {"nombre_archivo": file.filename, "movimientos": len(movs)})
        # El extracto NO genera asiento propio: el UM importa los movimientos
        # con cuentas correctas (um_lote: 1-1-1-3-1 / 2-1-1-1).
        # registrar_extracto() usaba cuentas madre y duplicaba con um_lote.
        return extracto
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error("upload error: %s", e)
        raise HTTPException(400, "Error al procesar el archivo. Verificá el formato y volvé a intentar.")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


@router.patch("/{extracto_id}/renombrar")
def renombrar_extracto(extracto_id: int, payload: dict,
                       db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    """Cambia el nombre visible del extracto."""
    extracto = _extracto_for_user(db, extracto_id, current_user)
    nuevo_nombre = (payload.get("nombre") or "").strip()
    if not nuevo_nombre:
        raise HTTPException(400, "El nombre no puede estar vacío")
    antes = extracto.nombre_archivo
    extracto.nombre_archivo = nuevo_nombre
    db.commit()
    registrar_log(db, current_user.id, "extractos_bancarios", extracto_id,
                  "UPDATE", {"antes": antes, "despues": nuevo_nombre})
    return {"ok": True, "nombre_archivo": extracto.nombre_archivo}


@router.delete("/{extracto_id}")
def delete_extracto(extracto_id: int, db: Session = Depends(get_db),
                    current_user: User = Depends(require_permission("delete_records"))):
    """Soft delete: marca el extracto como eliminado (deleted_at = now()).
    Los movimientos y la relación con planillas se conservan, así si se restaura
    el extracto vuelve completo. Para borrar definitivo: usar /admin/papelera/purgar."""
    extracto = _extracto_for_user(db, extracto_id, current_user)

    try:
        extracto.deleted_at = datetime.now(_ARG).replace(tzinfo=None)
        db.commit()
        registrar_log(db, current_user.id, "extractos_bancarios", extracto_id,
                      "SOFT_DELETE", {"nombre": extracto.nombre_archivo})
        return {
            "ok": True,
            "mensaje": f"Extracto #{extracto_id} enviado a papelera. Restaurable desde /admin/papelera.",
            "soft_deleted": True,
        }
    except Exception as e:
        db.rollback()
        logger.error("soft delete extracto: %s", e)
        raise HTTPException(500, "Error al eliminar el extracto. Intentá de nuevo.")


@router.delete("")
def delete_todos_extractos(db: Session = Depends(get_db),
                           current_user: User = Depends(get_current_user)):
    """Elimina TODOS los extractos, planillas y movimientos DE LA ORG ACTUAL.
    Solo superadmin. Aislado por organizacion_id — no toca otras orgs."""
    if not current_user.is_superadmin:
        raise HTTPException(403, "Solo superadmin puede ejecutar limpieza masiva")

    from app.models.planilla import PlanillaRow, Planilla
    oid = current_user.organizacion_id

    try:
        # FKs: PlanillaRow → Planilla → ExtractoBancario / MovimientoBanco → ExtractoBancario
        planilla_ids = [p.id for p in db.query(Planilla.id).filter(Planilla.organizacion_id == oid).all()]
        extracto_ids = [e.id for e in db.query(ExtractoBancario.id).filter(ExtractoBancario.organizacion_id == oid).all()]

        if planilla_ids:
            db.query(PlanillaRow).filter(PlanillaRow.planilla_id.in_(planilla_ids)).update(
                {"orden_movimiento_acreditado": None}, synchronize_session=False
            )
        db.flush()
        n_planillas = db.query(Planilla).filter(Planilla.organizacion_id == oid).delete(synchronize_session=False)
        db.flush()
        n_movs = 0
        if extracto_ids:
            n_movs = db.query(MovimientoBanco).filter(MovimientoBanco.extracto_id.in_(extracto_ids)).delete(synchronize_session=False)
        db.flush()
        n_extractos = db.query(ExtractoBancario).filter(ExtractoBancario.organizacion_id == oid).delete(synchronize_session=False)
        db.commit()

        registrar_log(db, current_user.id, "extractos_bancarios", 0, "DELETE_ALL",
                      {"extractos": n_extractos, "movimientos": n_movs, "planillas": n_planillas, "org_id": oid})
        return {"ok": True, "mensaje": f"Limpieza completa de la org: {n_extractos} extractos, {n_planillas} planillas, {n_movs} movimientos"}
    except Exception as e:
        db.rollback()
        logger.error("limpiar extractos: %s", e)
        raise HTTPException(500, "Error al limpiar. Intentá de nuevo.")


@router.delete("/{extracto_id}/movimientos-um")
def eliminar_movimientos_um(extracto_id: int,
                            forzar_todo: bool = Query(False),
                            db: Session = Depends(get_db),
                            current_user: User = Depends(require_permission("delete_records"))):
    """Elimina solo el ultimo lote de UM. Si no hay lotes asignados requiere forzar_todo=true."""
    extracto = _extracto_for_user(db, extracto_id, current_user)
    try:
        from sqlalchemy import func
        from app.models.planilla import PlanillaRow

        max_lote = db.query(func.max(MovimientoBanco.um_lote)).filter(
            MovimientoBanco.extracto_id == extracto_id,
            MovimientoBanco.source == 'um',
            MovimientoBanco.um_lote.isnot(None),
        ).scalar()

        if max_lote is None:
            if not forzar_todo:
                total_um = db.query(func.count(MovimientoBanco.id)).filter(
                    MovimientoBanco.extracto_id == extracto_id,
                    MovimientoBanco.source == 'um'
                ).scalar() or 0
                raise HTTPException(409, {
                    "code": "sin_lotes",
                    "total": total_um,
                    "mensaje": f"Este extracto tiene {total_um} movimientos UM cargados antes del sistema de lotes. No se puede identificar solo el último. ¿Borrar todos?"
                })
            movs_query = db.query(MovimientoBanco).filter(
                MovimientoBanco.extracto_id == extracto_id,
                MovimientoBanco.source == 'um'
            )
        else:
            movs_query = db.query(MovimientoBanco).filter(
                MovimientoBanco.extracto_id == extracto_id,
                MovimientoBanco.um_lote == max_lote,
            )

        movs_a_borrar = movs_query.all()
        ids_a_borrar = [m.id for m in movs_a_borrar]

        # Revertir asientos contables antes de eliminar los movimientos
        # (fault-tolerant: un fallo aquí no debe bloquear la baja del lote UM,
        # pero se reporta para que no quede silencioso)
        contabilidad_ok = True
        try:
            from app.services.motor_contable import reversar_asientos as _rev
            _org_id = extracto.organizacion_id or 1
            for _mv in movs_a_borrar:
                _rev(db, "um_mov", _mv.id, _org_id, current_user.id, "Baja lote UM")
            if movs_a_borrar:
                _rev(db, "um_lote", movs_a_borrar[0].id, _org_id, current_user.id, "Baja lote UM")
        except Exception as _mc_ex:
            logger.error("reverso UM asientos falló (extracto %s, lote %s): %s", extracto_id, max_lote, _mc_ex)
            contabilidad_ok = False

        if ids_a_borrar:
            db.query(PlanillaRow).filter(
                PlanillaRow.orden_movimiento_acreditado.in_(ids_a_borrar)
            ).update({"orden_movimiento_acreditado": None}, synchronize_session="fetch")
            db.flush()

        n = movs_query.delete(synchronize_session="fetch")
        db.commit()
        registrar_log(db, current_user.id, "extractos_bancarios", extracto_id, "DELETE_UM",
                      {"eliminados": n, "lote": max_lote})
        return {"ok": True, "eliminados": n, "lote": max_lote, "contabilidad_ok": contabilidad_ok}
    except Exception as e:
        db.rollback()
        logger.error("eliminar UM: %s", e)
        raise HTTPException(500, "Error al eliminar movimientos UM. Intentá de nuevo.")


@router.post("/{extracto_id}/agregar-um", response_model=MergeUMResponse)
@limiter.limit("20/minute")
async def agregar_ultimos_movimientos(
    request: Request,
    extracto_id: int,
    file: UploadFile = File(...),
    corte_saldo: Optional[float] = Form(None),
    modo_asiento: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    extracto = _extracto_for_user(db, extracto_id, current_user)
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        raise HTTPException(400, "Formatos aceptados: .xlsx, .xls, .csv")
    tmp_path = None
    try:
        content = await file.read()
        if len(content) > settings.max_file_size:
            raise HTTPException(413, f"El archivo supera el límite de {settings.max_file_size // (1024 * 1024)} MB.")
        # Validación de magic bytes
        _XLSX_MAGIC = b'PK\x03\x04'
        _XLS_MAGIC  = b'\xd0\xcf\x11\xe0'
        header = content[:8]
        if ext == '.xlsx' and not header.startswith(_XLSX_MAGIC):
            raise HTTPException(400, "El archivo no es un Excel válido (.xlsx debe ser un archivo ZIP/OOXML)")
        if ext == '.xls' and not header.startswith(_XLS_MAGIC):
            raise HTTPException(400, "El archivo no es un Excel válido (.xls debe ser formato OLE2)")
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        parsed = parsear_extracto_bancario(tmp_path)
        stats = mergear_movimientos(db, extracto_id, parsed["movimientos"], corte_saldo=corte_saldo)
        registrar_log(db, current_user.id, "extractos_bancarios", extracto_id, "APPEND_UM",
                      {"archivo": file.filename, **stats})

        # Asiento contable automático para los movimientos UM nuevos
        # (fault-tolerant: la importación de movimientos ya se guardó arriba,
        # un fallo aquí no debe bloquearla, pero se reporta para que no quede silencioso)
        contabilidad_ok = True
        if stats["agregados"] > 0 and stats.get("nuevo_lote"):
            try:
                from app.services.motor_contable import registrar_um_import
                from app.models.organizacion import Organizacion as _Org
                _org_id = extracto.organizacion_id or 1
                movs_nuevos = db.query(MovimientoBanco).filter(
                    MovimientoBanco.extracto_id == extracto_id,
                    MovimientoBanco.um_lote == stats["nuevo_lote"],
                ).all()
                _org = db.query(_Org).filter(_Org.id == _org_id).first()
                _modo_config = ((_org.configuracion or {}).get("modo_asiento_um", "agrupado")) if _org else "agrupado"
                _modo = modo_asiento if modo_asiento in ("individual", "agrupado") else _modo_config
                registrar_um_import(
                    db=db,
                    extracto_id=extracto_id,
                    org_id=_org_id,
                    usuario_id=current_user.id,
                    movimientos_nuevos=movs_nuevos,
                    modo=_modo,
                )
            except Exception as _mc_ex:
                logger.error("motor_contable UM falló (extracto %s, lote %s): %s", extracto_id, stats.get("nuevo_lote"), _mc_ex)
                contabilidad_ok = False

        return {"extracto_id": extracto_id, **stats, "contabilidad_ok": contabilidad_ok}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error("procesar UM: %s", e)
        raise HTTPException(400, "Error al procesar el archivo UM. Verificá el formato y volvé a intentar.")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


def _build_mov_query(db, extracto_id, cliente, cuit, titular, desde, hasta, fecha_desde, fecha_hasta, sin_acreditar):
    q = db.query(MovimientoBanco).filter(MovimientoBanco.extracto_id == extracto_id)
    if cliente: q = q.filter(MovimientoBanco.cliente_acreditado.ilike(f"%{cliente}%"))
    if cuit: q = q.filter(MovimientoBanco.titular.ilike(f"%{cuit.replace('-','').replace(' ','')}%"))
    if titular: q = q.filter(MovimientoBanco.titular.ilike(f"%{titular}%"))
    if desde: q = q.filter(MovimientoBanco.fecha_acred >= desde)
    if hasta: q = q.filter(MovimientoBanco.fecha_acred <= hasta)
    if fecha_desde: q = q.filter(MovimientoBanco.fecha >= fecha_desde)
    if fecha_hasta: q = q.filter(MovimientoBanco.fecha <= fecha_hasta)
    if sin_acreditar is True:
        q = q.filter(or_(MovimientoBanco.cliente_acreditado.is_(None),
                         MovimientoBanco.cliente_acreditado == "",
                         MovimientoBanco.cliente_acreditado.ilike("no identificado")))
    elif sin_acreditar is False:
        q = q.filter(and_(MovimientoBanco.cliente_acreditado.isnot(None),
                          MovimientoBanco.cliente_acreditado != "",
                          ~MovimientoBanco.cliente_acreditado.ilike("no identificado")))
    return q


@router.get("/{extracto_id}/movimientos", response_model=MovimientosFiltradosResponse)
def listar_movimientos(extracto_id: int,
                       cliente: Optional[str] = Query(None), cuit: Optional[str] = Query(None),
                       titular: Optional[str] = Query(None), desde: Optional[date] = Query(None),
                       hasta: Optional[date] = Query(None), fecha_desde: Optional[date] = Query(None),
                       fecha_hasta: Optional[date] = Query(None), sin_acreditar: Optional[bool] = Query(None),
                       skip: int = 0, limit: int = 0,
                       db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    _extracto_for_user(db, extracto_id, current_user, include_deleted=True)
    q = _build_mov_query(db, extracto_id, cliente, cuit, titular, desde, hasta, fecha_desde, fecha_hasta, sin_acreditar)
    total = q.count()
    q = q.order_by(MovimientoBanco.orden.desc().nulls_last(), MovimientoBanco.id.desc()).offset(skip)
    if limit > 0:
        q = q.limit(limit)
    items = q.all()
    return {"extracto_id": extracto_id, "total": total, "items": items}


@router.get("/{extracto_id}/movimientos/export")
def export_movimientos_xlsx(extracto_id: int,
                            cliente: Optional[str] = Query(None), cuit: Optional[str] = Query(None),
                            titular: Optional[str] = Query(None), desde: Optional[date] = Query(None),
                            hasta: Optional[date] = Query(None), fecha_desde: Optional[date] = Query(None),
                            fecha_hasta: Optional[date] = Query(None), sin_acreditar: Optional[bool] = Query(None),
                            db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    extracto = _extracto_for_user(db, extracto_id, current_user, include_deleted=True)
    q = _build_mov_query(db, extracto_id, cliente, cuit, titular, desde, hasta, fecha_desde, fecha_hasta, sin_acreditar)
    rows = q.order_by(MovimientoBanco.orden.desc().nulls_last(), MovimientoBanco.id.desc()).all()
    movs = [{"orden": m.orden, "fecha": m.fecha, "mes": m.mes, "titular": m.titular,
              "monto": m.monto, "saldo": m.saldo, "cliente_acreditado": m.cliente_acreditado,
              "fecha_acred": m.fecha_acred} for m in rows]
    data = export_movimientos(extracto.nombre_archivo, movs)
    filename = f"movimientos_{datetime.now(_ARG).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{extracto_id}/export-contador")
def export_para_contador(
    extracto_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Descarga el extracto completo conciliado en Excel profesional para el contador.
    Hoja 1: todos los movimientos con acreditaciones coloreadas en verde.
    Hoja 2: resumen estadístico y detalle por cliente.
    """
    extracto = _extracto_for_user(db, extracto_id, current_user, include_deleted=True)

    movs = sorted(extracto.movimientos, key=lambda m: (m.fecha.isoformat() if m.fecha else "", m.orden or 0))
    data = [
        {
            "orden": m.orden,
            "fecha": m.fecha,
            "mes": m.mes,
            "titular": m.titular,
            "monto": m.monto,
            "saldo": m.saldo,
            "cliente_acreditado": m.cliente_acreditado,
            "fecha_acred": m.fecha_acred,
        }
        for m in movs
    ]

    xlsx = export_extracto_contador(extracto.nombre_archivo, data)
    fecha_str = datetime.now(_ARG).strftime('%Y%m%d')
    nombre_base = extracto.nombre_archivo.replace('.xlsx', '').replace('.XLSX', '').replace('.xls', '').replace('.XLS', '').strip()
    if not nombre_base:
        nombre_base = f"extracto_{extracto_id}"
    filename = f"{nombre_base}_conciliado_{fecha_str}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.patch("/{extracto_id}/movimientos/{mov_id}")
def update_movimiento(
    extracto_id: int,
    mov_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Edita un movimiento bancario (titular, monto, fecha, etc.) — para correcciones manuales."""
    extracto = _extracto_for_user(db, extracto_id, current_user, include_deleted=True)
    mov = db.query(MovimientoBanco).filter(
        MovimientoBanco.id == mov_id,
        MovimientoBanco.extracto_id == extracto_id
    ).first()
    if not mov:
        raise HTTPException(404, "Movimiento no encontrado")
    # Inmutabilidad de periodo cerrado
    from app.services.cierre_periodo import periodo_esta_cerrado
    fecha_check = mov.fecha_acred or mov.fecha
    if periodo_esta_cerrado(db, extracto.organizacion_id or 1, fecha_check):
        raise HTTPException(409, "El periodo ya está cerrado — este movimiento no se puede modificar")

    campos_editables = {"titular", "monto", "fecha", "mes", "saldo", "orden",
                        "cliente_acreditado", "fecha_acred"}
    acred_cambiado = False
    for campo, valor in payload.items():
        if campo not in campos_editables:
            continue
        if campo in ("fecha", "fecha_acred") and isinstance(valor, str) and valor:
            from datetime import date as date_type
            try:
                valor = date_type.fromisoformat(valor)
            except ValueError:
                continue
        if campo in ("fecha", "fecha_acred") and valor == "":
            valor = None
        if campo in ("cliente_acreditado", "fecha_acred"):
            acred_cambiado = True
        setattr(mov, campo, valor)

    db.commit()

    # Sync planilla rows — bidireccional
    if acred_cambiado:
        from app.models.planilla import PlanillaRow, Planilla
        from app.models.cliente import Cliente

        # 1. Actualizar filas ya vinculadas directamente a este movimiento
        linked = db.query(PlanillaRow).filter(
            PlanillaRow.orden_movimiento_acreditado == mov_id
        ).all()
        for row in linked:
            if mov.cliente_acreditado:
                row.status = "ok"
                if mov.fecha_acred:
                    row.fecha_acred = mov.fecha_acred
            elif row.status == "ok":
                row.status = "pendiente"
                row.fecha_acred = None
                row.orden_movimiento_acreditado = None

        # 2. Si se está acreditando a un cliente: buscar fila sin vincular
        #    en sus planillas de este extracto con el mismo monto
        if mov.cliente_acreditado and mov.monto is not None:
            cliente = db.query(Cliente).filter(
                func.lower(Cliente.nombre) == mov.cliente_acreditado.lower()
            ).first()
            if cliente:
                planillas = db.query(Planilla).filter(
                    Planilla.cliente_id == cliente.id,
                    Planilla.extracto_id == mov.extracto_id,
                    Planilla.deleted_at.is_(None),
                ).all()
                mov_monto = abs(mov.monto or 0)
                for planilla in planillas:
                    for row in planilla.rows:
                        if row.status == "ok" or row.orden_movimiento_acreditado is not None:
                            continue
                        try:
                            row_monto = abs(row.monto or 0)
                        except (ValueError, TypeError):
                            continue
                        if abs(row_monto - mov_monto) < 0.01:
                            row.status = "ok"
                            row.orden_movimiento_acreditado = mov_id
                            if mov.fecha_acred:
                                row.fecha_acred = mov.fecha_acred
                            break  # una fila por movimiento

        db.commit()

    db.refresh(mov)
    registrar_log(db, current_user.id, "movimientos_banco", mov_id, "UPDATE",
                  {"campos": list(payload.keys())})
    return {"ok": True, "id": mov_id}


@router.delete("/{extracto_id}/movimientos/{mov_id}")
def delete_movimiento(
    extracto_id: int,
    mov_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("delete_records"))
):
    from app.models.planilla import PlanillaRow
    extracto = _extracto_for_user(db, extracto_id, current_user, include_deleted=True)
    mov = db.query(MovimientoBanco).filter(
        MovimientoBanco.id == mov_id,
        MovimientoBanco.extracto_id == extracto_id
    ).first()
    if not mov:
        raise HTTPException(404, "Movimiento no encontrado")
    # Inmutabilidad de periodo cerrado
    from app.services.cierre_periodo import periodo_esta_cerrado
    fecha_check = mov.fecha_acred or mov.fecha
    if periodo_esta_cerrado(db, extracto.organizacion_id or 1, fecha_check):
        raise HTTPException(409, "El periodo ya está cerrado — este movimiento no se puede borrar")
    try:
        db.query(PlanillaRow).filter(
            PlanillaRow.orden_movimiento_acreditado == mov_id
        ).update({"orden_movimiento_acreditado": None}, synchronize_session="fetch")
        orden_borrado = mov.orden
        db.flush()
        db.delete(mov)
        db.flush()
        # Cerrar el hueco: todos los movimientos con orden mayor al borrado
        # dentro del mismo extracto decrementan en 1.
        if orden_borrado is not None:
            db.query(MovimientoBanco).filter(
                MovimientoBanco.extracto_id == extracto_id,
                MovimientoBanco.orden > orden_borrado,
            ).update(
                {MovimientoBanco.orden: MovimientoBanco.orden - 1},
                synchronize_session=False,
            )
        db.commit()
        registrar_log(db, current_user.id, "movimientos_banco", mov_id, "DELETE", {})
        return {"ok": True}
    except Exception as e:
        db.rollback()
        logger.error("borrar movimiento: %s", e)
        raise HTTPException(500, "Error al borrar el movimiento. Intentá de nuevo.")


@router.get("/{extracto_id}", response_model=ExtractoBancarioResponse)
def get_extracto(extracto_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    return _extracto_for_user(db, extracto_id, current_user, include_deleted=True)


# ─── Conciliaciones globales (cross-extracto) ───────────────────────────
# IMPORTANTE: estas rutas van con un prefijo distinto para que no choquen
# con /extractos/{id}. Se montan abajo como APIRouter separado.
from fastapi import APIRouter as _AR

conciliaciones_router = _AR(prefix="/conciliaciones", tags=["conciliaciones"])


def _conc_query(db, current_user, cliente, desde, hasta, monto_min, monto_max, titular, org_id=None):
    q = db.query(MovimientoBanco).filter(
        MovimientoBanco.cliente_acreditado.isnot(None),
        MovimientoBanco.cliente_acreditado != "",
        ~MovimientoBanco.cliente_acreditado.ilike("no identificado"),
    )
    # Scope por organizacion
    if can_switch_org(current_user, org_id) and org_id:
        q = q.filter(MovimientoBanco.organizacion_id == org_id)
    elif not current_user.is_superadmin:
        q = q.filter(MovimientoBanco.organizacion_id == (current_user.organizacion_id or 1))
    if cliente:
        q = q.filter(MovimientoBanco.cliente_acreditado.ilike(f"%{cliente}%"))
    if titular:
        q = q.filter(MovimientoBanco.titular.ilike(f"%{titular}%"))
    if desde:
        q = q.filter(MovimientoBanco.fecha_acred >= desde)
    if hasta:
        q = q.filter(MovimientoBanco.fecha_acred <= hasta)
    if monto_min is not None:
        q = q.filter(MovimientoBanco.monto >= monto_min)
    if monto_max is not None:
        q = q.filter(MovimientoBanco.monto <= monto_max)
    return q


@conciliaciones_router.get("")
def listar_conciliaciones(
    cliente: Optional[str] = Query(None),
    titular: Optional[str] = Query(None),
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    monto_min: Optional[float] = Query(None),
    monto_max: Optional[float] = Query(None),
    org_id: Optional[int] = Query(None),
    skip: int = 0, limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Lista todas las acreditaciones (movimientos con cliente_acreditado) cruzando
    TODOS los extractos del usuario. Permite filtrar por cliente, titular, rango
    de fecha de acreditacion y rango de monto.
    """
    q = _conc_query(db, current_user, cliente, desde, hasta, monto_min, monto_max, titular, org_id)
    total = q.count()
    rows = (
        q.order_by(MovimientoBanco.fecha_acred.desc().nulls_last(),
                   MovimientoBanco.fecha.desc().nulls_last(),
                   MovimientoBanco.id.desc())
        .offset(skip).limit(limit if limit > 0 else 50).all()
    )
    items = [{
        "id": m.id,
        "extracto_id": m.extracto_id,
        "orden": m.orden,
        "fecha": m.fecha,
        "titular": m.titular,
        "monto": m.monto,
        "saldo": m.saldo,
        "cliente_acreditado": m.cliente_acreditado,
        "fecha_acred": m.fecha_acred,
    } for m in rows]
    suma = sum(m.monto or 0 for m in rows)
    return {"total": total, "items": items, "suma": round(suma, 2)}


@conciliaciones_router.get("/export")
def export_conciliaciones(
    cliente: Optional[str] = Query(None),
    titular: Optional[str] = Query(None),
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    monto_min: Optional[float] = Query(None),
    monto_max: Optional[float] = Query(None),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = _conc_query(db, current_user, cliente, desde, hasta, monto_min, monto_max, titular, org_id)
    rows = q.order_by(MovimientoBanco.fecha_acred.desc().nulls_last(),
                      MovimientoBanco.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliaciones"

    titulo = "Conciliaciones"
    detalle = []
    if cliente: detalle.append(f"cliente={cliente}")
    if desde: detalle.append(f"desde={desde}")
    if hasta: detalle.append(f"hasta={hasta}")
    if titular: detalle.append(f"titular={titular}")
    ws.cell(row=1, column=1, value=titulo + (" — " + " · ".join(detalle) if detalle else ""))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    headers = ["Fecha mov.", "Orden", "Titular", "Importe", "Saldo", "Cliente", "Fecha acred.", "Extracto #"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")

    total_monto = 0
    for i, m in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=m.fecha).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=2, value=m.orden)
        ws.cell(row=i, column=3, value=m.titular)
        ws.cell(row=i, column=4, value=m.monto).number_format = "#,##0.00"
        ws.cell(row=i, column=5, value=m.saldo).number_format = "#,##0.00"
        ws.cell(row=i, column=6, value=m.cliente_acreditado)
        ws.cell(row=i, column=7, value=m.fecha_acred).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=8, value=m.extracto_id)
        # Alto 15 estandar Excel (= ~25px) y sin wrap_text para que el titular no agrande la fila
        for col in range(1, 9):
            c = ws.cell(row=i, column=col)
            c.alignment = Alignment(horizontal=c.alignment.horizontal or "general",
                                    vertical="center", wrap_text=False)
        ws.row_dimensions[i].height = 15
        total_monto += m.monto or 0

    n = len(rows)
    foot = 4 + n
    ws.cell(row=foot, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=foot, column=4, value=round(total_monto, 2)).font = Font(bold=True)
    ws.cell(row=foot, column=4).number_format = "#,##0.00"

    for col_letter, width in [("A",12),("B",8),("C",40),("D",14),("E",14),("F",18),("G",12),("H",12)]:
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_str = datetime.now(_ARG).strftime('%Y%m%d')
    filename = f"conciliaciones_{fecha_str}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})
