from datetime import date, datetime
from decimal import Decimal
from typing import Optional, List
import base64, io
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.middleware.auth import get_current_user, require_permission, can_switch_org
from app.models.user import User
from app.models.cheque import Cheque
from app.models.portador import Portador
from app.models.cliente import Cliente
from app.models.contabilidad import PlanCuenta
from app.services.motor_contable import registrar_cheque, acreditar_cheque, rechazar_cheque
from app.services.auditoria import registrar_log
from app.services.storage import upload_comprobante
from app.services.tz import hoy_art

router = APIRouter(prefix="/cheques", tags=["cheques"])


# ── Schemas ───────────────────────────────────────────────────────

class ChequeIn(BaseModel):
    cliente_id:          Optional[int]   = None
    portador_id:         Optional[int]   = None
    numero:              Optional[str]   = None
    banco_origen:        Optional[str]   = None
    librador:            Optional[str]   = None
    monto:               float           = Field(..., gt=0)
    comision:            float           = Field(0.0, ge=0)
    porcentaje_comision: Optional[float] = None
    codigo_postal:       Optional[str]   = None
    local_interior:      Optional[str]   = None
    fecha_emision:       Optional[date]  = None
    fecha_deposito:      Optional[date]  = None
    notas:               Optional[str]   = None


class AcreditarIn(BaseModel):
    fecha_acred:    Optional[date] = None
    banco_cuenta_id: int           = Field(..., description="ID de la cuenta de banco (plan_cuentas)")


class AcreditarMasivoIn(BaseModel):
    cheque_ids:     List[int]
    banco_cuenta_id: int
    fecha_acred:    Optional[date] = None


class RechazarIn(BaseModel):
    fecha_rechazo:    Optional[date] = None
    gastos_bancarios: float          = Field(0.0, ge=0)
    fisico:           bool           = False
    fecha_devolucion: Optional[date] = None


class PortadorIn(BaseModel):
    nombre: str


class BulkOcrItem(BaseModel):
    index:          int
    filename:       str
    numero:         Optional[str]   = None
    banco_origen:   Optional[str]   = None
    librador:       Optional[str]   = None
    monto:          Optional[float] = None
    fecha_emision:  Optional[str]   = None
    fecha_deposito: Optional[str]   = None
    codigo_postal:  Optional[str]   = None
    local_interior: Optional[str]   = None
    error:          bool            = False
    error_msg:      Optional[str]   = None


class BulkCrearItem(BaseModel):
    cliente_id:          Optional[int]   = None
    portador_id:         Optional[int]   = None
    numero:              Optional[str]   = None
    banco_origen:        Optional[str]   = None
    librador:            Optional[str]   = None
    monto:               float           = Field(..., gt=0)
    porcentaje_comision: Optional[float] = None
    codigo_postal:       Optional[str]   = None
    local_interior:      Optional[str]   = None
    fecha_emision:       Optional[str]   = None
    fecha_deposito:      Optional[str]   = None
    notas:               Optional[str]   = None


class BulkCrearIn(BaseModel):
    items:  List[BulkCrearItem]
    org_id: Optional[int] = None


# ── Helpers ───────────────────────────────────────────────────────

def _org_id(current_user: User, org_id: Optional[int]) -> int:
    if can_switch_org(current_user, org_id) and org_id:
        return org_id
    return current_user.organizacion_id or 1


def _local_interior(codigo_postal: Optional[str]) -> Optional[str]:
    if not codigo_postal:
        return None
    try:
        return "local" if int(codigo_postal) < 2000 else "interior"
    except (ValueError, TypeError):
        return None


def _cheque_dict(c: Cheque) -> dict:
    librador = c.librador or c.titular  # fallback para registros anteriores
    return {
        "id":                  c.id,
        "organizacion_id":     c.organizacion_id,
        "cliente_id":          c.cliente_id,
        "cliente_nombre":      c.cliente.nombre if c.cliente else None,
        "portador_id":         c.portador_id,
        "portador_nombre":     c.portador.nombre if c.portador else None,
        "numero":              c.numero,
        "banco_origen":        c.banco_origen,
        "librador":            librador,
        "monto":               c.monto,
        "comision":            c.comision,
        "porcentaje_comision": float(c.porcentaje_comision) if c.porcentaje_comision is not None else None,
        "codigo_postal":       c.codigo_postal,
        "local_interior":      c.local_interior,
        "fecha_emision":       c.fecha_emision,
        "fecha_deposito":      c.fecha_deposito,
        "fecha_acred":         c.fecha_acred,
        "estado":              c.estado,
        "fecha_rechazo":       c.fecha_rechazo,
        "fisico":              c.fisico,
        "fecha_devolucion":    c.fecha_devolucion,
        "notas":               c.notas,
        "tiene_foto":          bool(c.foto_comprobante),
        "banco_cuenta_id":     c.banco_cuenta_id,
        "created_at":          c.created_at,
    }


# ── Carga masiva OCR ──────────────────────────────────────────────

@router.post("/bulk-ocr")
async def bulk_ocr(
    fotos: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """Recibe hasta 30 fotos, aplica OCR a cada una en paralelo y devuelve los datos extraídos."""
    import asyncio
    import os
    import json
    import base64 as _base64

    MAX_FOTOS = 30
    if len(fotos) > MAX_FOTOS:
        raise HTTPException(400, f"Máximo {MAX_FOTOS} fotos por lote")
    if len(fotos) == 0:
        raise HTTPException(400, "Se requiere al menos una foto")

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(503, "OCR no configurado (falta GEMINI_API_KEY)")

    from routers.agente import _GEMINI_MODEL, _classify_gemini_error

    OCR_PROMPT = (
        "Extraé los datos de este cheque bancario argentino. "
        "Respondé SOLO con un JSON válido (sin texto extra, sin markdown), con estos campos "
        "(usá null si no está visible o no podés leerlo): "
        '{"numero": "string o null", "banco_origen": "string o null", "librador": "string o null", '
        '"monto": número_sin_formato_o_null, "fecha_emision": "YYYY-MM-DD o null", '
        '"fecha_deposito": "YYYY-MM-DD o null", "codigo_postal": "string o null"}'
    )

    async def _ocr_single(index: int, foto: UploadFile) -> dict:
        filename = foto.filename or f"foto_{index}"
        base_result = {
            "index": index,
            "filename": filename,
            "numero": None,
            "banco_origen": None,
            "librador": None,
            "monto": None,
            "fecha_emision": None,
            "fecha_deposito": None,
            "codigo_postal": None,
            "local_interior": None,
        }
        try:
            raw = await foto.read()
            mime_type = foto.content_type or "image/jpeg"
            # Llamada a Gemini en thread pool para no bloquear el event loop
            loop = asyncio.get_event_loop()

            def _call():
                import google.generativeai as genai
                import time as _time
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel(_GEMINI_MODEL)
                image_part = genai.protos.Part(
                    inline_data=genai.protos.Blob(mime_type=mime_type, data=raw)
                )
                for attempt in range(2):
                    try:
                        resp = model.generate_content([image_part, OCR_PROMPT])
                        try:
                            texto = resp.text.strip()
                        except Exception:
                            return {}
                        if not texto:
                            return {}
                        if texto.startswith("```"):
                            lines = [l for l in texto.split("\n") if not l.startswith("```")]
                            texto = "\n".join(lines).strip()
                        return json.loads(texto)
                    except Exception as ex:
                        msg = str(ex).upper()
                        is_transient = "RESOURCE_EXHAUSTED" in msg or "429" in str(ex)
                        if is_transient and attempt == 0:
                            _time.sleep(5)
                            continue
                        raise

            datos = await loop.run_in_executor(None, _call)
            cp = str(datos.get("codigo_postal") or "").strip() or None
            li = _local_interior(cp)
            return {
                **base_result,
                "numero":         str(datos["numero"]).strip()        if datos.get("numero")         else None,
                "banco_origen":   str(datos["banco_origen"]).strip()  if datos.get("banco_origen")   else None,
                "librador":       str(datos["librador"]).strip()      if datos.get("librador")        else None,
                "monto":          float(datos["monto"])               if datos.get("monto") is not None else None,
                "fecha_emision":  str(datos["fecha_emision"])         if datos.get("fecha_emision")   else None,
                "fecha_deposito": str(datos["fecha_deposito"])        if datos.get("fecha_deposito")  else None,
                "codigo_postal":  cp,
                "local_interior": li,
                "error":          False,
            }
        except json.JSONDecodeError:
            return {**base_result, "error": True, "error_msg": "Respuesta OCR inválida"}
        except Exception as ex:
            _, msg = _classify_gemini_error(ex)
            return {**base_result, "error": True, "error_msg": msg}

    tasks = [_ocr_single(i, foto) for i, foto in enumerate(fotos)]
    results = await asyncio.gather(*tasks)
    return {"items": list(results)}


@router.post("/bulk-crear")
def bulk_crear(
    body: BulkCrearIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Crea múltiples cheques en un solo request. Valida cliente por item; no aborta ante error parcial."""
    oid = _org_id(current_user, body.org_id)

    creados_cheques = []
    errores = []

    for idx, item in enumerate(body.items):
        try:
            cli = None
            if item.cliente_id:
                cli = db.query(Cliente).filter(
                    Cliente.id == item.cliente_id, Cliente.organizacion_id == oid
                ).first()
                if not cli:
                    errores.append({"index": idx, "msg": f"Cliente {item.cliente_id} no encontrado"})
                    continue
                if not cli.cuenta_contable_id:
                    errores.append({"index": idx, "msg": f"'{cli.nombre}' no tiene cuenta contable. Creala en Contabilidad → Clientes."})
                    continue

            if item.portador_id:
                if not db.query(Portador).filter(Portador.id == item.portador_id, Portador.organizacion_id == oid).first():
                    errores.append({"index": idx, "msg": f"Portador {item.portador_id} no encontrado"})
                    continue

            li = item.local_interior or _local_interior(item.codigo_postal)

            if item.porcentaje_comision is not None:
                pct_comision = Decimal(str(item.porcentaje_comision))
            elif cli:
                if li == "local" and cli.porcentaje_comision_local is not None:
                    pct_comision = cli.porcentaje_comision_local
                elif li == "interior" and cli.porcentaje_comision_interior is not None:
                    pct_comision = cli.porcentaje_comision_interior
                else:
                    pct_comision = cli.porcentaje_comision
            else:
                pct_comision = None

            monto_dec = Decimal(str(item.monto))
            comision_dec = Decimal("0")
            if pct_comision is not None:
                comision_dec = (monto_dec * pct_comision / Decimal("100")).quantize(Decimal("0.01"))

            fecha_dep = None
            if item.fecha_deposito:
                try:
                    from datetime import date as _date
                    fecha_dep = _date.fromisoformat(item.fecha_deposito[:10])
                except Exception:
                    fecha_dep = None
            fecha_dep = fecha_dep or hoy_art()

            fecha_emi = None
            if item.fecha_emision:
                try:
                    from datetime import date as _date
                    fecha_emi = _date.fromisoformat(item.fecha_emision[:10])
                except Exception:
                    fecha_emi = None

            c = Cheque(
                organizacion_id=oid,
                cliente_id=item.cliente_id,
                portador_id=item.portador_id,
                numero=item.numero,
                banco_origen=item.banco_origen,
                librador=item.librador,
                monto=monto_dec,
                comision=comision_dec,
                porcentaje_comision=pct_comision,
                codigo_postal=item.codigo_postal,
                local_interior=li,
                fecha_emision=fecha_emi,
                fecha_deposito=fecha_dep,
                estado="registrado",
                notas=item.notas,
                usuario_id=current_user.id,
            )
            db.add(c)
            db.flush()

            registrar_cheque(
                db=db, cheque_id=c.id, org_id=oid, usuario_id=current_user.id,
                titular=c.librador or "", monto=c.monto, comision=c.comision,
                fecha=fecha_dep,
            )
            registrar_log(db, current_user.id, "cheques", c.id, "INSERT",
                          {"monto": float(c.monto), "librador": c.librador,
                           "cliente_id": c.cliente_id, "bulk": True,
                           "fecha_deposito": str(fecha_dep)})
            creados_cheques.append(c)

        except Exception as ex:
            db.rollback()
            errores.append({"index": idx, "msg": str(ex)})
            continue

    db.commit()
    for c in creados_cheques:
        try:
            db.refresh(c)
        except Exception:
            pass

    return {
        "creados":  len(creados_cheques),
        "errores":  errores,
        "cheques":  [_cheque_dict(c) for c in creados_cheques],
    }


# ── Portadores ────────────────────────────────────────────────────

@router.get("/portadores")
def list_portadores(
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    items = db.query(Portador).filter(Portador.organizacion_id == oid).order_by(Portador.nombre).all()
    return [{"id": p.id, "nombre": p.nombre} for p in items]


@router.post("/portadores")
def crear_portador(
    body: PortadorIn,
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    nombre = body.nombre.strip()
    if not nombre:
        raise HTTPException(400, "Nombre requerido")
    p = Portador(organizacion_id=oid, nombre=nombre)
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"id": p.id, "nombre": p.nombre}


# ── Export general ───────────────────────────────────────────────

@router.get("/exportar")
def exportar_todos_excel(
    org_id:    Optional[int] = Query(None),
    estado:    Optional[str] = Query(None),
    cliente_id: Optional[int] = Query(None),
    desde:     Optional[str] = Query(None),
    hasta:     Optional[str] = Query(None),
    db:        Session       = Depends(get_db),
    current_user: User       = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    q = db.query(Cheque).filter(Cheque.organizacion_id == oid)
    if estado:
        q = q.filter(Cheque.estado == estado)
    if cliente_id:
        q = q.filter(Cheque.cliente_id == cliente_id)
    if desde:
        q = q.filter(Cheque.fecha_deposito >= desde)
    if hasta:
        q = q.filter(Cheque.fecha_deposito <= hasta)
    cheques = q.order_by(Cheque.fecha_deposito.desc().nullslast(), Cheque.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cheques"

    HEADER_FILL = PatternFill("solid", fgColor="1E1E2E")
    HEADER_FONT = Font(bold=True, color="A0A0C0", size=9)
    TOTAL_FONT  = Font(bold=True, color="FFFFFF")

    headers = [
        "Estado", "F. Depósito", "F. Acred.", "Cliente", "Portador",
        "Librador", "Banco", "Número", "CP", "L/I",
        "Monto", "Comisión", "Notas",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 14, 14, 18, 18, 22, 14, 14, 10, 8, 16, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_monto = 0.0
    total_comision = 0.0

    for c in cheques:
        nombre   = c.cliente.nombre if c.cliente else "Sin cliente"
        portador = c.portador.nombre if c.portador else ""
        librador = c.librador or c.titular or ""
        monto    = float(c.monto)
        comision = float(c.comision) if c.comision else 0.0

        ws.append([
            c.estado.capitalize(),
            str(c.fecha_deposito) if c.fecha_deposito else "",
            str(c.fecha_acred)    if c.fecha_acred    else "",
            nombre, portador, librador,
            c.banco_origen or "",
            c.numero or "",
            c.codigo_postal or "",
            (c.local_interior or "").capitalize(),
            monto,
            comision if comision else "",
            c.notas or "",
        ])
        ws.cell(row=ws.max_row, column=11).number_format = '#,##0.00'
        if comision:
            ws.cell(row=ws.max_row, column=12).number_format = '#,##0.00'
        total_monto    += monto
        total_comision += comision

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", "", "", "", "", total_monto, total_comision if total_comision else ""])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = TOTAL_FONT
    ws.cell(row=total_row, column=11).number_format = '#,##0.00'
    if total_comision:
        ws.cell(row=total_row, column=12).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "cheques.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Vista por depósito ────────────────────────────────────────────

@router.get("/deposito/exportar")
def exportar_deposito_excel(
    fecha: str = Query(...),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    cheques = (
        db.query(Cheque)
        .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito == fecha)
        .order_by(Cheque.cliente_id, Cheque.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Dep {fecha}"

    HEADER_FILL = PatternFill("solid", fgColor="1E1E2E")
    HEADER_FONT = Font(bold=True, color="A0A0C0", size=9)
    TOTAL_FONT  = Font(bold=True, color="FFFFFF")
    GRAY_FONT   = Font(color="888888", size=8)

    headers = [
        "Fecha Depósito", "Cliente", "Fecha Cheque", "Banco N",
        "Banco", "Librador", "Número", "Código P", "L/I", "Importe",
    ]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    col_widths = [14, 18, 14, 16, 16, 22, 14, 10, 8, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    por_cliente: dict[str, float] = {}
    por_tipo: dict[str, float] = {"local": 0.0, "interior": 0.0, "sin CP": 0.0}
    total = 0.0

    for c in cheques:
        nombre   = c.cliente.nombre if c.cliente else "Sin cliente"
        portador = c.portador.nombre if c.portador else ""
        librador = c.librador or c.titular or ""
        li       = (c.local_interior or "").lower()
        monto    = float(c.monto)

        ws.append([
            str(c.fecha_deposito) if c.fecha_deposito else "",
            nombre,
            str(c.fecha_emision) if c.fecha_emision else "",
            portador,
            c.banco_origen or "",
            librador,
            c.numero or "",
            c.codigo_postal or "",
            (c.local_interior or "").capitalize(),
            monto,
        ])
        ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'

        por_cliente[nombre] = por_cliente.get(nombre, 0.0) + monto
        key = li if li in ("local", "interior") else "sin CP"
        por_tipo[key] = por_tipo.get(key, 0.0) + monto
        total += monto

    ws.append([])
    ws.append(["RESUMEN POR CLIENTE", "", "", "", "", "", "", "", "", "TOTAL"])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    for cliente_nombre, subtotal in sorted(por_cliente.items()):
        ws.append(["", cliente_nombre, "", "", "", "", "", "", "", subtotal])
        ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'
        ws.cell(row=ws.max_row, column=2).font = GRAY_FONT

    ws.append([])
    ws.append(["RESUMEN LOCAL / INTERIOR", "", "", "", "", "", "", "", "", "TOTAL"])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    for tipo, subtotal in por_tipo.items():
        if subtotal:
            ws.append(["", tipo.capitalize(), "", "", "", "", "", "", "", subtotal])
            ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'
            ws.cell(row=ws.max_row, column=2).font = GRAY_FONT

    ws.append([])
    ws.append(["TOTAL GENERAL", "", "", "", "", "", "", "", "", total])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = TOTAL_FONT
    ws.cell(row=total_row, column=10).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cheques_deposito_{fecha}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/deposito")
def resumen_deposito(
    fecha: Optional[str] = Query(None),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)

    if not fecha:
        fechas = (
            db.query(Cheque.fecha_deposito)
            .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito.isnot(None))
            .distinct()
            .order_by(Cheque.fecha_deposito.desc())
            .all()
        )
        return {"fechas": [str(f[0]) for f in fechas]}

    cheques = (
        db.query(Cheque)
        .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito == fecha)
        .order_by(Cheque.cliente_id, Cheque.id)
        .all()
    )
    items = [_cheque_dict(c) for c in cheques]

    por_cliente: dict[str, float] = {}
    por_tipo: dict[str, float] = {}
    total = 0.0
    for c in cheques:
        nombre = c.cliente.nombre if c.cliente else "Sin cliente"
        li     = (c.local_interior or "sin CP").lower()
        monto  = float(c.monto)
        por_cliente[nombre] = por_cliente.get(nombre, 0.0) + monto
        por_tipo[li]        = por_tipo.get(li, 0.0) + monto
        total += monto

    return {
        "fecha":   fecha,
        "items":   items,
        "resumen": {"por_cliente": por_cliente, "por_tipo": por_tipo, "total": total, "cantidad": len(cheques)},
    }


# ── CRUD principal ────────────────────────────────────────────────

@router.get("")
def list_cheques(
    org_id:     Optional[int] = Query(None),
    estado:     Optional[str] = Query(None),
    cliente_id: Optional[int] = Query(None),
    desde:      Optional[str] = Query(None),
    hasta:      Optional[str] = Query(None),
    skip:       int = 0,
    limit:      int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    q = db.query(Cheque).filter(Cheque.organizacion_id == oid)
    if estado:
        q = q.filter(Cheque.estado == estado)
    if cliente_id:
        q = q.filter(Cheque.cliente_id == cliente_id)
    if desde:
        q = q.filter(Cheque.fecha_deposito >= desde)
    if hasta:
        q = q.filter(Cheque.fecha_deposito <= hasta)
    total = q.count()
    items = q.order_by(Cheque.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_cheque_dict(c) for c in items]}


@router.post("")
def crear_cheque(
    body: ChequeIn,
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    if body.cliente_id:
        cli = db.query(Cliente).filter(Cliente.id == body.cliente_id, Cliente.organizacion_id == oid).first()
        if not cli:
            raise HTTPException(404, "Cliente no encontrado")
        if not cli.cuenta_contable_id:
            raise HTTPException(
                400,
                f"El cliente '{cli.nombre}' no tiene cuenta contable configurada. "
                "Creala en Contabilidad → Clientes antes de registrar cheques."
            )
    if body.portador_id:
        if not db.query(Portador).filter(Portador.id == body.portador_id, Portador.organizacion_id == oid).first():
            raise HTTPException(404, "Portador no encontrado")

    li = body.local_interior or _local_interior(body.codigo_postal)

    # % comisión: usa el del body si vino; si no, lo hereda del cliente según
    # local/interior (con fallback al % general del cliente).
    if body.porcentaje_comision is not None:
        pct_comision = Decimal(str(body.porcentaje_comision))
    elif body.cliente_id and cli:
        if li == "local" and cli.porcentaje_comision_local is not None:
            pct_comision = cli.porcentaje_comision_local
        elif li == "interior" and cli.porcentaje_comision_interior is not None:
            pct_comision = cli.porcentaje_comision_interior
        else:
            pct_comision = cli.porcentaje_comision
    else:
        pct_comision = None

    c = Cheque(
        organizacion_id=oid,
        cliente_id=body.cliente_id,
        portador_id=body.portador_id,
        numero=body.numero,
        banco_origen=body.banco_origen,
        librador=body.librador,
        monto=body.monto,
        comision=body.comision,
        porcentaje_comision=pct_comision,
        codigo_postal=body.codigo_postal,
        local_interior=li,
        fecha_emision=body.fecha_emision,
        fecha_deposito=body.fecha_deposito or hoy_art(),
        estado="registrado",
        notas=body.notas,
        usuario_id=current_user.id,
    )
    db.add(c)
    db.flush()

    registrar_cheque(
        db=db, cheque_id=c.id, org_id=oid, usuario_id=current_user.id,
        titular=c.librador or "", monto=c.monto, comision=c.comision,
        fecha=c.fecha_deposito or hoy_art(),
    )
    registrar_log(db, current_user.id, "cheques", c.id, "INSERT",
                  {"monto": c.monto, "librador": c.librador, "cliente_id": c.cliente_id,
                   "fecha_deposito": str(c.fecha_deposito) if c.fecha_deposito else None})
    db.commit()
    db.refresh(c)
    return _cheque_dict(c)


@router.get("/{cheque_id}")
def get_cheque(
    cheque_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    return _cheque_dict(c)


@router.patch("/{cheque_id}")
def editar_cheque(
    cheque_id: int,
    body: ChequeIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    if c.estado not in ("registrado", "pendiente"):
        raise HTTPException(400, "Solo se pueden editar cheques en estado registrado")

    cambios = {}
    for field in ("cliente_id", "portador_id", "numero", "banco_origen", "librador",
                  "monto", "comision", "codigo_postal", "fecha_emision", "fecha_deposito", "notas"):
        val = getattr(body, field, None)
        if val is not None:
            cambios[field] = {"de": str(getattr(c, field)), "a": str(val)}
            setattr(c, field, val)
    if body.codigo_postal:
        c.local_interior = body.local_interior or _local_interior(body.codigo_postal)
    if body.porcentaje_comision is not None:
        c.porcentaje_comision = Decimal(str(body.porcentaje_comision))
    if cambios:
        registrar_log(db, current_user.id, "cheques", c.id, "UPDATE", cambios)
    db.commit()
    db.refresh(c)
    return _cheque_dict(c)


@router.post("/{cheque_id}/acreditar")
def acreditar(
    cheque_id: int,
    body: AcreditarIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    if c.estado not in ("registrado", "depositado", "pendiente"):
        raise HTTPException(400, f"Cheque ya está {c.estado}")

    # Verificar que el banco elegido existe
    banco_cuenta = db.query(PlanCuenta).filter(PlanCuenta.id == body.banco_cuenta_id).first()
    if not banco_cuenta:
        raise HTTPException(404, "Cuenta de banco no encontrada")

    # Verificar que el cliente tiene cuenta contable
    cli = db.query(Cliente).filter(Cliente.id == c.cliente_id).first() if c.cliente_id else None
    if not cli or not cli.cuenta_contable_id:
        raise HTTPException(400, "El cliente no tiene cuenta contable configurada")

    c.estado          = "acreditado"
    c.fecha_acred     = body.fecha_acred or hoy_art()
    c.banco_cuenta_id = body.banco_cuenta_id
    db.flush()

    neto = Decimal(str(c.monto)) - Decimal(str(c.comision or 0))
    acreditar_cheque(
        db=db, cheque_id=c.id, org_id=c.organizacion_id, usuario_id=current_user.id,
        titular=c.librador or c.titular or "",
        monto=Decimal(str(c.monto)), neto=neto,
        banco_cuenta_id=body.banco_cuenta_id,
        cliente_cuenta_id=cli.cuenta_contable_id,
        fecha=c.fecha_acred,
    )
    registrar_log(db, current_user.id, "cheques", c.id, "ACREDITAR",
                  {"monto": float(c.monto), "fecha_acred": str(c.fecha_acred),
                   "banco": banco_cuenta.nombre})
    db.commit()
    db.refresh(c)
    return _cheque_dict(c)


@router.post("/acreditar")
def acreditar_masivo(
    body: AcreditarMasivoIn,
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Acredita uno o varios cheques de una sola vez (flujo Por depósito)."""
    oid = _org_id(current_user, org_id)

    banco_cuenta = db.query(PlanCuenta).filter(PlanCuenta.id == body.banco_cuenta_id).first()
    if not banco_cuenta:
        raise HTTPException(404, "Cuenta de banco no encontrada")

    fecha = body.fecha_acred or hoy_art()
    resultados = []
    for cheque_id in body.cheque_ids:
        c = db.query(Cheque).filter(Cheque.id == cheque_id, Cheque.organizacion_id == oid).first()
        if not c:
            resultados.append({"id": cheque_id, "ok": False, "error": "No encontrado"})
            continue
        if c.estado not in ("registrado", "depositado", "pendiente"):
            resultados.append({"id": cheque_id, "ok": False, "error": f"Estado {c.estado}"})
            continue
        cli = db.query(Cliente).filter(Cliente.id == c.cliente_id).first() if c.cliente_id else None
        if not cli or not cli.cuenta_contable_id:
            resultados.append({"id": cheque_id, "ok": False,
                               "error": f"Cliente sin cuenta contable"})
            continue

        c.estado          = "acreditado"
        c.fecha_acred     = fecha
        c.banco_cuenta_id = body.banco_cuenta_id
        db.flush()

        neto = Decimal(str(c.monto)) - Decimal(str(c.comision or 0))
        acreditar_cheque(
            db=db, cheque_id=c.id, org_id=c.organizacion_id, usuario_id=current_user.id,
            titular=c.librador or c.titular or "",
            monto=Decimal(str(c.monto)), neto=neto,
            banco_cuenta_id=body.banco_cuenta_id,
            cliente_cuenta_id=cli.cuenta_contable_id,
            fecha=fecha,
        )
        registrar_log(db, current_user.id, "cheques", c.id, "ACREDITAR",
                      {"monto": float(c.monto), "banco": banco_cuenta.nombre})
        resultados.append({"id": cheque_id, "ok": True})

    db.commit()
    ok_count = sum(1 for r in resultados if r["ok"])
    return {"acreditados": ok_count, "total": len(body.cheque_ids), "detalle": resultados}


@router.post("/{cheque_id}/rechazar")
def rechazar(
    cheque_id: int,
    body: RechazarIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    if c.estado != "acreditado":
        raise HTTPException(400, "Solo se pueden rechazar cheques acreditados")

    cli = db.query(Cliente).filter(Cliente.id == c.cliente_id).first() if c.cliente_id else None
    if not cli or not cli.cuenta_contable_id:
        raise HTTPException(400, "El cliente no tiene cuenta contable configurada")
    if not c.banco_cuenta_id:
        raise HTTPException(400, "El cheque no tiene banco registrado de la acreditación")

    c.estado           = "rechazado"
    c.fecha_rechazo    = body.fecha_rechazo or hoy_art()
    c.fisico           = body.fisico
    c.fecha_devolucion = body.fecha_devolucion
    db.flush()

    rechazar_cheque(
        db=db, cheque_id=c.id, org_id=c.organizacion_id, usuario_id=current_user.id,
        titular=c.librador or c.titular or "",
        monto=Decimal(str(c.monto)),
        gastos=Decimal(str(body.gastos_bancarios)),
        banco_cuenta_id=c.banco_cuenta_id,
        cliente_cuenta_id=cli.cuenta_contable_id,
        fecha=c.fecha_rechazo,
    )
    registrar_log(db, current_user.id, "cheques", c.id, "RECHAZAR",
                  {"monto": float(c.monto), "gastos": body.gastos_bancarios,
                   "fecha_rechazo": str(c.fecha_rechazo), "fisico": c.fisico})
    db.commit()
    db.refresh(c)
    return _cheque_dict(c)


@router.delete("/{cheque_id}")
def eliminar_cheque(
    cheque_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("delete_records")),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    if c.estado not in ("registrado", "pendiente"):
        raise HTTPException(400, "Solo se pueden eliminar cheques en estado registrado")

    from app.services.motor_contable import reversar_asientos
    motivo = f"Cheque #{cheque_id} eliminado por {current_user.email}"
    # Revertir asiento nuevo (cheque_registro) y también los legacy por si existieran
    for mod in ("cheque_registro", "cheque_carga", "cheque_comision"):
        reversar_asientos(db, modulo=mod, referencia_id=cheque_id,
                          org_id=c.organizacion_id, usuario_id=current_user.id, motivo=motivo)

    registrar_log(db, current_user.id, "cheques", cheque_id, "DELETE",
                  {"monto": c.monto, "librador": c.librador, "estado": c.estado})
    db.delete(c)
    db.commit()
    return {"ok": True}


# ── Foto comprobante ──────────────────────────────────────────────

class FotoIn(BaseModel):
    foto_base64: str


@router.post("/{cheque_id}/foto")
def subir_foto(
    cheque_id: int,
    body: FotoIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    c.foto_comprobante = upload_comprobante(body.foto_base64, prefix=f"cheque/{c.organizacion_id}")
    db.commit()
    return {"ok": True, "tiene_foto": True}


@router.get("/{cheque_id}/foto")
def ver_foto(
    cheque_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    if not c.foto_comprobante:
        raise HTTPException(404, "Sin foto")
    return {"cheque_id": cheque_id, "foto_base64": c.foto_comprobante}


@router.delete("/{cheque_id}/foto")
def eliminar_foto(
    cheque_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("delete_records")),
):
    oid = current_user.organizacion_id or 1
    c = db.query(Cheque).filter(Cheque.id == cheque_id).first()
    if not c:
        raise HTTPException(404, "Cheque no encontrado")
    if not current_user.is_superadmin and c.organizacion_id != oid:
        raise HTTPException(403, "Sin acceso")
    c.foto_comprobante = None
    db.commit()
    return {"ok": True}


# ── Importación masiva por Excel ──────────────────────────────────

def _parse_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date):
        return v
    try:
        return date.fromisoformat(str(v)[:10])
    except Exception:
        return None


@router.post("/importar")
async def importar_excel(
    file: UploadFile = File(...),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    ext = (file.filename or '').lower().split('.')[-1]
    if ext not in ('xlsx', 'xls'):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")

    if not rows:
        raise HTTPException(400, "El archivo está vacío")

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    COL = {
        'librador':       next((i for i, h in enumerate(headers) if 'librador' in h or 'titular' in h), None),
        'banco_origen':   next((i for i, h in enumerate(headers) if 'banco' in h and 'nuestro' not in h), None),
        'numero':         next((i for i, h in enumerate(headers) if 'numer' in h or 'cheque' in h), None),
        'monto':          next((i for i, h in enumerate(headers) if 'monto' in h or 'importe' in h), None),
        'comision':       next((i for i, h in enumerate(headers) if 'comis' in h), None),
        'fecha_emision':  next((i for i, h in enumerate(headers) if 'emisi' in h), None),
        'fecha_deposito': next((i for i, h in enumerate(headers) if 'deposito' in h or 'depósito' in h), None),
        'cliente':        next((i for i, h in enumerate(headers) if 'cliente' in h), None),
        'codigo_postal':  next((i for i, h in enumerate(headers) if 'codigo' in h or 'postal' in h), None),
        'notas':          next((i for i, h in enumerate(headers) if 'nota' in h or 'observ' in h), None),
    }
    if COL['monto'] is None:
        raise HTTPException(400, "Columna 'monto' requerida no encontrada")

    clientes_cache: dict[str, int] = {
        c.nombre.lower(): c.id
        for c in db.query(Cliente).filter(Cliente.organizacion_id == oid).all()
    }

    importados, errores = 0, []
    for i, row in enumerate(rows[1:], start=2):
        try:
            monto_raw = row[COL['monto']] if COL['monto'] is not None else None
            if not monto_raw:
                continue
            monto = Decimal(str(monto_raw).replace(',', '.').replace('$', '').strip())
            if monto <= 0:
                continue

            cliente_id = None
            if COL['cliente'] is not None and row[COL['cliente']]:
                cv = str(row[COL['cliente']]).strip()
                cliente_id = int(cv) if cv.isdigit() else clientes_cache.get(cv.lower())

            comision = Decimal("0")
            if COL['comision'] is not None and row[COL['comision']]:
                try:
                    comision = Decimal(str(row[COL['comision']]).replace(',', '.').replace('$', '').strip())
                except Exception:
                    pass

            cp = str(row[COL['codigo_postal']]).strip() if COL['codigo_postal'] is not None and row[COL['codigo_postal']] else None
            c = Cheque(
                organizacion_id=oid,
                cliente_id=cliente_id,
                librador=str(row[COL['librador']]).strip() if COL['librador'] is not None and row[COL['librador']] else None,
                banco_origen=str(row[COL['banco_origen']]).strip() if COL['banco_origen'] is not None and row[COL['banco_origen']] else None,
                numero=str(row[COL['numero']]).strip() if COL['numero'] is not None and row[COL['numero']] else None,
                monto=monto, comision=comision,
                codigo_postal=cp, local_interior=_local_interior(cp),
                fecha_emision=_parse_date(row[COL['fecha_emision']]) if COL['fecha_emision'] is not None else None,
                fecha_deposito=_parse_date(row[COL['fecha_deposito']]) if COL['fecha_deposito'] is not None else hoy_art(),
                notas=str(row[COL['notas']]).strip() if COL['notas'] is not None and row[COL['notas']] else None,
                estado="registrado", usuario_id=current_user.id,
            )
            db.add(c)
            db.flush()
            registrar_cheque(
                db=db, cheque_id=c.id, org_id=oid, usuario_id=current_user.id,
                titular=c.librador or "", monto=c.monto, comision=c.comision,
                fecha=c.fecha_deposito or hoy_art(),
            )
            importados += 1
        except Exception as ex:
            errores.append(f"Fila {i}: {ex}")

    db.commit()
    return {"importados": importados, "errores": errores}
