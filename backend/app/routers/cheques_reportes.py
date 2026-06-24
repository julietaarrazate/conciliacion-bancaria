"""Router cheques — exportación Excel y vista por depósito.

Rutas expuestas (bajo el prefix /cheques del router padre):
  GET /exportar
  GET /deposito/exportar
  GET /deposito
"""
from typing import Optional
import io

from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.middleware.auth import get_current_user
from app.models.user import User
from app.models.cheque import Cheque

from .cheques_common import _org_id, _cheque_dict

router = APIRouter(tags=["cheques"])


# ── Export general ───────────────────────────────────────────────

@router.get("/exportar")
def exportar_todos_excel(
    org_id:    Optional[int] = Query(None),
    estado:    Optional[str] = Query(None),
    cliente_id: Optional[int] = Query(None),
    desde:     Optional[str] = Query(None),
    hasta:     Optional[str] = Query(None),
    db:        Session       = Depends(get_db),
    current_user: User       = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    q = db.query(Cheque).filter(Cheque.organizacion_id == oid)
    if estado:
        q = q.filter(Cheque.estado == estado)
    if cliente_id:
        q = q.filter(Cheque.cliente_id == cliente_id)
    if desde:
        q = q.filter(Cheque.fecha_deposito >= desde)
    if hasta:
        q = q.filter(Cheque.fecha_deposito <= hasta)
    cheques = q.order_by(Cheque.fecha_deposito.desc().nullslast(), Cheque.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cheques"

    HEADER_FILL = PatternFill("solid", fgColor="1E1E2E")
    HEADER_FONT = Font(bold=True, color="A0A0C0", size=9)
    TOTAL_FONT  = Font(bold=True, color="FFFFFF")

    headers = [
        "Estado", "F. Depósito", "F. Acred.", "Cliente", "Portador",
        "Librador", "Banco", "Número", "CP", "L/I",
        "Monto", "Comisión", "Notas",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 14, 14, 18, 18, 22, 14, 14, 10, 8, 16, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_monto = 0.0
    total_comision = 0.0

    for c in cheques:
        nombre   = c.cliente.nombre if c.cliente else "Sin cliente"
        portador = c.portador.nombre if c.portador else ""
        librador = c.librador or c.titular or ""
        monto    = float(c.monto)
        comision = float(c.comision) if c.comision else 0.0

        ws.append([
            c.estado.capitalize(),
            str(c.fecha_deposito) if c.fecha_deposito else "",
            str(c.fecha_acred)    if c.fecha_acred    else "",
            nombre, portador, librador,
            c.banco_origen or "",
            c.numero or "",
            c.codigo_postal or "",
            (c.local_interior or "").capitalize(),
            monto,
            comision if comision else "",
            c.notas or "",
        ])
        ws.cell(row=ws.max_row, column=11).number_format = '#,##0.00'
        if comision:
            ws.cell(row=ws.max_row, column=12).number_format = '#,##0.00'
        total_monto    += monto
        total_comision += comision

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", "", "", "", "", total_monto, total_comision if total_comision else ""])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = TOTAL_FONT
    ws.cell(row=total_row, column=11).number_format = '#,##0.00'
    if total_comision:
        ws.cell(row=total_row, column=12).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "cheques.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Vista por depósito ────────────────────────────────────────────

@router.get("/deposito/exportar")
def exportar_deposito_excel(
    fecha: str = Query(...),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)
    cheques = (
        db.query(Cheque)
        .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito == fecha)
        .order_by(Cheque.cliente_id, Cheque.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Dep {fecha}"

    HEADER_FILL = PatternFill("solid", fgColor="1E1E2E")
    HEADER_FONT = Font(bold=True, color="A0A0C0", size=9)
    TOTAL_FONT  = Font(bold=True, color="FFFFFF")
    GRAY_FONT   = Font(color="888888", size=8)

    headers = [
        "Fecha Depósito", "Cliente", "Fecha Cheque", "Banco N",
        "Banco", "Librador", "Número", "Código P", "L/I", "Importe",
    ]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    col_widths = [14, 18, 14, 16, 16, 22, 14, 10, 8, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    por_cliente: dict[str, float] = {}
    por_tipo: dict[str, float] = {"local": 0.0, "interior": 0.0, "sin CP": 0.0}
    total = 0.0

    for c in cheques:
        nombre   = c.cliente.nombre if c.cliente else "Sin cliente"
        portador = c.portador.nombre if c.portador else ""
        librador = c.librador or c.titular or ""
        li       = (c.local_interior or "").lower()
        monto    = float(c.monto)

        ws.append([
            str(c.fecha_deposito) if c.fecha_deposito else "",
            nombre,
            str(c.fecha_emision) if c.fecha_emision else "",
            portador,
            c.banco_origen or "",
            librador,
            c.numero or "",
            c.codigo_postal or "",
            (c.local_interior or "").capitalize(),
            monto,
        ])
        ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'

        por_cliente[nombre] = por_cliente.get(nombre, 0.0) + monto
        key = li if li in ("local", "interior") else "sin CP"
        por_tipo[key] = por_tipo.get(key, 0.0) + monto
        total += monto

    ws.append([])
    ws.append(["RESUMEN POR CLIENTE", "", "", "", "", "", "", "", "", "TOTAL"])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    for cliente_nombre, subtotal in sorted(por_cliente.items()):
        ws.append(["", cliente_nombre, "", "", "", "", "", "", "", subtotal])
        ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'
        ws.cell(row=ws.max_row, column=2).font = GRAY_FONT

    ws.append([])
    ws.append(["RESUMEN LOCAL / INTERIOR", "", "", "", "", "", "", "", "", "TOTAL"])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    for tipo, subtotal in por_tipo.items():
        if subtotal:
            ws.append(["", tipo.capitalize(), "", "", "", "", "", "", "", subtotal])
            ws.cell(row=ws.max_row, column=10).number_format = '#,##0.00'
            ws.cell(row=ws.max_row, column=2).font = GRAY_FONT

    ws.append([])
    ws.append(["TOTAL GENERAL", "", "", "", "", "", "", "", "", total])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = TOTAL_FONT
    ws.cell(row=total_row, column=10).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cheques_deposito_{fecha}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/deposito")
def resumen_deposito(
    fecha: Optional[str] = Query(None),
    org_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oid = _org_id(current_user, org_id)

    if not fecha:
        fechas = (
            db.query(Cheque.fecha_deposito)
            .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito.isnot(None))
            .distinct()
            .order_by(Cheque.fecha_deposito.desc())
            .all()
        )
        return {"fechas": [str(f[0]) for f in fechas]}

    cheques = (
        db.query(Cheque)
        .filter(Cheque.organizacion_id == oid, Cheque.fecha_deposito == fecha)
        .order_by(Cheque.cliente_id, Cheque.id)
        .all()
    )
    items = [_cheque_dict(c) for c in cheques]

    por_cliente: dict[str, float] = {}
    por_tipo: dict[str, float] = {}
    total = 0.0
    for c in cheques:
        nombre = c.cliente.nombre if c.cliente else "Sin cliente"
        li     = (c.local_interior or "sin CP").lower()
        monto  = float(c.monto)
        por_cliente[nombre] = por_cliente.get(nombre, 0.0) + monto
        por_tipo[li]        = por_tipo.get(li, 0.0) + monto
        total += monto

    return {
        "fecha":   fecha,
        "items":   items,
        "resumen": {"por_cliente": por_cliente, "por_tipo": por_tipo, "total": total, "cantidad": len(cheques)},
    }
