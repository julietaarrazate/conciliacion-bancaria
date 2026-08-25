"""
Genera archivos de prueba realistas:
- extracto_prueba.xlsx    -> 30 movimientos, orden DESCENDENTE (30 arriba, 1 abajo)
- ultimos_movimientos.xlsx-> UM con 5 filas: 2 ya estan en extracto + 3 nuevos (31,32,33)
- planilla_green.xlsx     -> 8 pagos (5 OK, 2 faltan_datos, 1 no_esta)
- planilla_alojando.xlsx  -> 6 pagos (5 OK, 1 no_esta)
- planilla_tucu.xlsx      -> 7 pagos (6 OK, 1 faltan_datos)

Salida (configurable con la variable de entorno DATOS_PRUEBA_DIR):
    <carpeta del usuario>/Desktop/INBOX/datos_prueba/
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date, timedelta
import os

DESKTOP  = os.path.join(os.path.expanduser("~"), "Desktop")
OUT_DIR  = os.getenv("DATOS_PRUEBA_DIR", os.path.join(DESKTOP, "INBOX", "datos_prueba"))
os.makedirs(OUT_DIR, exist_ok=True)

FECHA_BASE = date(2026, 4, 15)

# (dias_atras, cuit, nombre_titular, monto)
# Orden 1 = mas antiguo (index 0 = 15 dias atras)
# Orden 30 = mas reciente (index 29 = 1 dia atras)
MOVIMIENTOS = [
    (15, "27112233447", "LOZANO BEATRIZ",    310_000),   # orden 1  - mas antiguo
    (15, "20001122336", "CABRERA OSCAR",     175_000),   # orden 2
    (14, "30990011225", "DISTRIBUIDORA SA",  650_000),   # orden 3
    (14, "20889900114", "FLORES ANDRES",     98_000),    # orden 4
    (13, "27778899003", "SILVA PATRICIA",    240_000),   # orden 5
    (13, "20667788992", "ROJAS SERGIO",      160_000),   # orden 6
    (12, "27556677891", "PAREDES ELENA",     410_000),   # orden 7
    (12, "20445566780", "AGUIRRE MARTIN",    55_000),    # orden 8
    (11, "27334455669", "IBARRA CECILIA",    290_000),   # orden 9
    (11, "20223344557", "CASTILLO PEDRO",    380_000),   # orden 10
    (10, "27112233446", "MENDOZA NATALIA",   125_000),   # orden 11
    (10, "20001122335", "VARGAS ROBERTO",    190_000),   # orden 12
    (9,  "30990011224", "INVERSIONES SRL",   750_000),   # orden 13
    (9,  "20889900113", "RUIZ ALBERTO",      67_800),    # orden 14
    (8,  "27778899002", "JIMENEZ CLAUDIA",   220_000),   # orden 15
    (8,  "20667788991", "MORENO JOSE",       340_000),   # orden 16
    (7,  "27556677890", "DIAZ VALERIA",      115_000),   # orden 17
    (7,  "20445566779", "HERRERA LUIS",      180_000),   # orden 18
    (6,  "27334455668", "RAMIREZ SOFIA",     430_000),   # orden 19
    (6,  "20223344556", "TORRES MIGUEL",     260_000),   # orden 20
    (5,  "27112233445", "SANCHEZ PAULA",     95_000),    # orden 21
    (5,  "20001122334", "RODRIGUEZ DIEGO",   78_300),    # orden 22
    (4,  "30990011223", "EMPRESA SA",        500_000),   # orden 23 - monto comun 1/3
    (4,  "20889900112", "GOMEZ CARLOS",      145_000),   # orden 24
    (3,  "27778899001", "PEREZ LAURA",       500_000),   # orden 25 - monto comun 2/3
    (3,  "20667788990", "MARTINEZ JUAN",     320_000),   # orden 26
    (2,  "30556677889", "COMERCIO SRL",      500_000),   # orden 27 - monto comun 3/3
    (2,  "20445566778", "FERNANDEZ PABLO",   87_500),    # orden 28
    (1,  "27334455667", "LOPEZ ANDREA",      200_000),   # orden 29
    (1,  "20112233440", "GARCIA MARIA",      150_000),   # orden 30 - mas reciente
]

# Movimientos NUEVOS para los Ultimos Movimientos (van a ser 31, 32, 33)
MOVIMIENTOS_NUEVOS = [
    (0,  "20334455668", "QUIROGA RAMON",     285_000),   # orden 31 - hoy
    (0,  "27445566779", "ACOSTA MIRIAM",     167_500),   # orden 32 - hoy
    (0,  "20556677890", "BENITES OSCAR",     92_000),    # orden 33 - hoy
]


def _escribir_headers(ws):
    headers = [None, "Orden", "Fecha", "Mes", "titular", "Importe Pesos", "Saldo", "cliente", "fecha acred"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        if h:
            c.fill = PatternFill("solid", fgColor="3483FA")
            c.font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16


def _escribir_fila(ws, row, orden, dias_atras, cuit, nombre, monto, saldo_actual):
    fecha = FECHA_BASE - timedelta(days=dias_atras)
    mes   = fecha.strftime("%B").upper()
    titular = f"TRANSF {cuit} {nombre}"

    ws.cell(row=row, column=2, value=orden)
    ws.cell(row=row, column=3, value=fecha).number_format = "DD/MM/YYYY"
    ws.cell(row=row, column=4, value=mes)
    ws.cell(row=row, column=5, value=titular)
    ws.cell(row=row, column=6, value=monto).number_format = "#,##0.00"
    ws.cell(row=row, column=7, value=round(saldo_actual, 2)).number_format = "#,##0.00"
    ws.cell(row=row, column=8, value=None)
    ws.cell(row=row, column=9, value=None)


def crear_extracto():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracto"

    ws["A1"] = "Extracto Bancario - Banco Macro (PRUEBA)"
    ws["A1"].font = Font(bold=True, size=12)

    _escribir_headers(ws)

    # Calcular saldo base (suma de todos los montos partiendo de 5.000.000)
    saldo_base = 5_000_000.0

    # DESCENDENTE: orden 30 en fila 3, orden 1 en fila 32
    total = len(MOVIMIENTOS)
    for i, (dias_atras, cuit, nombre, monto) in enumerate(reversed(MOVIMIENTOS)):
        orden = total - i          # 30, 29, 28, ..., 1
        row   = i + 3              # fila 3, 4, 5, ...

        # Saldo: el mov mas reciente (orden 30) tiene el saldo mas alto
        # Saldo de fila = saldo_base menos la suma de montos con orden < este
        montos_anteriores = sum(m[3] for j, m in enumerate(MOVIMIENTOS) if j < (total - 1 - i))
        saldo_fila = saldo_base - montos_anteriores

        _escribir_fila(ws, row, orden, dias_atras, cuit, nombre, monto, saldo_fila)

    path = os.path.join(OUT_DIR, "extracto_prueba.xlsx")
    wb.save(path)
    print("[OK] extracto_prueba.xlsx      30 movs, orden 30 arriba / 1 abajo (descendente)")
    return path


def crear_ultimos_movimientos():
    """
    UM con 5 filas en estructura identica al extracto:
    - filas 1-2: los 2 ultimos del extracto (orden 29 y 30) -> seran DUPLICADOS
    - filas 3-5: 3 movimientos nuevos (orden 31, 32, 33) -> se agregaran al extracto

    Al subirlo con 'Agregar UM' el sistema debe detectar:
      2 duplicados (ya existian) y agregar solo los 3 nuevos.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ultimos Movimientos"

    ws["A1"] = "Ultimos Movimientos - Banco Macro (PRUEBA)"
    ws["A1"].font = Font(bold=True, size=12)

    _escribir_headers(ws)

    # Saldo base continuado desde el extracto
    saldo_continuado = 5_000_000.0 - sum(m[3] for m in MOVIMIENTOS)

    # -- Filas 1-2: los 2 ultimos movimientos del extracto (DUPLICADOS) --
    # En el extracto, los dos mas recientes son los ultimos de MOVIMIENTOS:
    # MOVIMIENTOS[-1] = orden 30 (GARCIA MARIA), MOVIMIENTOS[-2] = orden 29 (LOPEZ ANDREA)
    duplicados = [
        (30, MOVIMIENTOS[-1]),   # orden 30
        (29, MOVIMIENTOS[-2]),   # orden 29
    ]
    for row_idx, (orden, (dias_atras, cuit, nombre, monto)) in enumerate(duplicados):
        row = row_idx + 3
        saldo_dup = saldo_continuado + monto * (2 - row_idx)
        _escribir_fila(ws, row, orden, dias_atras, cuit, nombre, monto, saldo_dup)

    # -- Filas 3-5: movimientos NUEVOS (ordenes 31, 32, 33) --
    saldo_new = saldo_continuado
    for row_idx, (dias_atras, cuit, nombre, monto) in enumerate(MOVIMIENTOS_NUEVOS):
        orden = 31 + row_idx
        row   = row_idx + 5    # filas 5, 6, 7
        saldo_new -= monto
        _escribir_fila(ws, row, orden, dias_atras, cuit, nombre, monto, saldo_new)

    path = os.path.join(OUT_DIR, "ultimos_movimientos_prueba.xlsx")
    wb.save(path)
    print("[OK] ultimos_movimientos_prueba.xlsx  2 dup (29,30) + 3 nuevos (31,32,33)")
    return path


def crear_planilla_green():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla Green"

    headers = ["Fecha", "Descripcion", "Importe", "CUIT", "Observaciones"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="00AA44")
        c.font = Font(bold=True, color="FFFFFF")

    filas = [
        (date(2026,4,14), "Pago cuota 1", 150_000, "20112233440"),   # OK  - orden 30
        (date(2026,4,13), "Pago cuota 2", 200_000, "27334455667"),   # OK  - orden 29
        (date(2026,4,12), "Pago cuota 3", 87_500,  "20445566778"),   # OK  - orden 28
        (date(2026,4,11), "Pago cuota 4", 320_000, "20667788990"),   # OK  - orden 26
        (date(2026,4,10), "Pago cuota 5", 145_000, "20889900112"),   # OK  - orden 24
        (date(2026,4,9),  "Pago cuota 6", 500_000, "20001122334"),   # FALTAN DATOS - comun, CUIT no coincide
        (date(2026,4,8),  "Pago cuota 7", 999_999, ""),              # NO ESTA - monto inexistente
        (date(2026,4,7),  "Pago cuota 8", 500_000, "20112233440"),   # FALTAN DATOS - comun, CUIT no coincide
    ]

    for row, (fecha, desc, importe, cuit) in enumerate(filas, start=2):
        ws.cell(row=row, column=1, value=fecha).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=importe).number_format = "#,##0.00"
        ws.cell(row=row, column=4, value=cuit)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    path = os.path.join(OUT_DIR, "planilla_green.xlsx")
    wb.save(path)
    print("[OK] planilla_green.xlsx       8 filas: 5 OK, 2 faltan_datos, 1 no_esta")
    return path


def crear_planilla_alojando():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alojando ABR"

    # Header corrido en fila 2 (como el real de Alojando)
    ws["A1"] = "ALOJANDO SA - Rendicion Abril 2026"
    ws["A1"].font = Font(bold=True, size=11)

    headers = ["N", "Fecha pago", "Concepto", "Monto", "CUIT ordenante", "Estado"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="E6B800")
        c.font = Font(bold=True)

    filas = [
        (1, date(2026,4,12), "Transferencia abono", 78_300,  "20001122334"),   # OK - orden 22
        (2, date(2026,4,11), "Transferencia abono", 95_000,  "27112233445"),   # OK - orden 21
        (3, date(2026,4,10), "Transferencia abono", 260_000, "20223344556"),   # OK - orden 20
        (4, date(2026,4,9),  "Transferencia abono", 430_000, "27334455668"),   # OK - orden 19
        (5, date(2026,4,8),  "Transferencia abono", 180_000, "20445566779"),   # OK - orden 18
        (6, date(2026,4,7),  "Transferencia abono", 123_456, ""),              # NO ESTA - monto inexistente
    ]

    for n, fecha, conc, monto, cuit in filas:
        row = n + 2
        ws.cell(row=row, column=1, value=n)
        ws.cell(row=row, column=2, value=fecha).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=3, value=conc)
        ws.cell(row=row, column=4, value=monto).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=cuit)

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    path = os.path.join(OUT_DIR, "planilla_alojando.xlsx")
    wb.save(path)
    print("[OK] planilla_alojando.xlsx    6 filas: 5 OK, 1 no_esta")
    return path


def crear_planilla_tucu():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tucu"

    headers = ["Importe", "Titular", "Fecha"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="1B3F73")
        c.font = Font(bold=True, color="FFFFFF")

    filas = [
        (115_000, "DIAZ VALERIA",      date(2026,4,8)),   # OK  - orden 17
        (340_000, "MORENO JOSE",       date(2026,4,7)),   # OK  - orden 16
        (220_000, "JIMENEZ CLAUDIA",   date(2026,4,6)),   # OK  - orden 15
        (67_800,  "RUIZ ALBERTO",      date(2026,4,5)),   # OK  - orden 14
        (750_000, "INVERSIONES SRL",   date(2026,4,4)),   # OK  - orden 13
        (500_000, "GARCIA MARIA",      date(2026,4,3)),   # FALTAN DATOS - comun, titular no coincide
        (190_000, "VARGAS ROBERTO",    date(2026,4,2)),   # OK  - orden 12
    ]

    for row, (monto, titular, fecha) in enumerate(filas, start=2):
        ws.cell(row=row, column=1, value=monto).number_format = "#,##0.00"
        ws.cell(row=row, column=2, value=titular)
        ws.cell(row=row, column=3, value=fecha).number_format = "DD/MM/YYYY"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14

    path = os.path.join(OUT_DIR, "planilla_tucu.xlsx")
    wb.save(path)
    print("[OK] planilla_tucu.xlsx        7 filas: 6 OK, 1 faltan_datos")
    return path


if __name__ == "__main__":
    print()
    print("Generando archivos de prueba...")
    print()
    crear_extracto()
    crear_ultimos_movimientos()
    crear_planilla_green()
    crear_planilla_alojando()
    crear_planilla_tucu()
    print()
    print("=" * 60)
    print(f"  Archivos en: {OUT_DIR}")
    print("=" * 60)
    print()
    print("  PASO 1 - Subir extracto:")
    print("    extracto_prueba.xlsx        (30 movs, orden 30 arriba)")
    print()
    print("  PASO 1b - Probar Ultimos Movimientos (boton + Agregar UM):")
    print("    ultimos_movimientos_prueba.xlsx")
    print("    -> 2 duplicados detectados (ord 29 y 30)")
    print("    -> 3 nuevos agregados (ord 31, 32, 33)")
    print("    -> total extracto queda con 33 movs")
    print()
    print("  PASO 2 - Conciliar planillas:")
    print("    planilla_green.xlsx     cliente: Green")
    print("    planilla_alojando.xlsx  cliente: Alojando")
    print("    planilla_tucu.xlsx      cliente: Tucu")
    print()
    print("  RESULTADOS ESPERADOS:")
    print("    Green:    5 OK | 2 faltan_datos | 1 no_esta")
    print("    Alojando: 5 OK | 1 no_esta")
    print("    Tucu:     6 OK | 1 faltan_datos")
    print()
    print("  MONTO COMUN: 500.000 aparece 3 veces (ord 23, 25, 27)")
    print("  -> requiere CUIT o titular exacto para acreditar")
    print()
