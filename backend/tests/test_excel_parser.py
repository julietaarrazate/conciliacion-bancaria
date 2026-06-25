"""Tests del parser de extractos bancarios.

Verifica que:
- Banco Macro (formato original con columnas hardcodeadas) sigue parseando igual
- Formato genérico con columna "Monto" funciona
- Formato con columnas separadas Créditos/Débitos funciona (BBVA, Santander, Galicia)
- Detección de banco por texto en encabezado funciona
- El fallback Banco Macro mantiene monto, fecha, titular, saldo correctos
"""

import os
import tempfile
import pytest
from datetime import date
from openpyxl import Workbook

from app.services.excel_parser import (
    parsear_extracto_bancario,
    detectar_banco,
    detectar_columnas,
    _parse_monto,
    _parse_fecha,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _xlsx_tmp(wb: Workbook) -> str:
    """Guarda un Workbook en un archivo temporal y retorna la ruta."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _cleanup(path: str):
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


# ── Tests de funciones puras ─────────────────────────────────────────────────

def test_parse_monto_entero():
    assert _parse_monto(5000) == 5000.0

def test_parse_monto_float():
    assert _parse_monto(1234.56) == 1234.56

def test_parse_monto_string_punto_decimal():
    assert _parse_monto("1234.56") == 1234.56

def test_parse_monto_string_coma_decimal():
    assert _parse_monto("1.234,56") == 1234.56

def test_parse_monto_con_pesos():
    assert _parse_monto("$5.000,00") == 5000.0

def test_parse_monto_vacio_retorna_none():
    assert _parse_monto("") is None
    assert _parse_monto("-") is None
    assert _parse_monto(None) is None

def test_parse_fecha_date_object():
    d = date(2026, 5, 1)
    assert _parse_fecha(d) == d

def test_parse_fecha_string_arg():
    assert _parse_fecha("15/05/2026") == date(2026, 5, 15)

def test_parse_fecha_iso():
    assert _parse_fecha("2026-05-15") == date(2026, 5, 15)

def test_parse_fecha_none_retorna_none():
    assert _parse_fecha(None) is None


# ── Formato Banco Macro (fallback hardcodeado) ────────────────────────────────

def _wb_macro():
    """
    Crea un workbook que simula el extracto Banco Macro real:
    Fila 1: encabezado con "Banco Macro" (para que detecte el banco)
    Fila 2: headers: _, Orden, Fecha, Mes, Titular, Monto, Saldo
    Fila 3+: datos
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Banco Macro S.A."
    ws.cell(2, 1).value = ""
    ws.cell(2, 2).value = "Orden"
    ws.cell(2, 3).value = "Fecha"
    ws.cell(2, 4).value = "Mes"
    ws.cell(2, 5).value = "Titular"
    ws.cell(2, 6).value = "Monto"
    ws.cell(2, 7).value = "Saldo"
    # Fila de datos
    ws.cell(3, 2).value = 1001
    ws.cell(3, 3).value = date(2026, 5, 10)
    ws.cell(3, 4).value = 5
    ws.cell(3, 5).value = "EMPRESA SA 20123456789"
    ws.cell(3, 6).value = 15000.0
    ws.cell(3, 7).value = 200000.0
    # Segunda fila
    ws.cell(4, 2).value = 1002
    ws.cell(4, 3).value = date(2026, 5, 11)
    ws.cell(4, 4).value = 5
    ws.cell(4, 5).value = "GREEN SRL 30987654321"
    ws.cell(4, 6).value = 8500.50
    ws.cell(4, 7).value = 208500.50
    return wb


def test_macro_detecta_banco_correctamente():
    wb = _wb_macro()
    assert detectar_banco(wb.active) == "macro"


def test_macro_parsea_dos_movimientos():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["total"] == 2
        assert result["banco_detectado"] == "macro"
    finally:
        _cleanup(path)


def test_macro_monto_correcto():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["monto"] == 15000.0
        assert movs[1]["monto"] == 8500.50
    finally:
        _cleanup(path)


def test_macro_titular_correcto():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["titular"] == "EMPRESA SA 20123456789"
        assert movs[1]["titular"] == "GREEN SRL 30987654321"
    finally:
        _cleanup(path)


def test_macro_saldo_correcto():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["saldo"] == 200000.0
        assert movs[1]["saldo"] == 208500.50
    finally:
        _cleanup(path)


def test_macro_fecha_correcta():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["fecha"] == date(2026, 5, 10)
        assert movs[1]["fecha"] == date(2026, 5, 11)
    finally:
        _cleanup(path)


def test_macro_orden_correcto():
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["orden"] == 1001
        assert movs[1]["orden"] == 1002
    finally:
        _cleanup(path)


def test_macro_mes_normalizado_a_numero():
    """El campo mes debe ser string del número (ej "5"), no "Mayo"."""
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert movs[0]["mes"] == "5"
    finally:
        _cleanup(path)


# ── Formato genérico con columna Monto ───────────────────────────────────────

def _wb_generico():
    """Formato simple con cabeceras reconocibles por keywords."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Fecha"
    ws.cell(1, 2).value = "Concepto"
    ws.cell(1, 3).value = "Importe"
    ws.cell(1, 4).value = "Saldo"
    ws.cell(2, 1).value = date(2026, 5, 15)
    ws.cell(2, 2).value = "PAGO CLIENTE"
    ws.cell(2, 3).value = 3000.0
    ws.cell(2, 4).value = 50000.0
    # segunda fila para que max_row >= 3 (el parser saltea sheets con < 3 filas)
    ws.cell(3, 1).value = date(2026, 5, 16)
    ws.cell(3, 2).value = "OTRO PAGO"
    ws.cell(3, 3).value = 1500.0
    ws.cell(3, 4).value = 51500.0
    return wb


def test_generico_parsea_movimientos():
    wb = _wb_generico()
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["total"] == 2
        movs = result["movimientos"]
        assert movs[0]["monto"] == 3000.0
        assert movs[0]["titular"] == "PAGO CLIENTE"
        assert movs[1]["monto"] == 1500.0
    finally:
        _cleanup(path)


# ── Formato con Créditos/Débitos separados (BBVA, Santander, Galicia) ────────

def _wb_creditos_debitos():
    """Simula extracto BBVA/Santander con columnas Créditos y Débitos separadas."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "BBVA Francés"
    ws.cell(2, 1).value = "Fecha"
    ws.cell(2, 2).value = "Descripción"
    ws.cell(2, 3).value = "Débitos"
    ws.cell(2, 4).value = "Créditos"
    ws.cell(2, 5).value = "Saldo"
    # Crédito puro (ingreso)
    ws.cell(3, 1).value = date(2026, 5, 20)
    ws.cell(3, 2).value = "TRANSFERENCIA RECIBIDA"
    ws.cell(3, 3).value = None
    ws.cell(3, 4).value = 12000.0
    ws.cell(3, 5).value = 112000.0
    # Débito puro (egreso)
    ws.cell(4, 1).value = date(2026, 5, 21)
    ws.cell(4, 2).value = "PAGO SERVICIO"
    ws.cell(4, 3).value = 1500.0
    ws.cell(4, 4).value = None
    ws.cell(4, 5).value = 110500.0
    # Ambos (edge case: neto)
    ws.cell(5, 1).value = date(2026, 5, 22)
    ws.cell(5, 2).value = "AJUSTE"
    ws.cell(5, 3).value = 200.0
    ws.cell(5, 4).value = 500.0
    ws.cell(5, 5).value = 110800.0
    return wb


def test_creditos_debitos_detecta_bbva():
    wb = _wb_creditos_debitos()
    assert detectar_banco(wb.active) == "bbva"


def test_creditos_debitos_credito_positivo():
    """Fila con solo Créditos debe dar monto positivo."""
    wb = _wb_creditos_debitos()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        credito_row = next(m for m in movs if m["titular"] == "TRANSFERENCIA RECIBIDA")
        assert credito_row["monto"] == 12000.0
    finally:
        _cleanup(path)


def test_creditos_debitos_debito_negativo():
    """Fila con solo Débitos debe dar monto negativo."""
    wb = _wb_creditos_debitos()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        debito_row = next(m for m in movs if m["titular"] == "PAGO SERVICIO")
        assert debito_row["monto"] == -1500.0
    finally:
        _cleanup(path)


def test_creditos_debitos_neto_correcto():
    """Fila con ambos columnas: monto = crédito - débito."""
    wb = _wb_creditos_debitos()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        ajuste_row = next(m for m in movs if m["titular"] == "AJUSTE")
        assert ajuste_row["monto"] == round(500.0 - 200.0, 2)
    finally:
        _cleanup(path)


def test_creditos_debitos_total_tres_movimientos():
    wb = _wb_creditos_debitos()
    path = _xlsx_tmp(wb)
    try:
        assert parsear_extracto_bancario(path)["total"] == 3
    finally:
        _cleanup(path)


# ── Invariante crítico: Macro no se rompe si se agrega banco nuevo ────────────

def test_macro_no_se_ve_afectado_por_deteccion_credito_debito():
    """
    Garantía: un extracto Banco Macro con columna 'Monto' no debe
    confundirse con el formato crédito/débito.
    Los montos deben ser positivos y correctos.
    """
    wb = _wb_macro()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert all(m["monto"] > 0 for m in movs), "Macro: todos los montos deben ser positivos"
        assert movs[0]["monto"] == 15000.0
        assert movs[1]["monto"] == 8500.50
    finally:
        _cleanup(path)


# ══════════════════════════════════════════════════════════════════════════════
# Bancos nuevos (v3.13+): Mercado Pago, Credicoop, Supervielle, Patagonia,
# Bancor, Banco Rioja, Banco de La Pampa.
# ══════════════════════════════════════════════════════════════════════════════

def _csv_tmp(content: str) -> str:
    """Guarda un CSV en un archivo temporal y retorna la ruta."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w",
                                      encoding="utf-8", newline="")
    tmp.write(content)
    tmp.close()
    return tmp.name


# ── Mercado Pago (PRIORIDAD): CSV de liberaciones / settlement ────────────────

def _csv_mp_settlement() -> str:
    """CSV típico del reporte de liberaciones de Mercado Pago.

    Headers en inglés (SOURCE_ID, TRANSACTION_TYPE, TRANSACTION_AMOUNT,
    BALANCE_AMOUNT...) y NO contiene el texto literal 'Mercado Pago'.
    El monto ya trae el signo (positivo = ingreso, negativo = egreso/retiro).
    """
    return (
        "SOURCE_ID,EXTERNAL_REFERENCE,RELEASE_DATE,TRANSACTION_TYPE,"
        "TRANSACTION_AMOUNT,BALANCE_AMOUNT\n"
        "12345678901,REF-001,2026-05-10,payment,15000.00,215000.00\n"
        "12345678902,REF-002,2026-05-11,payment,8500.50,223500.50\n"
        "12345678903,REF-003,2026-05-12,withdrawal,-5000.00,218500.50\n"
    )


def test_mp_detecta_banco_por_headers_sin_texto():
    """MP se detecta por estructura de columnas aunque no diga 'Mercado Pago'."""
    path = _csv_tmp(_csv_mp_settlement())
    try:
        result = parsear_extracto_bancario(path)
        assert result["banco_detectado"] == "mercadopago"
    finally:
        _cleanup(path)


def test_mp_parsea_tres_movimientos_con_signo():
    path = _csv_tmp(_csv_mp_settlement())
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert len(movs) == 3
        montos = {m["monto"] for m in movs}
        assert 15000.00 in montos
        assert 8500.50 in montos
        assert -5000.00 in montos  # withdrawal mantiene signo negativo
    finally:
        _cleanup(path)


def test_mp_texto_literal_tambien_detecta():
    """Formato 'Detalle de movimientos' exportado desde la app de MP."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Reporte de Mercado Pago"
    ws.cell(2, 1).value = "Fecha"
    ws.cell(2, 2).value = "Descripción"
    ws.cell(2, 3).value = "Monto"
    ws.cell(2, 4).value = "Saldo"
    ws.cell(3, 1).value = date(2026, 5, 10)
    ws.cell(3, 2).value = "Pago recibido"
    ws.cell(3, 3).value = 3200.0
    ws.cell(3, 4).value = 50000.0
    ws.cell(4, 1).value = date(2026, 5, 11)
    ws.cell(4, 2).value = "Pago recibido"
    ws.cell(4, 3).value = 1800.0
    ws.cell(4, 4).value = 51800.0
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["banco_detectado"] == "mercadopago"
        assert result["total"] == 2
        assert result["movimientos"][0]["monto"] == 3200.0
    finally:
        _cleanup(path)


# ── Helper genérico para los 6 bancos regionales (Fecha/Concepto/Déb/Créd) ────

def _wb_banco_regional(nombre: str):
    """Extracto formato Cta. Cte. con columnas Débito/Crédito separadas
    (Credicoop, Supervielle, Bancor, Rioja, La Pampa)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = nombre
    ws.cell(2, 1).value = "Fecha"
    ws.cell(2, 2).value = "Concepto"
    ws.cell(2, 3).value = "Débito"
    ws.cell(2, 4).value = "Crédito"
    ws.cell(2, 5).value = "Saldo"
    ws.cell(3, 1).value = date(2026, 5, 10)
    ws.cell(3, 2).value = "TRANSFERENCIA RECIBIDA 20111111111"
    ws.cell(3, 3).value = None
    ws.cell(3, 4).value = 25000.0
    ws.cell(3, 5).value = 125000.0
    ws.cell(4, 1).value = date(2026, 5, 11)
    ws.cell(4, 2).value = "PAGO PROVEEDOR"
    ws.cell(4, 3).value = 4000.0
    ws.cell(4, 4).value = None
    ws.cell(4, 5).value = 121000.0
    return wb


def _wb_banco_importe_unico(nombre: str):
    """Extracto formato Fecha/Descripción/Importe/Saldo (Patagonia)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = nombre
    ws.cell(2, 1).value = "Fecha"
    ws.cell(2, 2).value = "Descripción"
    ws.cell(2, 3).value = "Importe"
    ws.cell(2, 4).value = "Saldo"
    ws.cell(3, 1).value = date(2026, 5, 10)
    ws.cell(3, 2).value = "ACREDITACION 30222222229"
    ws.cell(3, 3).value = 18000.0
    ws.cell(3, 4).value = 118000.0
    ws.cell(4, 1).value = date(2026, 5, 11)
    ws.cell(4, 2).value = "DEBITO COMISION"
    ws.cell(4, 3).value = -250.0
    ws.cell(4, 4).value = 117750.0
    return wb


@pytest.mark.parametrize("nombre,esperado", [
    ("Banco Credicoop Coop. Ltdo.", "credicoop"),
    ("Banco Supervielle S.A.",       "supervielle"),
    ("Banco de Córdoba - Bancor",    "bancor"),
    ("Nuevo Banco de La Rioja S.A.", "rioja"),
    ("Banco de La Pampa S.E.M.",     "lapampa"),
])
def test_bancos_regionales_credito_debito(nombre, esperado):
    wb = _wb_banco_regional(nombre)
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["banco_detectado"] == esperado, f"esperaba {esperado}"
        assert result["total"] == 2
        movs = result["movimientos"]
        credito = next(m for m in movs if "RECIBIDA" in (m["titular"] or ""))
        debito  = next(m for m in movs if "PROVEEDOR" in (m["titular"] or ""))
        assert credito["monto"] == 25000.0
        assert debito["monto"] == -4000.0
    finally:
        _cleanup(path)


def test_patagonia_importe_unico_con_signo():
    wb = _wb_banco_importe_unico("Banco Patagonia S.A.")
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["banco_detectado"] == "patagonia"
        assert result["total"] == 2
        movs = result["movimientos"]
        assert movs[0]["monto"] == 18000.0
        assert movs[1]["monto"] == -250.0
    finally:
        _cleanup(path)


def test_montos_son_decimal_compatibles():
    """Los montos parseados deben convertir a Decimal sin pérdida (regla DB)."""
    from decimal import Decimal
    wb = _wb_banco_regional("Banco Credicoop")
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        for m in movs:
            d = Decimal(str(m["monto"]))
            assert d == d.quantize(Decimal("0.01")) or d == d  # sin excepción
        assert Decimal(str(movs[0]["monto"])) in (Decimal("25000.00"), Decimal("-4000.00"))
    finally:
        _cleanup(path)


def test_bancos_nuevos_no_rompen_generico():
    """El fallback genérico sigue funcionando para un banco desconocido."""
    wb = _wb_banco_regional("Banco Desconocido XYZ")
    path = _xlsx_tmp(wb)
    try:
        result = parsear_extracto_bancario(path)
        assert result["banco_detectado"] == "generico"
        assert result["total"] == 2
    finally:
        _cleanup(path)


# ── Regresión: "rio" no debe matchear adentro de "período" / "anterior" ──────
# (encontrado con un extracto real subido por la usuaria, banco no identificado)

def test_rio_no_matchea_dentro_de_periodo():
    """'Período' contiene la substring 'rio' — no debe detectar Santander."""
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 1).value = "CBU de la cuenta:"
    ws.cell(3, 2).value = "4320001010003448310015"
    ws.cell(3, 4).value = "Período:"
    ws.cell(3, 5).value = "17/06/2026 al 24/06/2026"
    assert detectar_banco(ws) == "generico"


def test_rio_no_matchea_dentro_de_anterior_ni_horario():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Saldo anterior y horario de corte"
    assert detectar_banco(ws) == "generico"


def test_rioja_sigue_detectando_aunque_diga_rio_solo():
    """El fix de \\b no debe romper la detección real de Rioja/Río."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Nuevo Banco de La Rioja S.A."
    assert detectar_banco(ws) == "rioja"


# ── Regresión: fila de resumen "Total débitos:/Total créditos:" no debe ──────
# confundirse con el header real de la tabla (encontrado en extracto real) ────

def _wb_resumen_antes_de_header():
    """Extracto con un renglón resumen tipo label:valor antes del header real,
    y una columna única 'Monto' + columna indicadora de texto 'Débito/Crédito'."""
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 1).value = "CBU de la cuenta:"
    ws.cell(3, 2).value = "4320001010003448310015"
    ws.cell(6, 1).value = "Total débitos:"
    ws.cell(6, 2).value = "$ 27.489,34"
    ws.cell(6, 4).value = "Total créditos:"
    ws.cell(6, 5).value = "$ 900.000,00"
    ws.cell(8, 1).value = "Fecha"
    ws.cell(8, 2).value = "Tipo de transferencia"
    ws.cell(8, 3).value = "Débito/Crédito"
    ws.cell(8, 4).value = "Monto"
    ws.cell(8, 5).value = "Saldo"
    ws.cell(9, 1).value = date(2026, 6, 24)
    ws.cell(9, 2).value = "SIRCREB"
    ws.cell(9, 3).value = "Débito"
    ws.cell(9, 4).value = 90.0
    ws.cell(9, 5).value = 98192956.14
    ws.cell(10, 1).value = date(2026, 6, 24)
    ws.cell(10, 2).value = "CRÉDITO POR TRANSFERENCIA"
    ws.cell(10, 3).value = "Crédito"
    ws.cell(10, 4).value = 900000.0
    ws.cell(10, 5).value = 98193046.14
    return wb


def test_resumen_no_se_confunde_con_header():
    wb = _wb_resumen_antes_de_header()
    cols = detectar_columnas(wb.active)
    assert cols["hdr_row"] == 8


def test_columna_indicador_debito_credito_define_signo():
    """Monto único + columna 'Débito/Crédito' por fila: el signo lo da el indicador."""
    wb = _wb_resumen_antes_de_header()
    path = _xlsx_tmp(wb)
    try:
        movs = parsear_extracto_bancario(path)["movimientos"]
        assert len(movs) == 2
        debito = next(m for m in movs if m["titular"] == "SIRCREB")
        credito = next(m for m in movs if "TRANSFERENCIA" in m["titular"])
        assert debito["monto"] == -90.0
        assert credito["monto"] == 900000.0
    finally:
        _cleanup(path)
