"""
Tests del export de planilla conciliada (Excel + PDF).

Cubre:
- Excel: abre con openpyxl, tiene fila TOTAL con la suma correcta,
  y fila de total declarado + diferencia cuando la planilla tiene
  `total_declarado`.
- PDF: GET /planillas/{id}/export-pdf devuelve 200, content-type
  application/pdf y bytes que empiezan con %PDF.
- Multi-tenant: planilla de otra organización → 404 en ambos endpoints.
"""
import io
import pytest
from datetime import date, datetime
from decimal import Decimal

from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.database import Base, get_db
from app.services.auth import get_password_hash, create_access_token
from app.services.excel_export import export_planilla_conciliada
from app.models.organizacion import Organizacion
from app.models.user import User
from app.models.cliente import Cliente
from app.models.extracto import ExtractoBancario, MovimientoBanco
from app.models.planilla import Planilla, PlanillaRow


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    org1 = Organizacion(id=1, nombre="OrgExportPlanilla")
    org2 = Organizacion(id=2, nombre="OrgExportPlanillaOtra")
    session.add_all([org1, org2])
    session.commit()

    admin = User(
        email="admin@exportplan.test", full_name="Admin Export",
        hashed_password=get_password_hash("pw123"),
        organizacion_id=1, is_active=True, role="admin", is_superadmin=False,
    )
    session.add(admin)
    session.commit()

    yield session
    session.close()


@pytest.fixture
def client(db):
    def _override_db():
        yield db

    app.dependency_overrides[get_db] = _override_db
    c = TestClient(app, raise_server_exceptions=True)
    yield c
    app.dependency_overrides.pop(get_db, None)


def _token(db, email):
    user = db.query(User).filter(User.email == email).first()
    return create_access_token({"sub": user.email, "user_id": user.id, "role": user.role})


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _seed_planilla(db, org_id=1, total_declarado=None, cliente_nombre="ClienteExport"):
    cliente = Cliente(nombre=cliente_nombre, organizacion_id=org_id)
    db.add(cliente)
    db.flush()

    extracto = ExtractoBancario(
        nombre_archivo="extracto_export_test.xlsx",
        organizacion_id=org_id, creado_por=1,
        fecha_creacion=datetime.utcnow(),
    )
    db.add(extracto)
    db.flush()

    mov = MovimientoBanco(
        extracto_id=extracto.id, orden=1, fecha=date(2026, 6, 1), mes="junio",
        titular="CLIENTE EXPORT SA", monto=Decimal("5000.00"), saldo=Decimal("5000.00"),
        source="extracto", organizacion_id=org_id,
    )
    db.add(mov)
    db.flush()

    planilla = Planilla(
        cliente_id=cliente.id, extracto_id=extracto.id, usuario_id=1,
        nombre_archivo="planilla_export_test.xlsx", organizacion_id=org_id,
        fecha_carga=datetime.utcnow(), total_declarado=total_declarado,
    )
    db.add(planilla)
    db.flush()

    row_ok = PlanillaRow(
        planilla_id=planilla.id, monto=Decimal("5000.00"), cuit="20123456789",
        titular="Cliente Export SA", status="ok",
        orden_movimiento_acreditado=mov.id, organizacion_id=org_id,
    )
    row_no_esta = PlanillaRow(
        planilla_id=planilla.id, monto=Decimal("3000.00"), cuit=None,
        titular="Otro", status="no está", organizacion_id=org_id,
    )
    db.add_all([row_ok, row_no_esta])
    db.commit()
    db.refresh(planilla)
    return planilla


# ── Excel: contenido y fila de totales ────────────────────────────────────────

class TestExcelPlanillaConciliada:

    def test_excel_abre_con_openpyxl_y_total_correcto(self):
        """La suma de la fila TOTAL debe ser la suma de los montos de las filas."""
        import openpyxl

        planilla_data = {
            "cliente_nombre": "Cliente Test",
            "nombre_archivo": "planilla.xlsx",
            "rows": [
                {"monto": Decimal("5000.00"), "cuit": "20123456789", "titular": "Juan Perez",
                 "status": "ok", "orden_movimiento_acreditado": 1, "mov_titular": "Juan Perez",
                 "mov_fecha": date(2026, 6, 1), "mov_fecha_acred": date(2026, 6, 2)},
                {"monto": Decimal("3000.00"), "cuit": None, "titular": "Otro",
                 "status": "no está", "orden_movimiento_acreditado": None, "mov_titular": None,
                 "mov_fecha": None, "mov_fecha_acred": None},
            ],
        }
        xlsx_bytes = export_planilla_conciliada(planilla_data, [])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Planilla cliente"]

        # Fila 6 y 7 = datos (2 filas), fila 8 = TOTAL
        assert ws.cell(row=8, column=1).value == "TOTAL"
        assert ws.cell(row=8, column=2).value == Decimal("8000.00")

    def test_excel_con_total_declarado_agrega_filas_de_cuadre(self):
        """Con total_declarado, se agregan filas de 'Total declarado' y 'Diferencia'."""
        import openpyxl

        planilla_data = {
            "cliente_nombre": "Cliente Test",
            "nombre_archivo": "planilla.xlsx",
            "rows": [
                {"monto": Decimal("5000.00"), "cuit": "20123456789", "titular": "Juan Perez",
                 "status": "ok", "orden_movimiento_acreditado": 1, "mov_titular": "Juan Perez",
                 "mov_fecha": date(2026, 6, 1), "mov_fecha_acred": date(2026, 6, 2)},
            ],
            "total_declarado": Decimal("5500.00"),
        }
        xlsx_bytes = export_planilla_conciliada(planilla_data, [])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Planilla cliente"]

        # Fila 6 = dato, fila 7 = TOTAL, fila 8 = Total declarado, fila 9 = Diferencia
        assert ws.cell(row=7, column=1).value == "TOTAL"
        assert ws.cell(row=7, column=2).value == Decimal("5000.00")
        assert ws.cell(row=8, column=1).value == "Total declarado por el cliente"
        assert ws.cell(row=8, column=2).value == Decimal("5500.00")
        assert ws.cell(row=9, column=1).value == "Diferencia"
        assert ws.cell(row=9, column=2).value == Decimal("500.00")

    def test_excel_sin_total_declarado_no_agrega_filas_de_cuadre(self):
        """Sin total_declarado, no hay filas extra después del TOTAL."""
        import openpyxl

        planilla_data = {
            "cliente_nombre": "Cliente Test",
            "nombre_archivo": "planilla.xlsx",
            "rows": [
                {"monto": Decimal("1000.00"), "cuit": None, "titular": "X",
                 "status": "ok", "orden_movimiento_acreditado": None, "mov_titular": None,
                 "mov_fecha": None, "mov_fecha_acred": None},
            ],
        }
        xlsx_bytes = export_planilla_conciliada(planilla_data, [])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Planilla cliente"]

        assert ws.cell(row=7, column=1).value == "TOTAL"
        assert ws.cell(row=8, column=1).value is None


# ── Endpoint Excel: multi-tenant ──────────────────────────────────────────────

class TestDownloadEndpointMultiTenant:

    def test_download_planilla_otra_org_404(self, client, db):
        planilla_otra_org = _seed_planilla(db, org_id=2)
        token = _token(db, "admin@exportplan.test")
        r = client.get(f"/planillas/{planilla_otra_org.id}/download", headers=_auth(token))
        assert r.status_code == 404

    def test_download_planilla_propia_org_200(self, client, db):
        planilla = _seed_planilla(db, org_id=1)
        token = _token(db, "admin@exportplan.test")
        r = client.get(f"/planillas/{planilla.id}/download", headers=_auth(token))
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]


# ── Endpoint PDF ───────────────────────────────────────────────────────────────

class TestExportPdfEndpoint:

    def test_export_pdf_devuelve_200_y_pdf_valido(self, client, db):
        planilla = _seed_planilla(db, org_id=1, total_declarado=Decimal("8500.00"))
        token = _token(db, "admin@exportplan.test")
        r = client.get(f"/planillas/{planilla.id}/export-pdf", headers=_auth(token))
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    def test_export_pdf_sin_auth_403(self, client):
        r = client.get("/planillas/1/export-pdf")
        assert r.status_code in (401, 403)

    def test_export_pdf_planilla_inexistente_404(self, client, db):
        token = _token(db, "admin@exportplan.test")
        r = client.get("/planillas/99999/export-pdf", headers=_auth(token))
        assert r.status_code == 404

    def test_export_pdf_planilla_otra_org_404(self, client, db):
        """Aislamiento multi-tenant: planilla de otra org → 404, no 200 con datos ajenos."""
        planilla_otra_org = _seed_planilla(db, org_id=2)
        token = _token(db, "admin@exportplan.test")
        r = client.get(f"/planillas/{planilla_otra_org.id}/export-pdf", headers=_auth(token))
        assert r.status_code == 404
