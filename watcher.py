import os, re, time, shutil, datetime, tempfile, json
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ─── configuración desde config.json ─────────────────────────────────────────

_HERE = Path(__file__).parent
with open(_HERE / "config.json", encoding="utf-8") as _f:
    _CFG = json.load(_f)

INBOX         = Path(_CFG["inbox_path"])
EXTRACTO      = Path(_CFG["extracto_path"])
CLIENTES_BASE = Path(_CFG["clientes_path"])
CARPETA_MES   = _CFG.get("carpeta_mes", "")
PROCESADOS    = INBOX / "procesados"

def _resolver_fecha(v):
    if v == "hoy":  return datetime.date.today()
    if v == "ayer": return datetime.date.today() - datetime.timedelta(days=1)
    return datetime.date.fromisoformat(v)

FECHA_ACRED = _resolver_fecha(_CFG.get("fecha_acred", "hoy"))

CLIENTS = ['green','tucu','david','smt','gwinn','innova','camparo','alojando','pinares','paraguay']

DATE_FMT   = 'dd/mm/yyyy'
CURR_FMT   = r'"$"\ #,##0.00;"$"\ \-#,##0.00'
THIN       = Side(style='thin')
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL   = PatternFill('solid', fgColor='002060')
HDR_FONT   = Font(color='FFFFFF', bold=True)
PALETA     = ['FFF2CC','D9EAD3','CFE2F3','FCE5CD','EAD1DC','D9D2E9']

_planilla_counter = defaultdict(int)


# ─── helpers ──────────────────────────────────────────────────────────────────

def hoy():
    return FECHA_ACRED

def norm_fecha(v):
    if hasattr(v, 'date'): return v.date()
    return v

def es_libre(cli):
    if cli is None: return True
    if isinstance(cli, str) and cli.strip().lower() in ('no identificado', ''): return True
    return False

def parse_importe(v):
    """Acepta int, float, y strings con $, espacios, nbsp, formato ES o EN."""
    if isinstance(v, (int, float)): return round(float(v), 2)
    if isinstance(v, str):
        s = v.strip().replace('$','').replace('\xa0','').replace(' ','')
        if not s: return None
        if ',' in s and '.' in s:
            if s.rfind(',') > s.rfind('.'): s = s.replace('.','').replace(',','.')
            else: s = s.replace(',','')
        elif ',' in s: s = s.replace(',','.')
        try: return round(float(s), 2)
        except ValueError: pass
    return None

def norm_cuit(v):
    """Normaliza CUIT/CUIL a string de solo dígitos."""
    if v is None: return ''
    return re.sub(r'\D', '', str(v))

def extraer_cuit_titular(titular):
    """Extrae el CUIT del campo titular del extracto (busca 10-11 dígitos consecutivos)."""
    if not titular: return ''
    nums = re.findall(r'\d{10,11}', str(titular))
    return nums[0] if nums else ''

def detectar_header(ws):
    """
    Busca 'monto' o 'importe' en filas 1-5.
    Verifica que la col (o col+1) tenga datos numéricos reales.
    Retorna (header_row, imp_col).
    """
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(r, c).value or '').lower().strip()
            if 'monto' in h or 'importe' in h:
                data_start = r + 1
                for test_col in [c, c + 1]:
                    hits = sum(
                        1 for dr in range(data_start, min(data_start+6, ws.max_row+1))
                        if isinstance(ws.cell(dr, test_col).value, (int, float))
                        and 0 < ws.cell(dr, test_col).value < 5_000_000
                    )
                    if hits >= 1: return r, test_col
    return 2, 6

def detectar_cuit_col(ws, hdr_row):
    """Detecta columna de CUIT buscando 'cuit' en el header."""
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hdr_row, c).value or '').lower().strip()
        if 'cuit' in h: return c
    return None

def detectar_titular_col(ws, hdr_row):
    """Detecta columna de titular/nombre buscando 'titular' o 'nombre' en el header."""
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hdr_row, c).value or '').lower().strip()
        if 'titular' in h or 'nombre' in h: return c
    return None


# ─── extracto ─────────────────────────────────────────────────────────────────

def cargar_extracto():
    """
    Carga con data_only=True para obtener valores evaluados (no fórmulas).
    Solo requiere importe válido en col 6.
    Col 2 puede ser fórmula o entero: NO se filtra por tipo.
    """
    wb = openpyxl.load_workbook(EXTRACTO, data_only=True)
    ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        imp = parse_importe(ws.cell(r, 6).value)
        if imp is None: continue
        saldo = ws.cell(r, 7).value
        rows.append({
            "row":        r,
            "orden":      ws.cell(r, 2).value,   # número real gracias a data_only
            "importe":    imp,
            "titular":    str(ws.cell(r, 5).value or ""),
            "cliente":    ws.cell(r, 8).value,
            "fecha_acred": norm_fecha(ws.cell(r, 9).value),
            "fecha":      ws.cell(r, 3).value,
            "mes":        ws.cell(r, 4).value,
            "saldo":      float(saldo) if isinstance(saldo, (int, float)) else None,
            "cuit_ex":    extraer_cuit_titular(ws.cell(r, 5).value),
        })
    return wb, ws, rows

def guardar_extracto(wb_ex):
    """Guarda via tempfile para evitar PermissionError de Windows."""
    tmp = Path(tempfile.gettempdir()) / "extracto_watcher_tmp.xlsx"
    wb_ex.save(str(tmp))
    with open(str(tmp), 'rb') as f: data = f.read()
    tmp.unlink(missing_ok=True)
    try:
        with open(str(EXTRACTO), 'wb') as f: f.write(data)
    except PermissionError:
        alt = EXTRACTO.parent / (EXTRACTO.stem + "_NUEVO.xlsx")
        with open(str(alt), 'wb') as f: f.write(data)
        print(f"  ⚠ Extracto abierto → guardado como {alt.name}")


# ─── matching ─────────────────────────────────────────────────────────────────

UMBRAL_COMUN = 3  # si hay ≥ N filas con mismo monto, exigir match de CUIT

def buscar_match(extracto, imp, cuit_cli, titular_cli, usados):
    """
    Busca la mejor fila libre para acreditar:
    - Monto poco frecuente (< UMBRAL_COMUN): match por monto solo.
    - Monto común (>= UMBRAL_COMUN): requiere match de CUIT o titular parcial.
    Retorna (match, razon) o (None, razon).
    """
    candidatos  = [e for e in extracto if e["importe"] == imp]
    no_usados   = [e for e in candidatos if e["row"] not in usados]
    libres      = [e for e in no_usados if es_libre(e["cliente"])]

    if not candidatos:
        return None, "no está"
    if not libres:
        if not no_usados:
            return None, "duplicado"
        match = no_usados[0]
        fa    = match["fecha_acred"]
        return match, f"acreditado {fa.strftime('%d/%m') if fa else ''}"

    # Hay filas libres — decidir si necesitamos validar CUIT
    if len(candidatos) < UMBRAL_COMUN:
        return libres[0], "ok"

    # Monto frecuente: preferir match de CUIT
    cuit_cli_norm = norm_cuit(cuit_cli)
    if cuit_cli_norm:
        por_cuit = [e for e in libres if e["cuit_ex"] == cuit_cli_norm]
        if por_cuit:
            return por_cuit[0], "ok"

        # Intentar con titular parcial (primeras 2 palabras)
        if titular_cli:
            partes = str(titular_cli).upper().split()[:2]
            por_tit = [e for e in libres
                       if all(p in e["titular"].upper() for p in partes)]
            if por_tit:
                return por_tit[0], "ok"

        return None, "faltan datos"

    # Sin CUIT en planilla y monto común → match directo al primero libre
    return libres[0], "ok"


# ─── conciliar ────────────────────────────────────────────────────────────────

def conciliar_hoja(ws, cliente, fecha, wb_ex, ws_ex, extracto,
                   idx_color, row_fill, usados, esta_planilla):
    """Procesa UNA hoja de la planilla de cliente."""
    hdr_row, imp_col = detectar_header(ws)
    status_col  = imp_col + 1
    data_start  = hdr_row + 1
    cuit_col    = detectar_cuit_col(ws, hdr_row)
    titular_col = detectar_titular_col(ws, hdr_row)
    print(f"    header_row={hdr_row}, imp_col={imp_col}, cuit_col={cuit_col}, data desde r{data_start}")

    RED  = Font(color='FF0000', bold=True)
    GRN  = Font(color='00B050', bold=True)
    DGRN = Font(color='006400', bold=True)
    stats = {"ok":0, "no_esta":0, "duplicado":0, "faltan_datos":0, "acreditado":0}

    for r in range(data_start, ws.max_row + 1):
        imp = parse_importe(ws.cell(r, imp_col).value)
        if imp is None: continue
        if imp > 5_000_000: continue    # totales / CUIT

        cuit_cli    = ws.cell(r, cuit_col).value    if cuit_col    else None
        titular_cli = ws.cell(r, titular_col).value if titular_col else None

        match, estado = buscar_match(extracto, imp, cuit_cli, titular_cli, usados)

        if estado == "ok" and match:
            usados.add(match["row"])
            ws_ex.cell(match["row"], 8).value = cliente
            fa_dt = datetime.datetime.combine(fecha, datetime.time())
            ws_ex.cell(match["row"], 9).value = fa_dt
            ws_ex.cell(match["row"], 9).number_format = DATE_FMT
            ws_ex.cell(match["row"], 8).fill = row_fill
            ws_ex.cell(match["row"], 9).fill = row_fill
            match["cliente"]     = cliente
            match["fecha_acred"] = fecha
            esta_planilla.append(match)
            stats["ok"] += 1
        elif estado == "no está":        stats["no_esta"] += 1
        elif estado == "duplicado":      stats["duplicado"] += 1
        elif estado == "faltan datos":   stats["faltan_datos"] += 1
        elif estado.startswith("acreditado"):
            stats["acreditado"] += 1
            if match: usados.add(match["row"])

        c = ws.cell(r, status_col)
        c.value = estado
        if estado == "ok":                                      c.font = GRN
        elif estado in ("no está","duplicado","faltan datos"):  c.font = RED
        elif estado.startswith("acreditado"):                   c.font = DGRN

    print(f"    {stats}")
    return stats


def conciliar_cliente(path, cliente):
    if not path.exists():
        print(f"  no encontrado: {path}"); return False

    fecha = hoy()
    print(f"\n>>> {cliente} | {path.name} ({fecha})")

    clave = (cliente, str(fecha))
    idx_color = _planilla_counter[clave]
    _planilla_counter[clave] += 1
    color_hex = PALETA[idx_color % len(PALETA)]
    row_fill  = PatternFill('solid', fgColor=color_hex)

    wb_ex, ws_ex, extracto = cargar_extracto()
    print(f"  Extracto: {len(extracto)} filas")
    usados = set()
    esta_planilla = []

    wb = openpyxl.load_workbook(path)

    # Siempre se procesa la primera solapa (índice 0) — es la más reciente
    # Aplica tanto a SMT (60+ solapas, la primera es el día actual) como al resto
    ws_cliente = wb.worksheets[0]
    conciliar_hoja(ws_cliente, cliente, fecha, wb_ex, ws_ex, extracto,
                   idx_color, row_fill, usados, esta_planilla)

    guardar_extracto(wb_ex)

    # ── Hoja2 ──────────────────────────────────────────────────────────────────
    cli_cap = cliente.capitalize()
    ws2 = wb.create_sheet("Hoja2")
    headers = ["Orden","Fecha","Mes","titular","Importe Pesos","Saldo","cliente","fecha acred"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(1, col, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center'); c.border = ALL_BORDER

    dr = 2
    for e in esta_planilla:
        fa = e["fecha_acred"]
        fa_dt = datetime.datetime.combine(fa, datetime.time()) if isinstance(fa, datetime.date) else fa
        ws2.cell(dr, 1, e.get("orden"))
        ws2.cell(dr, 2, e["fecha"]).number_format = DATE_FMT
        ws2.cell(dr, 3, e["mes"])
        ws2.cell(dr, 4, e["titular"])
        ws2.cell(dr, 5, e["importe"]).number_format = CURR_FMT
        if e["saldo"] is not None:
            ws2.cell(dr, 6, e["saldo"]).number_format = CURR_FMT
        ws2.cell(dr, 7, e["cliente"])
        ws2.cell(dr, 8, fa_dt).number_format = DATE_FMT
        for col in range(1, 9): ws2.cell(dr, col).border = ALL_BORDER
        dr += 1

    ic = get_column_letter(5)
    c = ws2.cell(dr, 5)
    c.value = f'=SUM({ic}2:{ic}{dr-1})'
    c.number_format = CURR_FMT; c.font = Font(bold=True); c.border = ALL_BORDER

    for ltr, w in zip('ABCDEFGH', [10,12,6,45,16,16,12,14]):
        ws2.column_dimensions[ltr].width = w

    # ── Guardar archivo cliente ─────────────────────────────────────────────────
    dest_dir = CLIENTES_BASE / cli_cap
    month_dir = None
    if dest_dir.is_dir():
        for d in sorted(dest_dir.iterdir()):
            if d.is_dir() and ('abr' in d.name.lower() or '-4' in d.name or '4 ' in d.name):
                month_dir = d; break
    if not month_dir:
        month_dir = dest_dir / CARPETA_MES
        month_dir.mkdir(parents=True, exist_ok=True)

    fecha_d  = fecha.strftime('%d.%m')
    suffix   = f" ({idx_color + 1})" if idx_color > 0 else ""
    out_name = f"{cli_cap} acreditado {fecha_d}{suffix}.xlsx"
    out_path = month_dir / out_name

    tmp_out = Path(tempfile.gettempdir()) / out_name
    wb.save(str(tmp_out))
    with open(str(tmp_out), 'rb') as f: d2 = f.read()
    tmp_out.unlink(missing_ok=True)
    with open(str(out_path), 'wb') as f: f.write(d2)
    print(f"  ✓ {len(esta_planilla)} acreditadas → {out_path}")
    return True


# ─── scan / main ──────────────────────────────────────────────────────────────

def scan():
    for cliente in CLIENTS:
        carpeta = INBOX / cliente
        if not carpeta.exists(): continue
        for f in sorted(carpeta.glob("*.xlsx")):
            try:
                ok = conciliar_cliente(f, cliente)
                if ok:
                    dest = PROCESADOS / f.name
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(f), str(dest))
            except Exception as exc:
                import traceback
                print(f"  ERROR {f.name}: {exc}")
                traceback.print_exc()

def main():
    if not EXTRACTO.exists():
        print(f"ERROR: extracto no encontrado:\n  {EXTRACTO}"); return
    if not INBOX.exists():
        print(f"ERROR: INBOX no encontrado:\n  {INBOX}"); return
    PROCESADOS.mkdir(parents=True, exist_ok=True)
    print(f"Extracto:  {EXTRACTO}")
    print(f"INBOX:     {INBOX}")
    print(f"Fecha acred: {FECHA_ACRED}